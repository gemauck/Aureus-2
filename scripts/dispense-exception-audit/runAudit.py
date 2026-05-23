#!/usr/bin/env python3
# CLI: audit dispense exception workbook and write Excel report with findings.
"""Run dispense exception audit and produce output workbook."""

from __future__ import annotations

import argparse
import json
import shutil
import sys
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook as load_openpyxl_workbook

from auditDecisions import run_all_audits
from parseWorkbook import detect_workbook, load_workbook


def write_audit_excel(input_path: str, output_path: str, audit_result: dict) -> None:
    shutil.copy2(input_path, output_path)
    wb = load_openpyxl_workbook(output_path)
    for sheet_name in ('Audit Findings', 'Audit Summary'):
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]
    wb.save(output_path)

    findings = audit_result['findings']
    summary = audit_result['summary']

    with pd.ExcelWriter(output_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        findings_df = pd.DataFrame(findings, columns=[
            'transaction_id', 'asset_number', 'check', 'severity',
            'manual_value', 'expected_value', 'evidence',
        ])
        findings_df.to_excel(writer, sheet_name='Audit Findings', index=False)

        summary_rows = [
            {'metric': 'transaction_count', 'value': summary['transaction_count']},
            {'metric': 'exception_flagged_count', 'value': summary['exception_flagged_count']},
            {'metric': 'exception_match_pct', 'value': summary['exception_match_pct']},
            {'metric': 'review_queue_count', 'value': summary['review_queue_count']},
            {'metric': 'finding_count', 'value': summary['finding_count']},
            {'metric': 'error_count', 'value': summary['error_count']},
            {'metric': 'warning_count', 'value': summary['warning_count']},
            {'metric': 'decision_pass_rate_pct', 'value': summary['decision_pass_rate_pct']},
        ]
        for check, count in summary.get('findings_by_check', {}).items():
            summary_rows.append({'metric': f'findings_{check}', 'value': count})
        pd.DataFrame(summary_rows).to_excel(writer, sheet_name='Audit Summary', index=False)


def main() -> int:
    parser = argparse.ArgumentParser(description='Audit dispense exception workbook manual decisions')
    parser.add_argument('input', help='Input .xlsx workbook path')
    parser.add_argument('-o', '--output', help='Output .xlsx path (default: input with -audit suffix)')
    parser.add_argument('--json', help='Also write JSON summary to this path')
    parser.add_argument('--economy-threshold', type=float, default=5.0, help='Abs variance threshold (5.0 = 500%%)')
    args = parser.parse_args()

    input_path = Path(args.input)
    if not input_path.exists():
        print(f'Input not found: {input_path}', file=sys.stderr)
        return 1

    detection = detect_workbook(str(input_path))
    if not detection['valid']:
        print(f"Invalid workbook: missing {detection['missing_sheets']}", file=sys.stderr)
        return 1

    data = load_workbook(str(input_path))
    audit_result = run_all_audits(data, economy_threshold=args.economy_threshold)

    output_path = args.output or str(input_path.with_name(input_path.stem + '-audit.xlsx'))
    write_audit_excel(str(input_path), output_path, audit_result)

    payload = {
        'summary': audit_result['summary'],
        'findings': audit_result['findings'],
        'output_path': output_path,
    }
    print(json.dumps(payload['summary']))

    if args.json:
        with open(args.json, 'w', encoding='utf-8') as f:
            json.dump(payload, f, indent=2, default=str)

    print(f'Audit report written to: {output_path}')
    return 0


if __name__ == '__main__':
    raise SystemExit(main())
