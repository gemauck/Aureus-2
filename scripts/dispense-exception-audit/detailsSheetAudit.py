# Write audit review columns onto the Details as Assets sheet in audit workbooks.
"""Inline audit columns for grouped Details as Assets transaction rows."""

from __future__ import annotations

from typing import Any

from openpyxl import load_workbook as load_openpyxl_workbook

from parseWorkbook import (
    is_column_header_row,
    is_transaction_row,
    map_header_row,
)

DETAILS_SHEET = 'Details as Assets'
AUDIT_HEADERS = ('Audit Severity', 'Audit Findings', 'Auditor Comment')


def format_findings_for_cell(findings: list[dict[str, Any]] | None) -> str:
    parts: list[str] = []
    for finding in findings or []:
        if finding.get('severity') == 'info':
            continue
        check = str(finding.get('check') or '').strip()
        detail = str(finding.get('expected_value') or finding.get('manual_value') or '').strip()
        if check and detail:
            parts.append(f'{check}: {detail}')
        elif check:
            parts.append(check)
        elif detail:
            parts.append(detail)
    return '; '.join(parts)


def build_txn_audit_lookup(
    review_transactions: list[dict[str, Any]] | None,
    comments: dict[str, str] | None = None,
) -> dict[str, dict[str, str]]:
    comment_map = {str(k): (v or '').strip() for k, v in (comments or {}).items()}
    lookup: dict[str, dict[str, str]] = {}
    for txn in review_transactions or []:
        tid = str(txn.get('transaction_id') or '').strip()
        if not tid:
            continue
        severity = str(txn.get('max_severity') or '').strip()
        if severity == 'none':
            severity = ''
        lookup[tid] = {
            'severity': severity,
            'findings': format_findings_for_cell(txn.get('findings')),
            'comment': comment_map.get(tid, ''),
        }
    for tid, text in comment_map.items():
        if tid not in lookup:
            lookup[tid] = {'severity': '', 'findings': '', 'comment': text}
        else:
            lookup[tid]['comment'] = text
    return lookup


def _row_values(ws, row_idx: int, width: int) -> list[Any]:
    return [ws.cell(row=row_idx, column=col).value for col in range(1, width + 1)]


def _last_filled_col(values: list[Any]) -> int:
    for idx in range(len(values), 0, -1):
        val = values[idx - 1]
        if val is None:
            continue
        if isinstance(val, str) and not val.strip():
            continue
        return idx
    return 0


def _find_details_data_width(ws) -> int:
    scan_width = max(ws.max_column, 30)
    max_width = 0
    for row_idx in range(1, ws.max_row + 1):
        values = _row_values(ws, row_idx, scan_width)
        if not is_column_header_row(values):
            continue
        width = _last_filled_col(values)
        while width > 0 and str(values[width - 1] or '').strip() in AUDIT_HEADERS:
            width -= 1
        max_width = max(max_width, width)
    return max_width or 24


def _find_existing_audit_cols(ws) -> tuple[int, int, int] | None:
    scan_width = max(ws.max_column, 30)
    for row_idx in range(1, min(ws.max_row, 200) + 1):
        values = _row_values(ws, row_idx, scan_width)
        if not is_column_header_row(values):
            continue
        cols: dict[str, int] = {}
        for col_idx, value in enumerate(values, start=1):
            label = str(value or '').strip()
            if label in AUDIT_HEADERS:
                cols[label] = col_idx
        if len(cols) == len(AUDIT_HEADERS):
            return tuple(cols[header] for header in AUDIT_HEADERS)
    return None


def _transaction_id_from_row(values: list[Any]) -> str | None:
    if not is_transaction_row(values):
        return None
    txn_id = values[1] if len(values) > 1 else None
    if txn_id is None:
        return None
    text = str(txn_id).strip()
    return text or None


def apply_audit_columns_to_details_sheet(
    output_path: str,
    review_transactions: list[dict[str, Any]] | None,
    comments: dict[str, str] | None = None,
    *,
    comments_only: bool = False,
) -> None:
    """Add or refresh audit columns on Details as Assets for each transaction row."""
    lookup = build_txn_audit_lookup(review_transactions, comments)
    wb = load_openpyxl_workbook(output_path)
    if DETAILS_SHEET not in wb.sheetnames:
        wb.save(output_path)
        return

    ws = wb[DETAILS_SHEET]
    audit_cols = _find_existing_audit_cols(ws)
    if not audit_cols:
        data_width = _find_details_data_width(ws)
        audit_cols = (data_width + 1, data_width + 2, data_width + 3)
    scan_width = max(audit_cols[2], ws.max_column, 30)

    for row_idx in range(1, ws.max_row + 1):
        values = _row_values(ws, row_idx, scan_width)
        if is_column_header_row(values):
            for col_idx, header in zip(audit_cols, AUDIT_HEADERS):
                ws.cell(row=row_idx, column=col_idx, value=header)
            continue

        tid = _transaction_id_from_row(values)
        if not tid:
            continue

        audit = lookup.get(tid, {})
        if comments_only:
            ws.cell(row=row_idx, column=audit_cols[2], value=audit.get('comment', ''))
            continue

        ws.cell(row=row_idx, column=audit_cols[0], value=audit.get('severity', ''))
        ws.cell(row=row_idx, column=audit_cols[1], value=audit.get('findings', ''))
        ws.cell(row=row_idx, column=audit_cols[2], value=audit.get('comment', ''))

    wb.save(output_path)
