"""Write analyst-ready dispense exception workbook."""
from __future__ import annotations

import re
import shutil
from copy import copy
from datetime import datetime
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from parse_workbook import HEADER_ALIASES, normalize_header

DETAILS_SHEET = "Details as Assets"

INELIGIBLE_HEADERS = [
    "Date & Time",
    "Transaction ID",
    "Asset Description",
    "Asset Number",
    "Asset Tag",
    "Asset Group",
    "Asset Tank Size (L)",
    "Asset Meter Type (Hr/Km)",
    "Storage Tank",
    "Fuel Pump",
    "Litres",
    "Opening Odo",
    "Closing Odo",
    "Total Usage Km/Hr",
    "Exception Reason",
    "Review Reason",
    "Abco Comment",
    "Operation Description / Comment",
    "Refund Eligibility",
    "Average Economy (180 Days)",
    "Economy",
    "% Variance",
    "Economy Type",
    "Department",
]

INELIGIBLE_FIELD_TO_HEADER = {
    "date_time": "Date & Time",
    "transaction_id": "Transaction ID",
    "asset_description": "Asset Description",
    "asset_number": "Asset Number",
    "asset_tag": "Asset Tag",
    "asset_group": "Asset Group",
    "tank_size_l": "Asset Tank Size (L)",
    "meter_type": "Asset Meter Type (Hr/Km)",
    "storage_tank": "Storage Tank",
    "fuel_pump": "Fuel Pump",
    "litres": "Litres",
    "opening_odo": "Opening Odo",
    "closing_odo": "Closing Odo",
    "total_usage": "Total Usage Km/Hr",
    "exception_reason": "Exception Reason",
    "review_reason": "Review Reason",
    "abco_comment": "Abco Comment",
    "operation_comment": "Operation Description / Comment",
    "refund_eligibility": "Refund Eligibility",
    "avg_economy_180d": "Average Economy (180 Days)",
    "economy": "Economy",
    "pct_variance": "% Variance",
    "economy_type": "Economy Type",
    "department": "Department",
}

EXCEPTION_GLOSSARY = [
    (
        "Fill outside of one hour from start",
        "Dispensing Point Error?",
        "Second or later dispense in a batch started more than 60 minutes after the first.",
        "Usually stays eligible — verify pump timing and batch start.",
    ),
    (
        "Odo difference <= 0",
        "AVR?",
        "Meter did not advance (or went backwards) between opening and closing readings.",
        "Often a split fill or AVR sync issue.",
    ),
    (
        "Odo difference <= 0, Consecutive dispenses within 60 minutes",
        "AVR?",
        "Zero odo movement on a consecutive dispense inside the 60-minute window.",
        "Review when part of a chain starting with odo <= 0.",
    ),
    (
        "Odo difference > 50 hrs",
        "AVR?",
        "Unusually large hour-meter jump for the period between dispenses.",
        "Check meter reset, AVR, or data entry. Use Abco Comment 'Just ok' to skip if valid.",
    ),
    (
        "Consecutive dispenses within 60 minutes",
        "AVR?",
        "Another dispense to the same asset within 60 minutes.",
        "Routine on split fills — only escalates when chained after odo <= 0.",
    ),
    (
        "AVR Sync",
        "AVR",
        "Transaction matched AVR Sync lookup upload.",
        "Confirm sync is legitimate before approving.",
    ),
]

FILL_GREEN = PatternFill("solid", fgColor="D9EAD3")
FILL_ORANGE = PatternFill("solid", fgColor="FF9900")
FILL_YELLOW = PatternFill("solid", fgColor="FFFF00")
HEADER_FILL = PatternFill("solid", fgColor="C9DAF8")
HEADER_FONT = Font(bold=True, color="000000")

SHEET_ORDER = [
    "Details as Assets",
    "Transactions deemed ineligible",
    "Possible Cause Summary",
    "Summary Per Asset",
    "Non-Mining Excluded",
    "Exception Reason Glossary",
    "Asset Info Lookup",
    "AVR Sync Lookup",
]


def _row_highlight(row: dict[str, Any]) -> PatternFill | None:
    comment = str(row.get("abco_comment") or "").lower()
    if row.get("_avr_sync") or comment == "avr sync":
        return FILL_YELLOW
    if row.get("_review"):
        return FILL_ORANGE
    if row.get("exception_60") or row.get("exception_120"):
        return FILL_GREEN
    return None


def _map_header_row(values: list[Any]) -> dict[str, int]:
    col_map: dict[str, int] = {}
    for idx, value in enumerate(values):
        norm = normalize_header(value)
        if not norm:
            continue
        for field, aliases in HEADER_ALIASES.items():
            if norm in aliases:
                col_map[field] = idx + 1
                break
    return col_map


def _is_column_header_row(values: list[Any]) -> bool:
    joined = " ".join(str(v or "").lower() for v in values[:6])
    return "date" in joined and "transaction" in joined


def _merge_col_maps(previous: dict[str, int], new_map: dict[str, int]) -> dict[str, int]:
    merged = dict(previous)
    merged.update(new_map)
    return merged


def _apply_light_touch_details(ws, txn_by_id: dict[str, dict[str, Any]]) -> None:
    """Preserve InsightWare layout; apply highlights and exception updates only."""
    ws.conditional_formatting._cf_rules.clear()
    max_col = ws.max_column
    col_map: dict[str, int] = {}

    for row_idx in range(1, ws.max_row + 1):
        values = [ws.cell(row_idx, col).value for col in range(1, max_col + 1)]

        if _is_column_header_row(values):
            new_map = _map_header_row(values)
            if new_map:
                col_map = _merge_col_maps(col_map, new_map)
            if row_idx == 1 and col_map.get("exception_60"):
                ws.cell(row_idx, col_map["exception_60"]).value = "Exception Reason"
            continue

        txn_col = col_map.get("transaction_id")
        if not txn_col:
            continue
        txn_id = str(ws.cell(row_idx, txn_col).value or "").strip()
        if not txn_id or txn_id.lower() == "transaction id":
            continue

        record = txn_by_id.get(txn_id)
        if not record:
            continue

        fill = _row_highlight(record)
        if fill:
            for col in range(1, max_col + 1):
                ws.cell(row_idx, col).fill = fill

        exc_col = col_map.get("exception_60")
        if exc_col and record.get("exception_60"):
            ws.cell(row_idx, exc_col).value = record["exception_60"]

        abco_col = col_map.get("abco_comment")
        if abco_col and not ws.cell(row_idx, abco_col).value:
            suggestion = record.get("suggested_abco_comment") or record.get("abco_comment")
            if suggestion:
                ws.cell(row_idx, abco_col).value = suggestion

    last_row = ws.max_row
    for col_idx in range(1, max_col + 1):
        header = normalize_header(ws.cell(1, col_idx).value)
        if header in {"% variance", "% varinace"}:
            var_col = get_column_letter(col_idx)
            ws.conditional_formatting.add(
                f"{var_col}2:{var_col}{last_row}",
                CellIsRule(operator="greaterThan", formula=['"60%"'], fill=PatternFill("solid", fgColor="B7E1CD")),
            )
            ws.conditional_formatting.add(
                f"{var_col}2:{var_col}{last_row}",
                CellIsRule(operator="lessThan", formula=['"-60%"'], fill=PatternFill("solid", fgColor="FFF2CC")),
            )
        if header in {
            "exception reason",
            "exception reason (60 min)",
            "exception reason (60 mins)",
        }:
            exc_col = get_column_letter(col_idx)
            ws.conditional_formatting.add(
                f"{exc_col}2:{exc_col}{last_row}",
                FormulaRule(formula=[f'LEN(TRIM({exc_col}2))>0'], fill=PatternFill("solid", fgColor="B7E1CD")),
            )


def _format_cell_value(field: str, value: Any) -> Any:
    if field == "date_time" and value is not None:
        return value
    if field == "pct_variance" and isinstance(value, (int, float)):
        return value
    if field in ("economy", "avg_economy_180d", "litres", "tank_size_l") and isinstance(
        value, (int, float)
    ):
        return value
    return value


def _write_row_from_record(
    ws,
    row_idx: int,
    record: dict[str, Any],
    headers: list[str],
    field_map: dict[str, str],
) -> None:
    header_to_field = {v: k for k, v in field_map.items()}
    fill = _row_highlight(record)
    for col, title in enumerate(headers, 1):
        field = header_to_field[title]
        if field == "exception_reason":
            value = _format_cell_value(field, record.get("exception_60"))
        else:
            value = _format_cell_value(field, record.get(field))
        if field == "pct_variance" and isinstance(value, float):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.number_format = "0.00%"
        else:
            cell = ws.cell(row=row_idx, column=col, value=value)
        if fill:
            cell.fill = fill


def _write_ineligible_sheet(ws, rows: list[dict[str, Any]]) -> None:
    ws.title = "Transactions deemed ineligible"
    ws.delete_rows(1, ws.max_row)
    ws.conditional_formatting._cf_rules.clear()

    for col, title in enumerate(INELIGIBLE_HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT

    for row_idx, record in enumerate(rows, 2):
        _write_row_from_record(ws, row_idx, record, INELIGIBLE_HEADERS, INELIGIBLE_FIELD_TO_HEADER)

    total_litres = sum(float(row.get("litres") or 0) for row in rows)
    footer_row = len(rows) + 2
    ws.cell(row=footer_row, column=1, value=f"Uploaded {datetime.now():%d %B %Y}")
    litres_col = INELIGIBLE_HEADERS.index("Litres") + 1
    ws.cell(row=footer_row, column=litres_col, value=round(total_litres, 2))

    last_row = max(2, len(rows) + 1)
    var_col = get_column_letter(INELIGIBLE_HEADERS.index("% Variance") + 1)
    exc_col = get_column_letter(INELIGIBLE_HEADERS.index("Exception Reason") + 1)
    ws.conditional_formatting.add(
        f"{var_col}2:{var_col}{last_row}",
        CellIsRule(operator="greaterThan", formula=['"60%"'], fill=PatternFill("solid", fgColor="B7E1CD")),
    )
    ws.conditional_formatting.add(
        f"{var_col}2:{var_col}{last_row}",
        CellIsRule(operator="lessThan", formula=['"-60%"'], fill=PatternFill("solid", fgColor="FFF2CC")),
    )
    ws.conditional_formatting.add(
        f"{exc_col}2:{exc_col}{last_row}",
        FormulaRule(formula=[f'LEN(TRIM({exc_col}2))>0'], fill=PatternFill("solid", fgColor="B7E1CD")),
    )

    for col in range(1, len(INELIGIBLE_HEADERS) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18


def _write_possible_cause_sheet(ws, rows: list[dict[str, Any]], site_title: str) -> None:
    ws.delete_rows(1, ws.max_row)
    ws.cell(row=1, column=1, value=site_title)
    ws.cell(row=2, column=1, value="Exception Reason")
    ws.cell(row=2, column=2, value="Possible Cause")
    ws.cell(row=2, column=3, value="Number of Transactions")
    ws.cell(row=2, column=4, value="SUM of Litres")
    ws.cell(row=2, column=5, value="Analyst Notes")
    for idx, row in enumerate(rows, 3):
        ws.cell(row=idx, column=1, value=row.get("exception_reason"))
        ws.cell(row=idx, column=2, value=row.get("possible_cause"))
        ws.cell(row=idx, column=3, value=row.get("transaction_count"))
        ws.cell(row=idx, column=4, value=row.get("litres"))
    total_row = len(rows) + 3
    ws.cell(row=total_row, column=1, value="Total")
    ws.cell(row=total_row, column=3, value=sum(int(row.get("transaction_count") or 0) for row in rows))
    ws.cell(row=total_row, column=4, value=sum(float(row.get("litres") or 0) for row in rows))


def _write_summary_per_asset_sheet(ws, rows: list[dict[str, Any]], site_title: str) -> None:
    ws.delete_rows(1, ws.max_row)
    ws.cell(row=1, column=1, value=site_title)
    ws.cell(row=2, column=1, value="Asset Number")
    ws.cell(row=2, column=2, value="Asset Description")
    ws.cell(row=2, column=3, value="Department")
    ws.cell(row=2, column=4, value="Number of Transactions")
    ws.cell(row=2, column=5, value="SUM of Litres")
    for idx, row in enumerate(rows, 3):
        ws.cell(row=idx, column=1, value=row.get("asset_number"))
        ws.cell(row=idx, column=2, value=row.get("asset_description"))
        ws.cell(row=idx, column=3, value=row.get("department"))
        ws.cell(row=idx, column=4, value=row.get("transaction_count"))
        ws.cell(row=idx, column=5, value=row.get("litres"))
    total_row = len(rows) + 3
    ws.cell(row=total_row, column=1, value="Total")
    ws.cell(row=total_row, column=4, value=sum(int(row.get("transaction_count") or 0) for row in rows))
    ws.cell(row=total_row, column=5, value=sum(float(row.get("litres") or 0) for row in rows))


def _write_non_mining_sheet(ws, rows: list[dict[str, Any]]) -> None:
    ws.title = "Non-Mining Excluded"
    ws.delete_rows(1, ws.max_row)
    for col, title in enumerate(INELIGIBLE_HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    for row_idx, record in enumerate(rows, 2):
        _write_row_from_record(ws, row_idx, record, INELIGIBLE_HEADERS, INELIGIBLE_FIELD_TO_HEADER)


def _write_glossary_sheet(ws) -> None:
    ws.title = "Exception Reason Glossary"
    ws.delete_rows(1, ws.max_row)
    headers = ["Exception Reason", "Typical Cause", "Meaning", "Analyst guidance"]
    for col, title in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    for row_idx, row in enumerate(EXCEPTION_GLOSSARY, 2):
        for col, value in enumerate(row, 1):
            ws.cell(row=row_idx, column=col, value=value)
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 28


def _copy_lookup_sheet(
    source_path: str | None,
    wb: openpyxl.Workbook,
    sheet_name: str,
    required_headers: set[str] | None = None,
) -> None:
    if not source_path:
        return
    src = openpyxl.load_workbook(source_path, data_only=False)
    src_ws = None
    if sheet_name in src.sheetnames:
        src_ws = src[sheet_name]
    elif required_headers:
        for name in src.sheetnames:
            ws = src[name]
            headers = {
                str(ws.cell(1, c).value or "").strip().lower()
                for c in range(1, ws.max_column + 1)
            }
            if required_headers.issubset(headers):
                src_ws = ws
                break
    if src_ws is None:
        src_ws = src[src.sheetnames[0]]
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    tgt = wb.create_sheet(sheet_name)
    for row in src_ws.iter_rows():
        for cell in row:
            tgt_cell = tgt.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                tgt_cell.font = copy(cell.font)
                tgt_cell.fill = copy(cell.fill)
                tgt_cell.border = copy(cell.border)
                tgt_cell.alignment = copy(cell.alignment)
                tgt_cell.number_format = cell.number_format


def _reorder_sheets(wb: openpyxl.Workbook) -> None:
    ordered = [wb[name] for name in SHEET_ORDER if name in wb.sheetnames]
    for name in wb.sheetnames:
        if name not in SHEET_ORDER:
            ordered.append(wb[name])
    wb._sheets = ordered


def build_summary_json(
    *,
    transactions: list[dict[str, Any]],
    review_queue: list[dict[str, Any]],
    possible_causes: list[dict[str, Any]],
    summary_per_asset: list[dict[str, Any]],
    excluded_non_mining: int,
    flagged_count: int,
    avr_sync_count: int,
    site_name: str,
    rule_profile: str | None = None,
    rule_profile_label: str | None = None,
    month_over_month: dict[str, Any] | None = None,
) -> dict[str, Any]:
    review_litres = round(sum(float(row.get("litres") or 0) for row in review_queue), 2)
    return {
        "site_name": site_name,
        "transaction_count": len(transactions),
        "review_queue_count": len(review_queue),
        "review_queue_litres": review_litres,
        "flagged_exception_count": flagged_count,
        "avr_sync_count": avr_sync_count,
        "excluded_non_mining_count": excluded_non_mining,
        "possible_cause_groups": len(possible_causes),
        "summary_asset_count": len(summary_per_asset),
        "rule_profile": rule_profile,
        "rule_profile_label": rule_profile_label,
        "possible_causes": possible_causes,
        "summary_per_asset": summary_per_asset[:10],
        "review_sample": [
            {
                "transaction_id": row.get("transaction_id"),
                "asset_number": row.get("asset_number"),
                "litres": row.get("litres"),
                "review_reason": row.get("review_reason"),
                "exception_reason": row.get("exception_60"),
                "suggested_abco_comment": row.get("suggested_abco_comment"),
            }
            for row in review_queue[:10]
        ],
        "month_over_month": month_over_month,
        "has_errors": False,
    }


def write_prepared_workbook(
    *,
    input_path: str,
    output_path: str,
    transactions: list[dict[str, Any]],
    review_queue: list[dict[str, Any]],
    possible_causes: list[dict[str, Any]],
    summary_per_asset: list[dict[str, Any]],
    excluded_non_mining_rows: list[dict[str, Any]],
    asset_lookup_path: str | None,
    avr_sync_path: str | None,
    site_name: str,
) -> None:
    shutil.copy2(input_path, output_path)
    wb = openpyxl.load_workbook(output_path)

    for sheet_name in list(wb.sheetnames):
        if sheet_name not in {DETAILS_SHEET, "Asset Info Lookup"}:
            del wb[sheet_name]

    txn_by_id = {
        str(row.get("transaction_id") or ""): row
        for row in transactions
        if row.get("transaction_id")
    }
    _apply_light_touch_details(wb[DETAILS_SHEET], txn_by_id)

    ineligible_ws = wb.create_sheet("Transactions deemed ineligible")
    _write_ineligible_sheet(ineligible_ws, review_queue)

    pcs = wb.create_sheet("Possible Cause Summary")
    _write_possible_cause_sheet(
        pcs,
        possible_causes,
        summary_sheet_title(site_name, "Possible Cause"),
    )

    spa = wb.create_sheet("Summary Per Asset")
    _write_summary_per_asset_sheet(
        spa,
        summary_per_asset,
        summary_sheet_title(site_name, "Per Asset"),
    )

    non_mining = wb.create_sheet("Non-Mining Excluded")
    _write_non_mining_sheet(non_mining, excluded_non_mining_rows)

    glossary = wb.create_sheet("Exception Reason Glossary")
    _write_glossary_sheet(glossary)

    if asset_lookup_path:
        _copy_lookup_sheet(
            asset_lookup_path,
            wb,
            "Asset Info Lookup",
            {"asset number", "asset group"},
        )
    _copy_lookup_sheet(
        avr_sync_path,
        wb,
        "AVR Sync Lookup",
        {"id", "code"},
    )

    _reorder_sheets(wb)
    wb.save(output_path)


def infer_site_name(input_path: str) -> str:
    stem = Path(input_path).stem
    stem = stem.replace(" - Transaction Exceptions - In Context", "")
    stem = re.sub(r"\s*\(\d+\)\s*$", "", stem)
    return stem.strip() or "Site"


def summary_sheet_title(site_name: str, kind: str) -> str:
    month_match = re.search(
        r"(January|February|March|April|May|June|July|August|September|October|November|December)\s+\d{4}",
        site_name,
        re.IGNORECASE,
    )
    month = month_match.group(0) if month_match else ""
    base = re.sub(
        r"\s*-\s*(January|February|March|April|May|June|July|August|September|October|November|December)\s+\d{4}\s*$",
        "",
        site_name,
        flags=re.IGNORECASE,
    ).strip()
    if month:
        return f"{base} - Summary of Dispense Review {kind} - {month}"
    return f"{base} - Summary of Dispense Review {kind}"
