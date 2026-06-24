"""Month-over-month comparison against a prior prepared workbook."""
from __future__ import annotations

from typing import Any

import openpyxl


def _review_ids_from_workbook(path: str) -> set[str]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if "Transactions deemed ineligible" not in wb.sheetnames:
        return set()
    ws = wb["Transactions deemed ineligible"]
    headers: dict[str, int] = {}
    ids: set[str] = set()
    for row in ws.iter_rows(values_only=True):
        if not row:
            continue
        if not headers:
            headers = {
                str(value or "").strip().lower(): idx
                for idx, value in enumerate(row)
                if value
            }
            continue
        txn_idx = headers.get("transaction id")
        if txn_idx is None:
            continue
        txn_id = str(row[txn_idx] or "").strip()
        if txn_id and txn_id.lower() not in {"transaction id", "total", "uploaded"}:
            if not txn_id.lower().startswith("uploaded"):
                ids.add(txn_id)
    return ids


def compare_month_over_month(
    review_queue: list[dict[str, Any]],
    prior_path: str | None,
) -> dict[str, Any] | None:
    if not prior_path:
        return None

    prior_ids = _review_ids_from_workbook(prior_path)
    current_ids = {str(row.get("transaction_id") or "") for row in review_queue}
    current_ids.discard("")

    current_assets: dict[str, float] = {}
    prior_assets: dict[str, float] = {}
    for row in review_queue:
        asset = str(row.get("asset_number") or "UNKNOWN")
        current_assets[asset] = current_assets.get(asset, 0.0) + float(row.get("litres") or 0)

    new_ids = sorted(current_ids - prior_ids)
    repeat_ids = sorted(current_ids & prior_ids)
    dropped_ids = sorted(prior_ids - current_ids)
    new_assets = sorted(set(current_assets) - set(prior_assets))

    repeat_litres = round(
        sum(float(row.get("litres") or 0) for row in review_queue if str(row.get("transaction_id")) in repeat_ids),
        2,
    )
    new_litres = round(
        sum(float(row.get("litres") or 0) for row in review_queue if str(row.get("transaction_id")) in new_ids),
        2,
    )

    return {
        "prior_review_count": len(prior_ids),
        "current_review_count": len(current_ids),
        "new_in_review_count": len(new_ids),
        "repeat_in_review_count": len(repeat_ids),
        "dropped_from_review_count": len(dropped_ids),
        "new_in_review_ids": new_ids[:25],
        "repeat_in_review_ids": repeat_ids[:25],
        "dropped_from_review_ids": dropped_ids[:25],
        "new_assets_in_review": new_assets[:25],
        "new_review_litres": new_litres,
        "repeat_review_litres": repeat_litres,
    }
