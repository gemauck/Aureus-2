"""Build fleet and owner routing indexes from a reference Transactions workbook."""
from __future__ import annotations

import re
from pathlib import Path
from typing import Any

import openpyxl

TRANSACTION_HEADERS = (
    "Fleet ID",
    "Registration Number",
    "Make",
    "Model",
    "TransactionDateTime",
    "Date",
    "Time",
    "Liters",
    "Location",
    "Device ID",
    "Pump",
    "Fleet Type",
    "VehicleCategory Description",
    "Responsibility Code",
    "BusinessRevenue Code",
    "Product Name",
    "Cost Centre",
    "TaxRebate Code",
    "Consumption Meter",
    "Consumption Type",
    "Hours/OD reading",
    "Group1",
    "Group2",
    "Group3",
    "Group4",
    "Group5",
    "Voucher Number",
    "Pump Controller",
    "Transaction Plate",
    "Operator ID",
    "Operator Name",
)

FLEET_FIELDS = (
    "Make",
    "Model",
    "Location",
    "Device ID",
    "Pump",
    "Pump Controller",
    "VehicleCategory Description",
    "Group1",
    "Group2",
    "Group3",
    "Group4",
    "Group5",
    "Product Name",
    "Consumption Meter",
    "Consumption Type",
    "Responsibility Code",
    "BusinessRevenue Code",
    "Cost Centre",
    "TaxRebate Code",
    "Fleet Type",
)


def _cell_str(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _normalize_pump_lookup_key(value: Any) -> str:
    return re.sub(r"\s+", " ", _cell_str(value)).lower()


def _pump_route_keys(device_id: str, pump: str) -> list[str]:
    """Keys used to match Sparrow Fuel Pump text to a Gilbarco location/device/pump."""
    keys: list[str] = []
    seen: set[str] = set()

    def add(raw: str) -> None:
        k = _normalize_pump_lookup_key(raw)
        if k and k not in seen:
            seen.add(k)
            keys.append(k)

    if device_id:
        add(device_id)
    if pump:
        add(pump)
        if " - " in pump:
            add(pump.split(" - ")[-1])
        parts = [p.strip() for p in pump.split() if p.strip()]
        if len(parts) >= 2:
            add(" ".join(parts[-2:]))
        if parts:
            add(parts[-1])
    return keys


def build_fuel_pump_routes(path: str | Path) -> dict[str, dict[str, Any]]:
    """Map normalized Fuel Pump / device strings to Gilbarco pump routes from a reference workbook."""
    path = Path(path)
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheet = (
        "All Transactions"
        if "All Transactions" in wb.sheetnames
        else wb.sheetnames[-1]
    )
    ws = wb[sheet]
    header = [_cell_str(c) for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())]

    routes: dict[str, dict[str, Any]] = {}
    for cells in ws.iter_rows(min_row=2, values_only=True):
        row = {header[i]: cells[i] if i < len(cells) else None for i in range(len(header))}
        device_id = _cell_str(row.get("Device ID"))
        pump = _cell_str(row.get("Pump"))
        if not device_id and not pump:
            continue
        route = {
            "location": row.get("Location"),
            "device_id": row.get("Device ID"),
            "pump": row.get("Pump"),
            "pump_controller": row.get("Pump Controller"),
        }
        for key in _pump_route_keys(device_id, pump):
            routes.setdefault(key, route)

    wb.close()
    return routes


def load_reference_indexes(
    path: str | Path,
) -> tuple[dict[str, dict[str, Any]], dict[str, dict[str, Any]], dict[str, dict[str, Any]]]:
    """Return (fleet_by_id, owner_pump_routes, bowser_by_fleet)."""
    path = Path(path)
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheet = (
        "All Transactions"
        if "All Transactions" in wb.sheetnames
        else wb.sheetnames[-1]
    )
    ws = wb[sheet]
    header = [_cell_str(c) for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())]

    fleet_by_id: dict[str, dict[str, Any]] = {}
    owner_routes: dict[str, dict[str, Any]] = {}
    bowser_by_fleet: dict[str, dict[str, Any]] = {}

    for cells in ws.iter_rows(min_row=2, values_only=True):
        row = {header[i]: cells[i] if i < len(cells) else None for i in range(len(header))}
        fleet_id = _cell_str(row.get("Fleet ID"))
        if not fleet_id:
            continue

        vcat = _cell_str(row.get("VehicleCategory Description"))
        if vcat == "BOWSER":
            if fleet_id not in bowser_by_fleet:
                bowser_by_fleet[fleet_id] = {k: row.get(k) for k in TRANSACTION_HEADERS}
            continue

        if fleet_id not in fleet_by_id:
            fleet_by_id[fleet_id] = {field: row.get(field) for field in FLEET_FIELDS}

        owner = _cell_str(row.get("Group3"))
        pump = _cell_str(row.get("Pump"))
        if owner and pump and owner not in owner_routes:
            owner_routes[owner] = {
                "location": row.get("Location"),
                "device_id": row.get("Device ID"),
                "pump": row.get("Pump"),
                "pump_controller": row.get("Pump Controller"),
            }

    wb.close()
    return fleet_by_id, owner_routes, bowser_by_fleet
