"""Convert Fuel Dispense Report rows to Transactions & Fuel Breakdown format."""
from __future__ import annotations

import json
import re
from datetime import date, datetime
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl import Workbook

from fleet_lookup import (
    TRANSACTION_HEADERS,
    build_fuel_pump_routes,
    load_reference_indexes,
)
from parse_dispense import (
    DISPENSE_HEADERS,
    GILBARCO_SECTION_TITLE,
    ORIGINAL_SECTION_TITLE,
    OUTPUT_SHEET_NAME,
    build_output_filename,
    dispense_period_range,
    extract_pump_code,
    normalize_asset_group,
    normalize_asset_owner,
    parse_consumption,
    parse_dispense_workbook,
    resolve_make_model,
    split_asset_description,
)
from winshuttle import (
    build_winshuttle_filename,
    format_winshuttle_report_date,
    winshuttle_config,
    write_winshuttle_workbook,
)

SCRIPT_DIR = Path(__file__).resolve().parent
DEFAULT_CONFIG_PATH = SCRIPT_DIR / "pump_config.json"
DEFAULT_GILBARCO_TEMPLATE = SCRIPT_DIR / "gilbarco-template.xlsx"


def _cell_str(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _parse_number(value: Any) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    s = _cell_str(value).replace(",", "")
    if not s:
        return None
    try:
        return float(s)
    except ValueError:
        return None


def load_pump_config(path: str | Path | None = None) -> dict[str, Any]:
    config_path = Path(path) if path else DEFAULT_CONFIG_PATH
    with config_path.open(encoding="utf-8") as fh:
        return json.load(fh)


def merge_fuel_pump_routes(
    reference_routes: dict[str, dict[str, Any]],
    config: dict[str, Any],
) -> dict[str, dict[str, Any]]:
    """Reference routes first; config sparrow_pump_to_route overrides."""
    merged = dict(reference_routes)
    for raw_key, route in config.get("sparrow_pump_to_route", {}).items():
        merged[_normalize_lookup_key(raw_key)] = route
    return merged


def _normalize_lookup_key(value: Any) -> str:
    return re.sub(r"\s+", " ", _cell_str(value)).lower()


def resolve_fleet_identity(
    asset_id: str,
    config: dict[str, Any],
    bowser_by_fleet: dict[str, dict[str, Any]],
) -> tuple[str | None, str | None]:
    """Return (vehicle_fleet_id for template lookup, bowser_fleet_id if bowser path)."""
    if not asset_id:
        return None, None

    entry = config.get("fleet_id_aliases", {}).get(asset_id)
    if isinstance(entry, dict):
        bowser_id = _cell_str(entry.get("bowser_fleet_id"))
        if bowser_id:
            return None, bowser_id
    elif isinstance(entry, str) and entry:
        asset_id = entry

    if asset_id in bowser_by_fleet:
        return None, asset_id
    return asset_id, None


def resolve_bowser_fleet_id(
    row: dict[str, Any],
    config: dict[str, Any],
    bowser_by_fleet: dict[str, dict[str, Any]],
) -> str | None:
    """Bowser path before vehicle conversion when asset or heuristics match a site bowser."""
    asset_id = _cell_str(row.get("Asset Number")) or _cell_str(row.get("Asset Registration"))
    _vehicle_id, bowser_id = resolve_fleet_identity(asset_id, config, bowser_by_fleet)
    if bowser_id:
        return bowser_id

    if asset_id and asset_id in bowser_by_fleet:
        return asset_id

    desc = _cell_str(row.get("Asset Description")).upper()
    if any(
        token in desc
        for token in ("BOWSER", "BULK TANK", "MOBILE DIESEL BOWSER", "BULK DIESEL TANK")
    ):
        base = re.sub(r"\s*-\s*BULK.*", "", asset_id, flags=re.IGNORECASE).strip()
        if base in bowser_by_fleet:
            return base

    if asset_id and re.search(r"\s*-\s*BULK", asset_id, re.IGNORECASE):
        base = re.sub(r"\s*-\s*BULK.*", "", asset_id, flags=re.IGNORECASE).strip()
        if base in bowser_by_fleet:
            return base

    fuel = _cell_str(row.get("Fuel Pump")).lower()
    if "service bay" in fuel:
        for candidate in ("OTK105M", "OTK8BULK", "OTK8"):
            if candidate in bowser_by_fleet:
                return candidate

    return None


def lookup_fuel_pump_route(
    fuel_pump: Any,
    config: dict[str, Any],
    fuel_pump_routes: dict[str, dict[str, Any]],
) -> dict[str, Any]:
    raw = _cell_str(fuel_pump)
    if not raw:
        return {}

    aliases = config.get("fuel_pump_aliases", {})
    candidates = [raw, extract_pump_code(raw, aliases)]
    for candidate in candidates:
        if not candidate:
            continue
        key = _normalize_lookup_key(candidate)
        if key in fuel_pump_routes:
            route = fuel_pump_routes[key]
            return {
                "location": route.get("location"),
                "device_id": route.get("device_id"),
                "pump": route.get("pump"),
                "pump_controller": route.get("pump_controller"),
            }
    return {}


def format_transaction_datetime(value: Any) -> tuple[str, str, str]:
    if isinstance(value, datetime):
        dt = value
    elif isinstance(value, date):
        dt = datetime.combine(value, datetime.min.time())
    else:
        s = _cell_str(value)
        for fmt in (
            "%Y-%m-%d %H:%M:%S",
            "%Y-%m-%d %H:%M",
            "%d/%m/%Y %H:%M:%S",
            "%d/%m/%Y %H:%M",
        ):
            try:
                dt = datetime.strptime(s, fmt)
                break
            except ValueError:
                dt = None
        if dt is None:
            raise ValueError(f"Unrecognized date/time: {value!r}")

    display = dt.strftime("%d/%m/%Y %H:%M:%S")
    return display, dt.strftime("%d/%m/%Y"), dt.strftime("%H:%M:%S")


def _route_matches(match: dict[str, Any], row: dict[str, Any]) -> bool:
    group1 = _cell_str(row.get("Group1"))
    owner = normalize_asset_owner(row.get("Asset Owner"))
    if "group1_contains" in match and match["group1_contains"].lower() not in group1.lower():
        return False
    if "asset_owner" in match:
        expected = normalize_asset_owner(match["asset_owner"])
        if owner != expected:
            return False
    return True


def resolve_vehicle_route(
    row: dict[str, Any],
    fleet_template: dict[str, Any] | None,
    owner_routes: dict[str, dict[str, Any]],
    config: dict[str, Any],
    fuel_pump_routes: dict[str, dict[str, Any]],
) -> dict[str, Any]:
    owner = normalize_asset_owner(row.get("Asset Owner"))

    if fleet_template and _cell_str(fleet_template.get("Pump")):
        return {
            "location": fleet_template.get("Location"),
            "device_id": fleet_template.get("Device ID"),
            "pump": fleet_template.get("Pump"),
            "pump_controller": fleet_template.get("Pump Controller"),
            "group1": fleet_template.get("Group1"),
        }

    fuel_route = lookup_fuel_pump_route(row.get("Fuel Pump"), config, fuel_pump_routes)
    if fuel_route.get("pump"):
        return {**fuel_route, "group1": row.get("Group1")}

    for route in config.get("vehicle_routes", []):
        if _route_matches(route.get("match", {}), row):
            return {
                "location": route.get("location"),
                "device_id": route.get("device_id"),
                "pump": route.get("pump"),
                "pump_controller": route.get("pump_controller"),
                "group1": route.get("group1"),
            }

    if owner and owner in owner_routes:
        ref = owner_routes[owner]
        return {
            "location": ref.get("location"),
            "device_id": ref.get("device_id"),
            "pump": ref.get("pump"),
            "pump_controller": ref.get("pump_controller"),
            "group1": row.get("Group1"),
        }

    if fleet_template:
        return {
            "location": fleet_template.get("Location"),
            "device_id": fleet_template.get("Device ID"),
            "pump": fleet_template.get("Pump"),
            "pump_controller": fleet_template.get("Pump Controller"),
            "group1": fleet_template.get("Group1"),
        }

    return {}


def convert_vehicle_row(
    row: dict[str, Any],
    fleet_by_id: dict[str, dict[str, Any]],
    owner_routes: dict[str, dict[str, Any]],
    config: dict[str, Any],
    warnings: list[str],
    fuel_pump_routes: dict[str, dict[str, Any]] | None = None,
) -> dict[str, Any] | None:
    asset_id = _cell_str(row.get("Asset Number")) or _cell_str(row.get("Asset Registration"))
    if not asset_id:
        return None

    vehicle_fleet_id, _bowser_id = resolve_fleet_identity(asset_id, config, {})
    fleet_id = vehicle_fleet_id if vehicle_fleet_id else asset_id

    litres = _parse_number(row.get("Litres"))
    if litres is None:
        warnings.append(f"No litres on asset row (API {_cell_str(row.get('API ID'))})")

    fleet_template = fleet_by_id.get(fleet_id)
    route = resolve_vehicle_route(
        row,
        fleet_template,
        owner_routes,
        config,
        fuel_pump_routes or {},
    )

    if fleet_template:
        make, model = resolve_make_model(row.get("Asset Description"), fleet_template)
        vcat = fleet_template.get("VehicleCategory Description")
        group2 = fleet_template.get("Group2")
        group3 = fleet_template.get("Group3")
        group4 = fleet_template.get("Group4")
        group5 = fleet_template.get("Group5")
        product = fleet_template.get("Product Name") or "DIESEL"
        meter = fleet_template.get("Consumption Meter")
        ctype = fleet_template.get("Consumption Type")
        group1 = route.get("group1") if route.get("group1") is not None else fleet_template.get("Group1")
    else:
        make, model = split_asset_description(row.get("Asset Description"))
        vcat = _cell_str(row.get("Operation")) or None
        group2 = normalize_asset_group(row.get("Asset Group")) or None
        group3 = normalize_asset_owner(row.get("Asset Owner")) or None
        group4 = row.get("Gilbarco Unique Asset ID")
        group5 = row.get("Internal Order Number")
        product = "DIESEL"
        meter, ctype, _ = parse_consumption(row.get("Odometer"), row.get("Economy"))
        group1 = route.get("group1") if route.get("group1") is not None else row.get("Group1")
        warnings.append(f"No fleet template for {fleet_id}; using dispense-derived fields")

    output_fleet_id = fleet_id

    meter, ctype, reading = parse_consumption(row.get("Odometer"), row.get("Economy"))
    if fleet_template and fleet_template.get("Consumption Meter"):
        meter = fleet_template.get("Consumption Meter") or meter
        ctype = fleet_template.get("Consumption Type") or ctype

    tx_dt, tx_date, tx_time = format_transaction_datetime(row.get("Date & Time"))
    location = route.get("location") or (fleet_template or {}).get("Location") or group3
    if not route.get("pump"):
        warnings.append(f"Missing pump route for {output_fleet_id} ({group3})")

    reg = _cell_str(row.get("Asset Registration")) or output_fleet_id

    return {
        "Fleet ID": output_fleet_id,
        "Registration Number": reg,
        "Make": make,
        "Model": model,
        "TransactionDateTime": tx_dt,
        "Date": tx_date,
        "Time": tx_time,
        "Liters": litres,
        "Location": location,
        "Device ID": route.get("device_id"),
        "Pump": route.get("pump"),
        "Fleet Type": (fleet_template or {}).get("Fleet Type"),
        "VehicleCategory Description": vcat,
        "Responsibility Code": (fleet_template or {}).get("Responsibility Code"),
        "BusinessRevenue Code": (fleet_template or {}).get("BusinessRevenue Code"),
        "Product Name": product,
        "Cost Centre": (fleet_template or {}).get("Cost Centre"),
        "TaxRebate Code": (fleet_template or {}).get("TaxRebate Code"),
        "Consumption Meter": meter,
        "Consumption Type": ctype,
        "Hours/OD reading": reading,
        "Group1": group1,
        "Group2": group2 or normalize_asset_group(row.get("Asset Group")) or None,
        "Group3": group3,
        "Group4": group4 if group4 is not None else (fleet_template or {}).get("Group4"),
        "Group5": group5 if group5 is not None else (fleet_template or {}).get("Group5"),
        "Voucher Number": str(row.get("API ID")) if row.get("API ID") is not None else None,
        "Pump Controller": route.get("pump_controller"),
        "Transaction Plate": output_fleet_id,
        "Operator ID": _cell_str(row.get("User Tag")) or None,
        "Operator Name": _cell_str(row.get("Operator")) or None,
    }


def convert_bowser_row_for_asset(
    row: dict[str, Any],
    bowser_fleet_id: str,
    bowser_by_fleet: dict[str, dict[str, Any]],
    config: dict[str, Any],
) -> dict[str, Any] | None:
    """Gilbarco bowser row when Sparrow still supplies an asset number on a bowser fill."""
    litres = _parse_number(row.get("Litres"))
    if litres is None:
        return None

    template_cfg = config.get("bowser_template", {})
    ref = bowser_by_fleet.get(bowser_fleet_id, {})
    ref_fields = {
        "Make": ref.get("Make") or template_cfg.get("make"),
        "Model": ref.get("Model") or template_cfg.get("model"),
    }
    make, model = resolve_make_model(row.get("Asset Description"), ref_fields)
    tx_dt, tx_date, tx_time = format_transaction_datetime(row.get("Date & Time"))

    return {
        "Fleet ID": bowser_fleet_id,
        "Registration Number": bowser_fleet_id,
        "Make": make or ref_fields.get("Make"),
        "Model": model or ref_fields.get("Model"),
        "TransactionDateTime": tx_dt,
        "Date": tx_date,
        "Time": tx_time,
        "Liters": litres,
        "Location": ref.get("Location") or template_cfg.get("location"),
        "Device ID": ref.get("Device ID") or template_cfg.get("device_id"),
        "Pump": ref.get("Pump") or template_cfg.get("pump"),
        "Fleet Type": ref.get("Fleet Type"),
        "VehicleCategory Description": ref.get("VehicleCategory Description")
        or template_cfg.get("vehicle_category"),
        "Responsibility Code": ref.get("Responsibility Code")
        or template_cfg.get("responsibility_code"),
        "BusinessRevenue Code": ref.get("BusinessRevenue Code")
        or template_cfg.get("business_revenue_code"),
        "Product Name": ref.get("Product Name") or template_cfg.get("product_name"),
        "Cost Centre": ref.get("Cost Centre") or template_cfg.get("cost_centre"),
        "TaxRebate Code": ref.get("TaxRebate Code") or template_cfg.get("tax_rebate_code"),
        "Consumption Meter": ref.get("Consumption Meter") or template_cfg.get("consumption_meter"),
        "Consumption Type": ref.get("Consumption Type") or template_cfg.get("consumption_type"),
        "Hours/OD reading": 0.0,
        "Group1": ref.get("Group1") or template_cfg.get("group1"),
        "Group2": ref.get("Group2") or template_cfg.get("group2"),
        "Group3": ref.get("Group3") or template_cfg.get("group3"),
        "Group4": ref.get("Group4") or template_cfg.get("group4"),
        "Group5": ref.get("Group5") or template_cfg.get("group5"),
        "Voucher Number": str(row.get("API ID")) if row.get("API ID") is not None else None,
        "Pump Controller": ref.get("Pump Controller") or template_cfg.get("pump_controller"),
        "Transaction Plate": bowser_fleet_id,
        "Operator ID": _cell_str(row.get("User Tag")) or None,
        "Operator Name": _cell_str(row.get("Operator")) or None,
    }


def convert_bowser_row(
    row: dict[str, Any],
    bowser_by_fleet: dict[str, dict[str, Any]],
    config: dict[str, Any],
) -> dict[str, Any] | None:
    litres = _parse_number(row.get("Litres"))
    if litres is None:
        return None

    aliases = config.get("fuel_pump_aliases", {})
    pump_code = extract_pump_code(row.get("Fuel Pump"), aliases)
    if not pump_code:
        return None

    template_cfg = config.get("bowser_template", {})
    ref = bowser_by_fleet.get(pump_code, {})

    tx_dt, tx_date, tx_time = format_transaction_datetime(row.get("Date & Time"))

    return {
        "Fleet ID": pump_code,
        "Registration Number": pump_code,
        "Make": ref.get("Make") or template_cfg.get("make"),
        "Model": ref.get("Model") or template_cfg.get("model"),
        "TransactionDateTime": tx_dt,
        "Date": tx_date,
        "Time": tx_time,
        "Liters": litres,
        "Location": ref.get("Location") or template_cfg.get("location"),
        "Device ID": ref.get("Device ID") or template_cfg.get("device_id"),
        "Pump": ref.get("Pump") or template_cfg.get("pump"),
        "Fleet Type": ref.get("Fleet Type"),
        "VehicleCategory Description": ref.get("VehicleCategory Description")
        or template_cfg.get("vehicle_category"),
        "Responsibility Code": ref.get("Responsibility Code")
        or template_cfg.get("responsibility_code"),
        "BusinessRevenue Code": ref.get("BusinessRevenue Code")
        or template_cfg.get("business_revenue_code"),
        "Product Name": ref.get("Product Name") or template_cfg.get("product_name"),
        "Cost Centre": ref.get("Cost Centre") or template_cfg.get("cost_centre"),
        "TaxRebate Code": ref.get("TaxRebate Code") or template_cfg.get("tax_rebate_code"),
        "Consumption Meter": ref.get("Consumption Meter") or template_cfg.get("consumption_meter"),
        "Consumption Type": ref.get("Consumption Type") or template_cfg.get("consumption_type"),
        "Hours/OD reading": 0.0,
        "Group1": ref.get("Group1") or template_cfg.get("group1"),
        "Group2": ref.get("Group2") or template_cfg.get("group2"),
        "Group3": ref.get("Group3") or template_cfg.get("group3"),
        "Group4": ref.get("Group4") or template_cfg.get("group4"),
        "Group5": ref.get("Group5") or template_cfg.get("group5"),
        "Voucher Number": str(row.get("API ID")) if row.get("API ID") is not None else None,
        "Pump Controller": ref.get("Pump Controller") or template_cfg.get("pump_controller"),
        "Transaction Plate": pump_code,
        "Operator ID": None,
        "Operator Name": _cell_str(row.get("Operator")) or None,
    }


def convert_unallocated_row(
    row: dict[str, Any],
    bowser_by_fleet: dict[str, dict[str, Any]],
    config: dict[str, Any],
    warnings: list[str],
) -> dict[str, Any]:
    """Gilbarco row for dispense lines without an asset (overrides, bowser fills, tests)."""
    bowser = convert_bowser_row(row, bowser_by_fleet, config)
    if bowser:
        return bowser

    litres = _parse_number(row.get("Litres"))
    aliases = config.get("fuel_pump_aliases", {})
    pump_code = extract_pump_code(row.get("Fuel Pump"), aliases) or "UNALLOCATED"
    template_cfg = config.get("bowser_template", {})

    try:
        tx_dt, tx_date, tx_time = format_transaction_datetime(row.get("Date & Time"))
    except ValueError:
        warnings.append(f"Unrecognized date on row API {_cell_str(row.get('API ID'))}")
        tx_dt = tx_date = tx_time = None

    return {
        "Fleet ID": pump_code,
        "Registration Number": pump_code,
        "Make": template_cfg.get("make"),
        "Model": template_cfg.get("model"),
        "TransactionDateTime": tx_dt,
        "Date": tx_date,
        "Time": tx_time,
        "Liters": litres,
        "Location": template_cfg.get("location"),
        "Device ID": template_cfg.get("device_id"),
        "Pump": template_cfg.get("pump"),
        "Fleet Type": None,
        "VehicleCategory Description": "UNALLOCATED",
        "Responsibility Code": template_cfg.get("responsibility_code"),
        "BusinessRevenue Code": template_cfg.get("business_revenue_code"),
        "Product Name": template_cfg.get("product_name") or "DIESEL",
        "Cost Centre": template_cfg.get("cost_centre"),
        "TaxRebate Code": template_cfg.get("tax_rebate_code"),
        "Consumption Meter": template_cfg.get("consumption_meter") or "None",
        "Consumption Type": template_cfg.get("consumption_type") or "None",
        "Hours/OD reading": 0.0,
        "Group1": template_cfg.get("group1"),
        "Group2": template_cfg.get("group2"),
        "Group3": template_cfg.get("group3"),
        "Group4": template_cfg.get("group4"),
        "Group5": template_cfg.get("group5"),
        "Voucher Number": str(row.get("API ID")) if row.get("API ID") is not None else None,
        "Pump Controller": template_cfg.get("pump_controller"),
        "Transaction Plate": pump_code,
        "Operator ID": None,
        "Operator Name": _cell_str(row.get("Operator")) or None,
    }


def convert_row_to_gilbarco(
    row: dict[str, Any],
    fleet_by_id: dict[str, dict[str, Any]],
    owner_routes: dict[str, dict[str, Any]],
    bowser_by_fleet: dict[str, dict[str, Any]],
    config: dict[str, Any],
    warnings: list[str],
    fuel_pump_routes: dict[str, dict[str, Any]] | None = None,
) -> dict[str, Any]:
    bowser_fleet_id = resolve_bowser_fleet_id(row, config, bowser_by_fleet)
    if bowser_fleet_id:
        converted = convert_bowser_row_for_asset(
            row, bowser_fleet_id, bowser_by_fleet, config
        )
        if converted:
            return converted

    fleet_id = _cell_str(row.get("Asset Number")) or _cell_str(row.get("Asset Registration"))
    if fleet_id:
        converted = convert_vehicle_row(
            row,
            fleet_by_id,
            owner_routes,
            config,
            warnings,
            fuel_pump_routes,
        )
        if converted:
            return converted
    return convert_unallocated_row(row, bowser_by_fleet, config, warnings)


def convert_dispense_rows(
    dispense_rows: list[dict[str, Any]],
    fleet_by_id: dict[str, dict[str, Any]],
    owner_routes: dict[str, dict[str, Any]],
    bowser_by_fleet: dict[str, dict[str, Any]],
    config: dict[str, Any],
    fuel_pump_routes: dict[str, dict[str, Any]] | None = None,
) -> tuple[list[dict[str, Any]], list[str]]:
    """Return (gilbarco_rows aligned 1:1 with dispense_rows, warnings)."""
    warnings: list[str] = []
    gilbarco_rows: list[dict[str, Any]] = []
    routes = fuel_pump_routes or {}
    for row in dispense_rows:
        gilbarco_rows.append(
            convert_row_to_gilbarco(
                row,
                fleet_by_id,
                owner_routes,
                bowser_by_fleet,
                config,
                warnings,
                routes,
            )
        )
    return gilbarco_rows, warnings


def write_side_by_side_workbook(
    dispense_rows: list[dict[str, Any]],
    gilbarco_rows: list[dict[str, Any]],
    output_path: Path,
) -> None:
    """One sheet: Gilbarco columns left, original dispense columns right (same row)."""
    if len(dispense_rows) != len(gilbarco_rows):
        raise ValueError("dispense_rows and gilbarco_rows must be the same length")

    wb = Workbook()
    ws = wb.active
    ws.title = OUTPUT_SHEET_NAME

    n_gilbarco = len(TRANSACTION_HEADERS)
    dispense_start_col = n_gilbarco + 2  # one blank separator column

    ws.cell(row=1, column=1, value=GILBARCO_SECTION_TITLE)
    ws.cell(row=1, column=dispense_start_col, value=ORIGINAL_SECTION_TITLE)

    for ci, header in enumerate(TRANSACTION_HEADERS, start=1):
        ws.cell(row=2, column=ci, value=header)
    for ci, header in enumerate(DISPENSE_HEADERS, start=dispense_start_col):
        ws.cell(row=2, column=ci, value=header)

    for ri, (dispense, gilbarco) in enumerate(zip(dispense_rows, gilbarco_rows), start=3):
        for ci, header in enumerate(TRANSACTION_HEADERS, start=1):
            ws.cell(row=ri, column=ci, value=gilbarco.get(header))
        for ci, header in enumerate(DISPENSE_HEADERS, start=dispense_start_col):
            ws.cell(row=ri, column=ci, value=dispense.get(header))

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    wb.close()


def _resolve_output_path(output_dir: Path, filename: str) -> Path:
    output_dir.mkdir(parents=True, exist_ok=True)
    candidate = output_dir / filename
    if not candidate.exists():
        return candidate
    stem = candidate.stem
    suffix = 1
    while True:
        alt = output_dir / f"{stem} ({suffix}).xlsx"
        if not alt.exists():
            return alt
        suffix += 1


def convert_workbook(
    dispense_path: str | Path,
    output_path: str | Path | None = None,
    *,
    output_dir: str | Path | None = None,
    template_path: str | Path | None = None,
    pump_config_path: str | Path | None = None,
    output_format: str = "gilbarco",
) -> dict[str, Any]:
    dispense_path = Path(dispense_path)
    template_path = Path(template_path) if template_path else DEFAULT_GILBARCO_TEMPLATE
    if not template_path.exists():
        raise FileNotFoundError(
            f"Gilbarco template not found: {template_path}. "
            "Expected bundled scripts/dispense-to-transactions/gilbarco-template.xlsx"
        )

    fmt = (output_format or "gilbarco").strip().lower()
    if fmt not in ("gilbarco", "winshuttle"):
        raise ValueError(f"Unsupported output_format: {output_format!r} (use gilbarco or winshuttle)")

    config = load_pump_config(pump_config_path)
    dispense_rows = parse_dispense_workbook(dispense_path)
    period = dispense_period_range(dispense_rows)
    output_file_name = (
        build_winshuttle_filename(period)
        if fmt == "winshuttle"
        else build_output_filename(period)
    )

    if output_path is None:
        base_dir = Path(output_dir) if output_dir else dispense_path.parent
        output_path = _resolve_output_path(base_dir, output_file_name)
    else:
        output_path = Path(output_path)

    fleet_by_id, owner_routes, bowser_by_fleet = load_reference_indexes(template_path)
    fuel_pump_routes = merge_fuel_pump_routes(
        build_fuel_pump_routes(template_path), config
    )
    gilbarco_rows, warnings = convert_dispense_rows(
        dispense_rows,
        fleet_by_id,
        owner_routes,
        bowser_by_fleet,
        config,
        fuel_pump_routes,
    )
    if fmt == "winshuttle":
        period_end = period[1] if period else None
        report_date = format_winshuttle_report_date(period_end)
        write_winshuttle_workbook(
            dispense_rows,
            gilbarco_rows,
            output_path,
            report_date=report_date,
            ws_cfg=winshuttle_config(config),
        )
    else:
        write_side_by_side_workbook(dispense_rows, gilbarco_rows, output_path)

    period_start, period_end = period if period else (None, None)
    unallocated = sum(
        1
        for r in dispense_rows
        if not _cell_str(r.get("Asset Number")) and not _cell_str(r.get("Asset Registration"))
    )
    return {
        "source": str(dispense_path.resolve()),
        "template": str(template_path.resolve()),
        "output": str(output_path.resolve()),
        "output_file_name": output_path.name,
        "output_format": fmt,
        "period_start": period_start,
        "period_end": period_end,
        "dispense_rows": len(dispense_rows),
        "gilbarco_rows": len(gilbarco_rows),
        "unallocated_rows": unallocated,
        "warnings": warnings,
    }
