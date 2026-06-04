"""Parse InsightWare / FuelTrack Fuel Dispense Report workbooks."""
from __future__ import annotations

import re
from datetime import date, datetime
from pathlib import Path
from typing import Any

import openpyxl

OUTPUT_SHEET_NAME = "Fuel Dispense Report"
GILBARCO_SECTION_TITLE = "Gilbarco (converted)"
ORIGINAL_SECTION_TITLE = "Original dispense (Sparrow)"

DISPENSE_HEADERS = (
    "Date & Time",
    "API ID",
    "Fuel Pump",
    "Asset Description",
    "Asset Number",
    "Asset Registration",
    "Asset Tag",
    "Group1",
    "Asset Group",
    "Asset Owner",
    "Gilbarco Unique Asset ID",
    "Internal Order Number",
    "Fuel Manager",
    "Operator",
    "User Tag",
    "Operation",
    "Location",
    "Override",
    "Pump Before",
    "Pump After",
    "Litres",
    "Odometer",
    "Economy",
    "Price/L",
    "Total",
)


def _cell_str(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _parse_number(value: Any) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    s = _cell_str(value).replace(",", "")
    if not s:
        return None
    try:
        return float(s)
    except ValueError:
        return None


def is_footer_row(row: dict[str, Any]) -> bool:
    pump_after = _cell_str(row.get("Pump After"))
    if pump_after.lower().startswith("total"):
        return True
    if row.get("Date & Time") is None and row.get("Asset Number") is None:
        litres = _parse_number(row.get("Litres"))
        if litres is None and not _cell_str(row.get("Fuel Pump")):
            return True
    return False


def is_bowser_fill(row: dict[str, Any]) -> bool:
    if row.get("Asset Number"):
        return False
    override = _cell_str(row.get("Override")).lower()
    if "bowser" in override:
        return True
    pump = _cell_str(row.get("Fuel Pump")).lower()
    if "fueltrack" in pump or "otk105" in pump:
        litres = _parse_number(row.get("Litres")) or 0
        return litres >= 50
    return False


def parse_dispense_workbook(path: str | Path) -> list[dict[str, Any]]:
    path = Path(path)
    if not path.exists():
        raise FileNotFoundError(path)

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheet_name = "Sheet1" if "Sheet1" in wb.sheetnames else wb.sheetnames[0]
    ws = wb[sheet_name]
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not header_row:
        wb.close()
        return []

    headers = [_cell_str(h) for h in header_row]
    if headers[0] != "Date & Time":
        raise ValueError(f"Unexpected dispense headers on {path.name}: {headers[:5]}")

    rows: list[dict[str, Any]] = []
    for cells in ws.iter_rows(min_row=2, values_only=True):
        if not cells or not any(c is not None and _cell_str(c) for c in cells):
            continue
        record = {
            headers[i]: cells[i] if i < len(cells) else None for i in range(len(headers))
        }
        if is_footer_row(record):
            continue
        rows.append(record)

    wb.close()
    return rows


def normalize_asset_owner(value: Any) -> str:
    s = _cell_str(value)
    if not s:
        return ""
    upper = s.upper()
    aliases = {
        "JANELAW ENTERPRISE": "JANELAW ENTERPRISE",
        "Janelaw Enterprise": "JANELAW ENTERPRISE",
    }
    return aliases.get(s, upper)


def normalize_asset_group(value: Any) -> str:
    s = _cell_str(value)
    if not s:
        return ""
    cleaned = re.sub(r"[\s-]+", " ", s).strip().lower()
    if cleaned in ("non eligible", "not eligible"):
        return "Not Eligible"
    if cleaned == "eligible":
        return "Eligible"
    return s


def split_asset_description(value: Any) -> tuple[str, str]:
    s = _cell_str(value)
    if not s:
        return "", ""
    if "-" in s:
        make, model = s.split("-", 1)
        return make.strip(), model.replace("-", " ").strip()
    return s, ""


def model_has_redundant_duplicate(model: Any) -> bool:
    """True when Model is like 'DOZER DOZER' or 'HILUX  HILUX' (repeated category token)."""
    parts = _cell_str(model).split()
    if len(parts) < 2:
        return False
    return parts[0].upper() == parts[1].upper()


def collapse_redundant_model(model: Any) -> str:
    """Collapse 'WORD WORD' to 'WORD'; normalize extra spaces."""
    s = re.sub(r"\s+", " ", _cell_str(model)).strip()
    if not s:
        return s
    parts = s.split()
    if len(parts) >= 2 and parts[0].upper() == parts[1].upper():
        if all(p.upper() == parts[0].upper() for p in parts):
            return parts[0]
        deduped: list[str] = []
        for part in parts:
            if not deduped or part.upper() != deduped[-1].upper():
                deduped.append(part)
        return " ".join(deduped)
    return s


def resolve_make_model(
    asset_description: Any,
    fleet_template: dict[str, Any] | None,
) -> tuple[str | None, str | None]:
    """
    Prefer Sparrow Asset Description when the reference template Model repeats one token
    (e.g. 'DOZER DOZER'); otherwise use template Make/Model.
    """
    if not fleet_template:
        make, model = split_asset_description(asset_description)
        return make or None, model or None

    make = _cell_str(fleet_template.get("Make"))
    model = _cell_str(fleet_template.get("Model"))

    if not model_has_redundant_duplicate(model):
        return make or None, model or None

    parsed_make, parsed_model = split_asset_description(asset_description)
    parsed_make = _cell_str(parsed_make)
    parsed_model = _cell_str(parsed_model)

    if parsed_model and not model_has_redundant_duplicate(parsed_model):
        return make or parsed_make or None, parsed_model

    collapsed = collapse_redundant_model(model)
    if parsed_make:
        return parsed_make, collapsed or parsed_make
    return make or None, collapsed or None


def parse_consumption(odometer: Any, economy: Any) -> tuple[str, str, float]:
    odo_s = _cell_str(odometer)
    econ_s = _cell_str(economy)
    reading = 0.0

    if odo_s:
        m = re.search(r"([\d,.]+)", odo_s)
        if m:
            try:
                reading = float(m.group(1).replace(",", ""))
            except ValueError:
                reading = 0.0

    if "km" in econ_s.lower() or ("km" in odo_s.lower() and "hr" not in odo_s.lower()):
        return "Odometer", "KM/L", reading
    if "hr" in odo_s.lower() or "l/hr" in econ_s.lower() or "l/h" in econ_s.lower():
        return "Hours", "L/HR", reading
    if odo_s:
        return "Hours", "L/HR", reading
    return "Hours", "L/HR", 0.0


def extract_pump_code(fuel_pump: Any, aliases: dict[str, str]) -> str:
    raw = _cell_str(fuel_pump)
    if not raw:
        return ""
    if raw in aliases:
        return aliases[raw]
    m = re.match(r"([A-Z0-9]+)", raw.upper())
    return m.group(1) if m else raw


def _parse_dispense_datetime(value: Any) -> datetime | None:
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime.combine(value, datetime.min.time())
    s = _cell_str(value)
    if not s:
        return None
    for fmt in (
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d %H:%M",
        "%d/%m/%Y %H:%M:%S",
        "%d/%m/%Y %H:%M",
    ):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    return None


def dispense_period_range(rows: list[dict[str, Any]]) -> tuple[str, str] | None:
    """Return (YYYYMMDD start, YYYYMMDD end) from dispense rows, or None."""
    dates: list[datetime] = []
    for row in rows:
        dt = _parse_dispense_datetime(row.get("Date & Time"))
        if dt:
            dates.append(dt)
    if not dates:
        return None
    start = min(dates).strftime("%Y%m%d")
    end = max(dates).strftime("%Y%m%d")
    return start, end


def build_output_filename(period: tuple[str, str] | None) -> str:
    if period:
        start, end = period
        return f"Fuel Dispense Report {start} - {end}.xlsx"
    return "Fuel Dispense Report.xlsx"
