"""WinShuttle SAP goods-movement upload format from Gilbarco-converted rows."""
from __future__ import annotations

from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

WINSHUTTLE_HEADERS = (
    "Number of External Material Slip",
    "Material Short Text with Field Label 'Material'",
    "Quantity in Unit of Entry",
    "Description of Storage Location",
    "Goods recipient",
    "Order Number",
    "Product",
    "Name",
)

WINSHUTTLE_SAP_FIELDS = (
    "GOHEAD-MTSNR",
    "GOITEM-MAKTX",
    "GOITEM-ERFMG",
    "GOITEM-LGOBE",
    "GOITEM-WEMPF",
    "COBL-AUFNR",
    None,
    "GOITEM-NAME1",
)

DEFAULT_WINSHUTTLE_CONFIG = {
    "material": "am0193746",
    "storage_location": "mf07",
    "goods_recipient": "mpho",
    "product": 261,
    "name": "tm01",
}

# Matches Book1.xlsx: accent1 (#156082) at Excel theme tint ~0.4
HEADER_FILL = PatternFill("solid", fgColor="729FB3")
HEADER_FONT = Font(bold=True, color="000000")
HEADER_FONT_PLAIN = Font(bold=False, color="000000")
DATE_FILL = PatternFill("solid", fgColor="00B050")
DATE_FONT = Font(bold=True, color="000000")


def _apply_header_style(cell, *, bold: bool = True) -> None:
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT if bold else HEADER_FONT_PLAIN


def _apply_date_style(cell) -> None:
    cell.fill = DATE_FILL
    cell.font = DATE_FONT


def winshuttle_config(config: dict[str, Any]) -> dict[str, Any]:
    """Merge pump_config winshuttle block with defaults."""
    raw = config.get("winshuttle") or {}
    merged = {**DEFAULT_WINSHUTTLE_CONFIG, **raw}
    return merged


def format_winshuttle_report_date(value: date | datetime | str | None) -> str | None:
    """Format as d/m/yyyy without zero-padding (e.g. 29/5/2026)."""
    if value is None:
        return None
    if isinstance(value, datetime):
        dt = value
    elif isinstance(value, date):
        dt = datetime.combine(value, datetime.min.time())
    elif isinstance(value, str) and len(value) == 8 and value.isdigit():
        dt = datetime.strptime(value, "%Y%m%d")
    else:
        return None
    return f"{dt.day}/{dt.month}/{dt.year}"


def resolve_winshuttle_order_number(
    gilbarco: dict[str, Any],
    dispense: dict[str, Any],
) -> Any:
    group5 = gilbarco.get("Group5")
    if group5 is not None and str(group5).strip() != "":
        return group5
    return dispense.get("Internal Order Number")


def convert_row_to_winshuttle(
    gilbarco: dict[str, Any],
    dispense: dict[str, Any],
    ws_cfg: dict[str, Any],
) -> tuple[Any, ...]:
    fleet_id = gilbarco.get("Fleet ID")
    litres = gilbarco.get("Liters")
    order = resolve_winshuttle_order_number(gilbarco, dispense)
    return (
        fleet_id,
        ws_cfg["material"],
        litres,
        ws_cfg["storage_location"],
        ws_cfg["goods_recipient"],
        order,
        ws_cfg["product"],
        ws_cfg["name"],
    )


def build_winshuttle_filename(period: tuple[str, str] | None) -> str:
    if period:
        start, end = period
        return f"WinShuttle Report {start} - {end}.xlsx"
    return "WinShuttle Report.xlsx"


def write_winshuttle_workbook(
    dispense_rows: list[dict[str, Any]],
    gilbarco_rows: list[dict[str, Any]],
    output_path: Path,
    *,
    report_date: str | None,
    ws_cfg: dict[str, Any],
) -> None:
    if len(dispense_rows) != len(gilbarco_rows):
        raise ValueError("dispense_rows and gilbarco_rows must be the same length")

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    for ci, header in enumerate(WINSHUTTLE_HEADERS, start=1):
        cell = ws.cell(row=1, column=ci, value=header)
        _apply_header_style(cell, bold=True)
    for ci, sap_field in enumerate(WINSHUTTLE_SAP_FIELDS, start=1):
        cell = ws.cell(row=2, column=ci, value=sap_field)
        _apply_header_style(cell, bold=sap_field is not None)

    if report_date:
        date_cell = ws.cell(row=3, column=2, value=report_date)
        _apply_date_style(date_cell)

    for ri, (dispense, gilbarco) in enumerate(
        zip(dispense_rows, gilbarco_rows), start=4
    ):
        for ci, value in enumerate(
            convert_row_to_winshuttle(gilbarco, dispense, ws_cfg), start=1
        ):
            ws.cell(row=ri, column=ci, value=value)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    wb.close()
