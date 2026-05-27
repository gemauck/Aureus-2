"""Write audit findings into a copy of the fuel refund workbook."""
from __future__ import annotations

import json
import shutil
from collections import defaultdict
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from rules import Finding, summarize_findings
from parse_workbook import ParsedWorkbook, _find_header_row

COMBINED_SHEET = "Combined Fuel Transactions"
AUDIT_HEADERS = [
    "Audit Result",
    "Audit Severity",
    "Findings Count",
    "Audit Comments",
    "Checks Failed",
]
SEVERITY_ORDER = {"error": 3, "warning": 2, "info": 1}

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF")
SEVERITY_FILLS = {
    "error": PatternFill("solid", fgColor="FFC7CE"),
    "warning": PatternFill("solid", fgColor="FFEB9C"),
    "info": PatternFill("solid", fgColor="DDEBF7"),
}
ROW_FILLS = {
    "error": PatternFill("solid", fgColor="E2D5F0"),
    "warning": PatternFill("solid", fgColor="EDE7F6"),
    "info": PatternFill("solid", fgColor="DDEBF7"),
}
PASS_AUDIT_FILL = PatternFill("solid", fgColor="E8F5E9")
AUDIT_COL_COUNT = len(AUDIT_HEADERS)


def row_severity(findings: list[Finding]) -> str | None:
    if not findings:
        return None
    return max(findings, key=lambda f: SEVERITY_ORDER.get(f.severity, 0)).severity


def audit_result_label(findings: list[Finding]) -> str:
    severity = row_severity(findings)
    if severity is None:
        return "Pass"
    if severity == "error":
        return "Error"
    return severity.capitalize()


def _combined_findings_by_row(findings: list[Finding]) -> dict[int, list[Finding]]:
    by_row: dict[int, list[Finding]] = defaultdict(list)
    for finding in findings:
        if finding.sheet == COMBINED_SHEET and finding.excel_row:
            by_row[finding.excel_row].append(finding)
    return by_row


def _annotate_combined_sheet(ws, findings: list[Finding], parsed: ParsedWorkbook) -> None:
    header_row = _find_header_row(ws) or 2
    ws.insert_cols(1, AUDIT_COL_COUNT)

    for col, title in enumerate(AUDIT_HEADERS, 1):
        cell = ws.cell(row=header_row, column=col, value=title)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT

    by_row = _combined_findings_by_row(findings)
    max_col = ws.max_column

    for row in parsed.combined_rows:
        excel_row = row.get("_excel_row")
        if not excel_row:
            continue
        row_findings = by_row.get(excel_row, [])
        severity = row_severity(row_findings)
        result = audit_result_label(row_findings)
        comments = "; ".join(f.message for f in row_findings if f.message)
        checks_failed = ", ".join(sorted({f.check_id for f in row_findings}))

        ws.cell(row=excel_row, column=1, value=result)
        ws.cell(row=excel_row, column=2, value=severity or "pass")
        ws.cell(row=excel_row, column=3, value=len(row_findings))
        ws.cell(row=excel_row, column=4, value=comments)
        ws.cell(row=excel_row, column=5, value=checks_failed)

        if severity is None:
            for col in range(1, AUDIT_COL_COUNT + 1):
                ws.cell(row=excel_row, column=col).fill = PASS_AUDIT_FILL
        else:
            row_fill = ROW_FILLS.get(severity)
            if row_fill:
                for col in range(1, max_col + 1):
                    ws.cell(row=excel_row, column=col).fill = row_fill

    for col in range(1, AUDIT_COL_COUNT + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18
    ws.column_dimensions["D"].width = 48

    ws.freeze_panes = "F3"


def write_audit_workbook(
    input_path: str | Path,
    output_path: str | Path,
    findings: list[Finding],
    parsed: ParsedWorkbook,
) -> None:
    input_path = Path(input_path)
    output_path = Path(output_path)
    shutil.copy2(input_path, output_path)

    wb = openpyxl.load_workbook(output_path)
    if "Audit Findings" in wb.sheetnames:
        del wb["Audit Findings"]
    if "Audit Summary" in wb.sheetnames:
        del wb["Audit Summary"]

    if COMBINED_SHEET in wb.sheetnames:
        _annotate_combined_sheet(wb[COMBINED_SHEET], findings, parsed)

    ws_findings = wb.create_sheet("Audit Findings", 0)
    finding_headers = [
        "Check ID",
        "Process Task",
        "Severity",
        "Sheet",
        "Excel Row",
        "Transaction ID",
        "Asset Number",
        "Message",
        "Fields (JSON)",
    ]
    for col, title in enumerate(finding_headers, 1):
        cell = ws_findings.cell(row=1, column=col, value=title)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT

    for ri, finding in enumerate(
        sorted(findings, key=lambda f: (f.severity != "error", f.check_id, f.excel_row or 0)),
        2,
    ):
        ws_findings.cell(row=ri, column=1, value=finding.check_id)
        ws_findings.cell(row=ri, column=2, value=finding.process_task or "")
        sev_cell = ws_findings.cell(row=ri, column=3, value=finding.severity)
        fill = SEVERITY_FILLS.get(finding.severity)
        if fill:
            for c in range(1, len(finding_headers) + 1):
                ws_findings.cell(row=ri, column=c).fill = fill
        ws_findings.cell(row=ri, column=4, value=finding.sheet)
        ws_findings.cell(row=ri, column=5, value=finding.excel_row)
        ws_findings.cell(row=ri, column=6, value=finding.transaction_id)
        ws_findings.cell(row=ri, column=7, value=finding.asset_number)
        ws_findings.cell(row=ri, column=8, value=finding.message)
        ws_findings.cell(row=ri, column=9, value=json.dumps(finding.fields, default=str))

    for col in range(1, len(finding_headers) + 1):
        ws_findings.column_dimensions[get_column_letter(col)].width = 18
    ws_findings.column_dimensions["H"].width = 60
    ws_findings.column_dimensions["I"].width = 40

    summary = summarize_findings(findings, parsed)
    ws_summary = wb.create_sheet("Audit Summary", 1)
    ws_summary["A1"] = "Fuel Refund Report Audit Summary"
    ws_summary["A1"].font = Font(bold=True, size=14)
    row = 3
    ws_summary.cell(row=row, column=1, value="Combined transaction rows")
    ws_summary.cell(row=row, column=2, value=summary["row_count_combined"])
    row += 1
    ws_summary.cell(row=row, column=1, value="Rows audited")
    ws_summary.cell(row=row, column=2, value=summary["rows_audited"])
    row += 1
    ws_summary.cell(row=row, column=1, value="Rows passed")
    ws_summary.cell(row=row, column=2, value=summary["rows_passed"])
    row += 1
    ws_summary.cell(row=row, column=1, value="Rows failed (error)")
    ws_summary.cell(row=row, column=2, value=summary["rows_failed"])
    row += 1
    ws_summary.cell(row=row, column=1, value="Rows warning only")
    ws_summary.cell(row=row, column=2, value=summary["rows_warning_only"])
    row += 1
    ws_summary.cell(row=row, column=1, value="Pass rate (%)")
    ws_summary.cell(row=row, column=2, value=summary["pass_rate_pct"])
    row += 1
    ws_summary.cell(row=row, column=1, value="Total findings")
    ws_summary.cell(row=row, column=2, value=summary["finding_count"])
    row += 2

    ws_summary.cell(row=row, column=1, value="By severity").font = Font(bold=True)
    row += 1
    for sev, count in summary.get("by_severity", {}).items():
        ws_summary.cell(row=row, column=1, value=sev)
        ws_summary.cell(row=row, column=2, value=count)
        row += 1
    row += 1

    ws_summary.cell(row=row, column=1, value="Checks passed").font = Font(bold=True)
    row += 1
    for check_id in summary.get("checks_passed", []):
        ws_summary.cell(row=row, column=1, value=check_id)
        row += 1
    row += 1

    ws_summary.cell(row=row, column=1, value="Checks failed").font = Font(bold=True)
    row += 1
    for item in summary.get("checks_failed", []):
        ws_summary.cell(row=row, column=1, value=item.get("check_id"))
        ws_summary.cell(row=row, column=2, value=item.get("count"))
        ws_summary.cell(row=row, column=3, value=item.get("process_task") or "")
        row += 1
    row += 1

    ws_summary.cell(row=row, column=1, value="By check").font = Font(bold=True)
    row += 1
    for check_id, count in summary.get("by_check", {}).items():
        ws_summary.cell(row=row, column=1, value=check_id)
        ws_summary.cell(row=row, column=2, value=count)
        row += 1
    row += 1

    ws_summary.cell(row=row, column=1, value="Refund rate(s)").font = Font(bold=True)
    row += 1
    for rate in parsed.refund_rates:
        ws_summary.cell(row=row, column=1, value=rate.get("label"))
        ws_summary.cell(row=row, column=2, value=rate.get("rate"))
        row += 1
    row += 1

    ws_summary.cell(row=row, column=1, value="Manual follow-up (not automated)").font = Font(bold=True)
    row += 1
    manual = [
        "Final formatting and VAT 201 insert",
        "Eligible Review tab assembly for final submission",
        "Stock adjustments and corrective deductions",
        "Asset rename pass for auto-created assets",
        "Compliance sign-off",
    ]
    for item in manual:
        ws_summary.cell(row=row, column=1, value=f"• {item}")
        row += 1

    if parsed.parse_warnings:
        row += 1
        ws_summary.cell(row=row, column=1, value="Parse warnings").font = Font(bold=True)
        row += 1
        for w in parsed.parse_warnings:
            ws_summary.cell(row=row, column=1, value=w)
            row += 1

    ws_summary.column_dimensions["A"].width = 42
    ws_summary.column_dimensions["B"].width = 24
    ws_summary.column_dimensions["C"].width = 36

    wb.save(output_path)
    wb.close()


def build_summary_json(findings: list[Finding], parsed: ParsedWorkbook) -> dict[str, Any]:
    base = summarize_findings(findings, parsed)
    base["findings"] = [f.to_dict() for f in findings]
    return base
