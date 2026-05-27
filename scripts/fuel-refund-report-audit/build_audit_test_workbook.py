#!/usr/bin/env python3
"""Inject labelled AUDIT-TEST violations into a DFRR workbook for manual QA upload."""
from __future__ import annotations

import argparse
import shutil
from datetime import datetime
from pathlib import Path

import openpyxl

GUIDE_SHEET = "Audit Test Guide"
COMBINED = "Combined Fuel Transactions"
RECEIPTS = "Fuel Receipts"
TANK_SUMMARY = "Combined Tank Summary"
ELIGIBLE_REVIEW = "Eligible Review - Dispenses"
ASSET_SHEET = "738-P1-DB03"

BASE_DT = datetime(2026, 4, 29, 10, 0, 0)


def _header_map(ws, header_row: int = 2) -> dict[str, int]:
    row = next(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True))
    return {str(h).strip(): i + 1 for i, h in enumerate(row) if h is not None and str(h).strip()}


def _set_row(ws, excel_row: int, col_map: dict[str, int], values: dict) -> None:
    for key, val in values.items():
        if key not in col_map:
            continue
        ws.cell(row=excel_row, column=col_map[key], value=val)


def _blank_combined_row() -> dict:
    return {
        "Transaction Type": None,
        "Date & Time": None,
        "Transaction ID": None,
        "Fuel Delivery Ref. No.": None,
        "Asset Description": None,
        "Asset Registration": None,
        "Asset Number": None,
        "Asset Tag": None,
        "Asset Group": None,
        "Asset Tank Size (L)": None,
        "Asset Meter Type (Hr/Km)": None,
        "Storage Tank": None,
        "Fuel Pump": None,
        "Fuel Dispensed or Received (L)": None,
        "≈ Tank Litres Before": None,
        "≈ Tank Litres After": None,
        "Pump Readings Before": None,
        "Pump Readings After": None,
        "Opening Odo": None,
        "Closing Odo": None,
        "Total Usage Km/Hr": None,
        "Total Fuel Used (L)": None,
        "Consumption": None,
        "Operation Description / Comment": None,
        "Refund Eligibility": None,
        "Eligible L": None,
        "Non-Eligible L": None,
        "Litres Transferred": None,
        "Operator": None,
        "Location": None,
        "Override": None,
        "Fuel Cost (R)": None,
        "Eligible Volume (L) (Claimable % of Total)": None,
        "Refund Price": None,
        "Refund Total": None,
    }


def _mining_base(**overrides) -> dict:
    row = _blank_combined_row()
    row.update(
        {
            "Transaction Type": "DISPENSE",
            "Date & Time": BASE_DT,
            "Transaction ID": "AUDIT-TEST-BASE",
            "Asset Description": "AUDIT TEST EXCAVATOR",
            "Asset Registration": "AUDIT-001",
            "Asset Number": "AUDIT-EX-001",
            "Asset Tag": "AUDIT-EX-001",
            "Asset Group": "Mining - Eligible",
            "Asset Tank Size (L)": 500,
            "Asset Meter Type (Hr/Km)": "hr",
            "Storage Tank": "HDV Bay",
            "Fuel Pump": "HDV P1",
            "Fuel Dispensed or Received (L)": -120.0,
            "Pump Readings Before": 1000,
            "Pump Readings After": 1120,
            "≈ Tank Litres Before": 5000,
            "≈ Tank Litres After": 4880,
            "Opening Odo": "100 hr",
            "Closing Odo": "110 hr",
            "Total Usage Km/Hr": "10.0 hr",
            "Total Fuel Used (L)": 120.0,
            "Consumption": 12.0,
            "Operation Description / Comment": "Audit test — normal mining dispense",
            "Refund Eligibility": "Eligible",
            "Eligible L": 100.0,
            "Operator": "Test Operator",
            "Location": "Pit 1",
            "Eligible Volume (L) (Claimable % of Total)": 100.0,
            "Refund Price": 3.66,
            "Refund Total": 366.0,
        }
    )
    row.update(overrides)
    return row


def _combined_test_rows() -> list[dict]:
    dt = lambda m: datetime(2026, 4, 29, 10, m, 0)
    rows: list[dict] = []

    rows.append(
        _mining_base(
            **{
                "Transaction Type": "INITIAL-DISPENSE",
                "Transaction ID": "AUDIT-TEST-01-initial-dispense-claim",
                "Fuel Dispensed or Received (L)": -50.0,
                "Refund Eligibility": "Non-Eligible",
                "Eligible L": 0,
                "Eligible Volume (L) (Claimable % of Total)": 50.0,
                "Refund Total": 183.0,
                "Operation Description / Comment": "Initial fill — must not have claim",
            }
        )
    )

    dup = _mining_base(
        **{
            "Transaction ID": "AUDIT-TEST-02-duplicate-A",
            "Date & Time": dt(5),
            "Asset Number": "AUDIT-DUP-001",
            "Fuel Dispensed or Received (L)": -80.0,
            "Total Fuel Used (L)": 80.0,
            "Eligible L": 80.0,
            "Eligible Volume (L) (Claimable % of Total)": 80.0,
            "Refund Total": 292.8,
        }
    )
    rows.append(dup)
    rows.append(
        {
            **dup,
            "Transaction ID": "AUDIT-TEST-02-duplicate-B",
        }
    )

    rows.append(
        _mining_base(
            **{
                "Transaction ID": "AUDIT-TEST-03-missing-claim",
                "Date & Time": dt(10),
                "Asset Number": "AUDIT-CLAIM-001",
                "Fuel Dispensed or Received (L)": -90.0,
                "Total Fuel Used (L)": 90.0,
                "Eligible L": 0,
                "Eligible Volume (L) (Claimable % of Total)": 0,
                "Refund Total": 0,
                "Refund Price": None,
            }
        )
    )

    rows.append(
        _mining_base(
            **{
                "Transaction ID": "AUDIT-TEST-04-missing-operator",
                "Date & Time": dt(12),
                "Asset Number": "AUDIT-OP-001",
                "Operator": None,
            }
        )
    )

    rows.append(
        _mining_base(
            **{
                "Transaction ID": "AUDIT-TEST-05-missing-location",
                "Date & Time": dt(14),
                "Asset Number": "AUDIT-LOC-001",
                "Location": None,
            }
        )
    )

    rows.append(
        _mining_base(
            **{
                "Transaction ID": "AUDIT-TEST-06-circular-bowser",
                "Date & Time": dt(16),
                "Asset Number": "LPD-AUDIT",
                "Asset Description": "MOBILE-BOWSER AUDIT TEST TANKER",
                "Storage Tank": "LPD-AUDIT",
                "Fuel Dispensed or Received (L)": -200.0,
                "Total Fuel Used (L)": 200.0,
            }
        )
    )

    rows.append(
        _mining_base(
            **{
                "Transaction ID": "AUDIT-TEST-07-exceeds-tank",
                "Date & Time": dt(18),
                "Asset Number": "AUDIT-TANK-001",
                "Asset Tank Size (L)": 100,
                "Fuel Dispensed or Received (L)": -250.0,
                "Total Fuel Used (L)": 250.0,
            }
        )
    )

    rows.append(
        _mining_base(
            **{
                "Transaction ID": "AUDIT-TEST-08-consecutive-1",
                "Date & Time": dt(20),
                "Asset Number": "AUDIT-CONSEC-001",
                "Asset Tank Size (L)": 100,
                "Fuel Dispensed or Received (L)": -60.0,
                "Total Fuel Used (L)": 60.0,
                "Eligible L": 60.0,
                "Eligible Volume (L) (Claimable % of Total)": 60.0,
                "Refund Total": 219.6,
            }
        )
    )
    rows.append(
        _mining_base(
            **{
                "Transaction ID": "AUDIT-TEST-08-consecutive-2",
                "Date & Time": dt(20),
                "Asset Number": "AUDIT-CONSEC-001",
                "Asset Tank Size (L)": 100,
                "Fuel Dispensed or Received (L)": -55.0,
                "Total Fuel Used (L)": 55.0,
                "Eligible L": 55.0,
                "Eligible Volume (L) (Claimable % of Total)": 55.0,
                "Refund Total": 201.3,
            }
        )
    )

    rows.append(
        _mining_base(
            **{
                "Transaction ID": "AUDIT-TEST-09-missing-pump",
                "Date & Time": dt(22),
                "Asset Number": "AUDIT-PUMP-001",
                "Pump Readings Before": None,
                "Pump Readings After": None,
            }
        )
    )

    rows.append(
        _mining_base(
            **{
                "Transaction ID": "AUDIT-TEST-10-missing-tank-litres",
                "Date & Time": dt(24),
                "Asset Number": "AUDIT-TANKREAD-001",
                "≈ Tank Litres Before": None,
                "≈ Tank Litres After": None,
            }
        )
    )

    rows.append(
        _mining_base(
            **{
                "Transaction ID": "AUDIT-TEST-11-negative-odo",
                "Date & Time": dt(26),
                "Asset Number": "AUDIT-ODO-NEG",
                "Eligible L": 100.0,
                "Eligible Volume (L) (Claimable % of Total)": 100.0,
                "Refund Total": 366.0,
                "Refund Price": 3.66,
                "Total Usage Km/Hr": "-27580.0 hr",
                "Total Fuel Used (L)": 150.0,
                "Fuel Dispensed or Received (L)": -150.0,
            }
        )
    )

    rows.append(
        _mining_base(
            **{
                "Transaction ID": "AUDIT-TEST-12-high-odo-hr",
                "Date & Time": dt(28),
                "Asset Number": "AUDIT-ODO-HIGH",
                "Total Usage Km/Hr": "999.0 hr",
                "Total Fuel Used (L)": 50.0,
                "Fuel Dispensed or Received (L)": -50.0,
            }
        )
    )

    rows.append(
        _mining_base(
            **{
                "Transaction ID": "AUDIT-TEST-13-high-odo-km",
                "Date & Time": dt(30),
                "Asset Number": "AUDIT-ODO-KM",
                "Asset Meter Type (Hr/Km)": "km",
                "Total Usage Km/Hr": "800 km",
                "Total Fuel Used (L)": 40.0,
                "Fuel Dispensed or Received (L)": -40.0,
            }
        )
    )

    rows.append(
        _mining_base(
            **{
                "Transaction ID": "AUDIT-TEST-14-unrealistic-consumption",
                "Date & Time": dt(32),
                "Asset Number": "AUDIT-CONS-001",
                "Consumption": 350.0,
                "Total Usage Km/Hr": "1.0 hr",
                "Total Fuel Used (L)": 350.0,
                "Fuel Dispensed or Received (L)": -350.0,
                "Eligible L": 350.0,
                "Eligible Volume (L) (Claimable % of Total)": 350.0,
                "Refund Total": 1281.0,
            }
        )
    )

    rows.append(
        _mining_base(
            **{
                "Transaction ID": "AUDIT-TEST-15-missing-op-desc",
                "Date & Time": dt(34),
                "Asset Number": "AUDIT-OPDESC-001",
                "Operation Description / Comment": None,
            }
        )
    )

    rows.append(
        _mining_base(
            **{
                "Transaction ID": "AUDIT-TEST-16-wrong-refund-rate",
                "Date & Time": dt(36),
                "Asset Number": "AUDIT-RATE-001",
                "Refund Price": 9.99,
                "Refund Total": 999.0,
                "Eligible Volume (L) (Claimable % of Total)": 100.0,
            }
        )
    )

    rows.append(
        _mining_base(
            **{
                "Transaction ID": "AUDIT-TEST-17-refund-math",
                "Date & Time": dt(38),
                "Asset Number": "AUDIT-MATH-001",
                "Eligible Volume (L) (Claimable % of Total)": 100.0,
                "Refund Price": 3.66,
                "Refund Total": 1.0,
            }
        )
    )

    rows.append(
        _mining_base(
            **{
                "Transaction Type": "FUEL-RECEIPT",
                "Transaction ID": "AUDIT-TEST-18-receipt-no-cost",
                "Date & Time": dt(40),
                "Asset Number": None,
                "Asset Group": None,
                "Fuel Dispensed or Received (L)": 5000.0,
                "Fuel Cost (R)": None,
                "Refund Eligibility": None,
                "Eligible L": None,
                "Operator": None,
                "Location": None,
            }
        )
    )

    rows.append(
        _mining_base(
            **{
                "Transaction ID": "AUDIT-TEST-19-auto-asset",
                "Date & Time": dt(42),
                "Asset Number": "AUTO-TEST-UNIT-001",
                "Asset Description": "UNKNOWN ASSET placeholder",
            }
        )
    )

    rows.append(
        _mining_base(
            **{
                "Transaction ID": "AUDIT-TEST-20-bowser-low-litre",
                "Date & Time": dt(44),
                "Asset Number": "LPD-LOW",
                "Asset Description": "MOBILE-BOWSER LOW LITRE TEST",
                "Storage Tank": "HDV Bay",
                "Fuel Dispensed or Received (L)": -15.0,
                "Total Fuel Used (L)": 15.0,
                "Eligible L": 0,
                "Eligible Volume (L) (Claimable % of Total)": 0,
                "Refund Total": 0,
                "Refund Eligibility": "Non-Eligible",
            }
        )
    )

    return rows


def _write_guide_sheet(wb) -> None:
    if GUIDE_SHEET in wb.sheetnames:
        del wb[GUIDE_SHEET]
    ws = wb.create_sheet(GUIDE_SHEET, 0)
    lines = [
        ["Fuel Refund Report Audit — test workbook"],
        [""],
        ["Upload this file in Tools → Fuel Refund Report Audit."],
        ["Enable ALL options for full coverage of optional checks:"],
        ["(Pump/tank/consumption will also flag thousands of real historical rows — filter by AUDIT-TEST)"],
        ["  ☑ Require pump readings"],
        ["  ☑ Require tank readings"],
        ["  ☑ Consumption assessment"],
        [""],
        ["Search Combined Fuel Transactions for Transaction ID starting with AUDIT-TEST-"],
        [""],
        ["Expected checks (one or more rows each):"],
        ["01  initial_dispense_no_claim", "AUDIT-TEST-01-..."],
        ["02  duplicate_transaction", "AUDIT-TEST-02-duplicate-A and B"],
        ["03  mining_eligible_missing_claim", "AUDIT-TEST-03-..."],
        ["04  mining_eligible_missing_operator", "AUDIT-TEST-04-..."],
        ["05  mining_eligible_missing_location", "AUDIT-TEST-05-..."],
        ["06  circular_storage_tank", "AUDIT-TEST-06-..."],
        ["07  dispense_exceeds_tank_size", "AUDIT-TEST-07-..."],
        ["08  consecutive_hour_exceeds_tank", "AUDIT-TEST-08-consecutive-1/2"],
        ["09  missing_pump_readings", "AUDIT-TEST-09-... (checkbox)"],
        ["10  missing_tank_readings", "AUDIT-TEST-10-... (checkbox)"],
        ["12  negative_odo_eligible", "AUDIT-TEST-11-..."],
        ["13  high_odo_eligible", "AUDIT-TEST-12- and 13-..."],
        ["14  unrealistic_consumption", "AUDIT-TEST-14-... (checkbox)"],
        ["15  mining_eligible_missing_operation_desc", "AUDIT-TEST-15-..."],
        ["16  refund_rate_summary", "AUDIT-TEST-16-..."],
        ["17  refund_total_math", "AUDIT-TEST-17-..."],
        ["18  receipt_missing_fuel_cost", "AUDIT-TEST-18- + Fuel Receipts rows"],
        ["19  receipt_duplicate", "Fuel Receipts AUDIT-REC-DUP-A/B"],
        ["20  tank_summary_imbalance", "738-P1-DB03 eligible volume on summary"],
        ["21  auto_created_asset_suspect", "AUDIT-TEST-19-..."],
        ["22  eligible_review_unmarked_consecutive", "Eligible Review rows (2 findings)"],
        ["23  bowser_low_litre", "AUDIT-TEST-20-..."],
    ]
    for ri, line in enumerate(lines, 1):
        ws.cell(row=ri, column=1, value=line[0] if len(line) == 1 else line[0])
        if len(line) > 1:
            ws.cell(row=ri, column=2, value=line[1])


def build(input_path: Path, output_path: Path) -> None:
    shutil.copy2(input_path, output_path)
    wb = openpyxl.load_workbook(output_path)

    ws = wb[COMBINED]
    col = _header_map(ws, 2)
    start_row = ws.max_row + 1
    for offset, row_data in enumerate(_combined_test_rows()):
        _set_row(ws, start_row + offset, col, row_data)

    # Fuel receipts: duplicate + missing cost
    if RECEIPTS in wb.sheetnames:
        fr = wb[RECEIPTS]
        fr_col = _header_map(fr, 2)
        r0 = fr.max_row + 1
        rec_dt = datetime(2026, 4, 29, 11, 0, 0)
        for i, eid in enumerate(("AUDIT-REC-DUP-A", "AUDIT-REC-DUP-B")):
            _set_row(
                fr,
                r0 + i,
                fr_col,
                {
                    "Date & Time": rec_dt,
                    "Event ID": eid,
                    "Fuel Pump": "HDV P1",
                    "Litres Received": 1000.0,
                    "Order Ref": "AUDIT-ORDER-1",
                    "Price/L": None,
                    "Fuel Cost (R)": None,
                },
            )
        _set_row(
            fr,
            r0 + 2,
            fr_col,
            {
                "Date & Time": datetime(2026, 4, 29, 11, 5, 0),
                "Event ID": "AUDIT-REC-NO-PRICE",
                "Fuel Pump": "HDV P1",
                "Litres Received": 500.0,
                "Order Ref": "AUDIT-ORDER-2",
                "Price/L": None,
                "Fuel Cost (R)": 0,
            },
        )

    # Tank summary imbalance: Eligible Volume > 0 but Refund Total = 0 (usage table section, row ~14)
    if TANK_SUMMARY in wb.sheetnames:
        ts = wb[TANK_SUMMARY]
        eligible_header_row = None
        for ri in range(1, 40):
            if (
                ts.cell(ri, 1).value == "Tank"
                and ts.cell(ri, 2).value
                and "Eligible Usage" in str(ts.cell(ri, 2).value)
            ):
                eligible_header_row = ri
                break
        if eligible_header_row:
            for ri in range(eligible_header_row + 1, eligible_header_row + 10):
                if ts.cell(ri, 1).value == "738-P1-DB03":
                    ts.cell(ri, 4, value=500.0)
                    ts.cell(ri, 5, value=0.0)
                    break

    # Eligible Review - Dispenses
    if ELIGIBLE_REVIEW in wb.sheetnames:
        del wb[ELIGIBLE_REVIEW]
    er = wb.create_sheet(ELIGIBLE_REVIEW)
    er_headers = [
        "Transaction Type",
        "Date & Time",
        "Transaction ID",
        "Asset Number",
        "Exception Reason",
        "Fuel Dispensed or Received (L)",
    ]
    for ci, h in enumerate(er_headers, 1):
        er.cell(row=1, column=ci, value=h)
    er.cell(row=2, column=1, value="DISPENSE")
    er.cell(row=2, column=2, value=datetime(2026, 4, 29, 12, 0, 0))
    er.cell(row=2, column=3, value="137-AUDIT-CONSEC-MISSING")
    er.cell(row=2, column=4, value="INK11")
    er.cell(row=2, column=5, value=None)
    er.cell(row=2, column=6, value=-50.0)
    er.cell(row=3, column=1, value="DISPENSE")
    er.cell(row=3, column=2, value=datetime(2026, 4, 29, 12, 5, 0))
    er.cell(row=3, column=3, value="137-AUDIT-SEQ-OK")
    er.cell(row=3, column=4, value="INK11")
    er.cell(row=3, column=5, value="Consecutive dispense within 1 hour")
    er.cell(row=3, column=6, value=-30.0)

    # Asset sheet: missing tank readings row + row with tank litres for final check
    if ASSET_SHEET in wb.sheetnames:
        aws = wb[ASSET_SHEET]
        acol = _header_map(aws, 2)
        ar = aws.max_row + 1
        _set_row(
            aws,
            ar,
            acol,
            _mining_base(
                **{
                    "Transaction ID": "AUDIT-TEST-10b-missing-tank-asset",
                    "Date & Time": datetime(2026, 4, 29, 10, 25, 0),
                    "Asset Number": "AUDIT-TANKREAD-ASSET",
                    "≈ Tank Litres Before": None,
                    "≈ Tank Litres After": None,
                }
            ),
        )
        _set_row(
            aws,
            ar + 1,
            acol,
            _mining_base(
                **{
                    "Transaction ID": "AUDIT-TEST-21-final-tank-litres-present",
                    "Date & Time": datetime(2026, 4, 29, 10, 50, 0),
                    "Asset Number": "AUDIT-FINAL-TANK",
                    "≈ Tank Litres Before": 9999,
                    "≈ Tank Litres After": 9900,
                }
            ),
        )

    _write_guide_sheet(wb)
    wb.save(output_path)
    wb.close()


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--input", "-i", required=True, type=Path)
    parser.add_argument("--output", "-o", required=True, type=Path)
    args = parser.parse_args()
    if not args.input.exists():
        raise SystemExit(f"Input not found: {args.input}")
    args.output.parent.mkdir(parents=True, exist_ok=True)
    build(args.input, args.output)
    print(f"Wrote audit test workbook: {args.output}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
