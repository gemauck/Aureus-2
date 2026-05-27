"""Normalize DFRR workbooks before audit (strip prior audit columns, remove audit sheets)."""
from __future__ import annotations

from pathlib import Path
import shutil

import openpyxl

from parse_workbook import _find_header_row, _cell_str

AUDIT_COLUMN_HEADERS = (
    "Audit Result",
    "Audit Severity",
    "Findings Count",
    "Audit Comments",
    "Checks Failed",
)
AUDIT_SHEETS = frozenset({"Audit Findings", "Audit Summary"})
AUDIT_COL_COUNT = len(AUDIT_COLUMN_HEADERS)


def _header_has_audit_prefix(ws, header_row: int) -> bool:
    row = next(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True), None)
    if not row:
        return False
    headers = [_cell_str(c) for c in row[:AUDIT_COL_COUNT]]
    return headers == list(AUDIT_COLUMN_HEADERS)


def normalize_workbook(wb: openpyxl.Workbook) -> list[str]:
    """Strip duplicate audit columns and remove audit output sheets. Returns warning messages."""
    warnings: list[str] = []

    for sheet_name in list(wb.sheetnames):
        if sheet_name in AUDIT_SHEETS:
            del wb[sheet_name]
            warnings.append(f"Removed prior sheet: {sheet_name}")

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        hr = _find_header_row(ws)
        if hr and _header_has_audit_prefix(ws, hr):
            ws.delete_cols(1, AUDIT_COL_COUNT)
            warnings.append(f"Removed prior audit columns A–E from sheet: {sheet_name}")

    return warnings


def prepare_output_workbook(input_path: str | Path, output_path: str | Path) -> list[str]:
    """Copy input to output and normalize for a fresh audit run."""
    input_path = Path(input_path)
    output_path = Path(output_path)
    shutil.copy2(input_path, output_path)
    wb = openpyxl.load_workbook(output_path)
    warnings = normalize_workbook(wb)
    wb.save(output_path)
    wb.close()
    return warnings
