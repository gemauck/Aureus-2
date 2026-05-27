"""Parse InsightWare / Exxaro Detailed Fuel Refund Report workbooks."""
from __future__ import annotations

import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

import openpyxl

SYSTEM_SHEETS = frozenset(
    {
        "Combined Tank Summary",
        "VAT 201 Figures",
        "Combined Fuel Transactions",
        "Tank Transfers",
        "Fuel Receipts",
        "Tank Configuration",
        "Transactions by Category",
        "Eligible Review - Dispenses",
        "Eligible Review - Operations",
        "Eligible Review - Bowsers",
        "Stock Adjustments",
        "Audit Findings",
        "Audit Summary",
    }
)

BANNER_MARKERS = ("Transaction Type", "Date & Time")

AUDIT_COLUMN_HEADERS = (
    "Audit Result",
    "Audit Severity",
    "Findings Count",
    "Audit Comments",
    "Checks Failed",
)


def _strip_leading_audit_headers(headers: list[str]) -> list[str]:
    """Drop prior audit columns A–E when re-parsing an annotated workbook."""
    if len(headers) >= len(AUDIT_COLUMN_HEADERS) and headers[: len(AUDIT_COLUMN_HEADERS)] == list(
        AUDIT_COLUMN_HEADERS
    ):
        return headers[len(AUDIT_COLUMN_HEADERS) :]
    return headers


def _cell_str(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _find_header_row(ws, max_scan: int = 12) -> int | None:
    for ri in range(1, max_scan + 1):
        row = next(ws.iter_rows(min_row=ri, max_row=ri, values_only=True), None)
        if not row:
            continue
        cells = [_cell_str(c) for c in row if c is not None]
        if "Transaction Type" in cells:
            return ri
        first = _cell_str(row[0])
        if first == "Transaction Type":
            return ri
        if first == "Date & Time" and "Fuel Pump" in cells:
            return ri
    return None


def _rows_from_sheet(ws, header_row: int) -> tuple[list[str], list[dict]]:
    header_cells = next(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True), ())
    headers_full = [_cell_str(h) for h in header_cells]
    headers = _strip_leading_audit_headers(headers_full)
    col_offset = len(headers_full) - len(headers)
    rows: list[dict] = []
    for excel_row, cells in enumerate(
        ws.iter_rows(min_row=header_row + 1, values_only=True),
        start=header_row + 1,
    ):
        if not cells or not any(c is not None and str(c).strip() != "" for c in cells):
            continue
        data_cells = cells[col_offset:] if col_offset else cells
        first = _cell_str(data_cells[0]) if data_cells else ""
        tx_col = headers.index("Transaction Type") if "Transaction Type" in headers else 0
        tx_val = _cell_str(data_cells[tx_col]) if tx_col < len(data_cells) else first
        if first in ("Totals:", "Total") or tx_val in ("Totals:", "Total"):
            continue
        if tx_val in ("Audit Result", "Transaction Type"):
            continue
        record = {
            headers[i]: data_cells[i] if i < len(data_cells) else None
            for i in range(len(headers))
            if headers[i]
        }
        record["_excel_row"] = excel_row
        record["_sheet"] = ws.title
        rows.append(record)
    return headers, rows


def _parse_fuel_receipts(ws) -> list[dict]:
    header_row = _find_header_row(ws) or 2
    _, rows = _rows_from_sheet(ws, header_row)
    return rows


def _parse_eligible_review(ws) -> list[dict]:
    header_row = 1
    row1 = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if row1 and _cell_str(row1[0]) != "Transaction Type":
        header_row = _find_header_row(ws) or 1
    _, rows = _rows_from_sheet(ws, header_row)
    return rows


def _parse_refund_rates(ws) -> list[dict]:
    rates: list[dict] = []
    for ri, row in enumerate(ws.iter_rows(min_row=1, max_row=40, values_only=True), 1):
        if not row:
            continue
        label = _cell_str(row[0])
        if label.lower().startswith("rate from") and len(row) > 1:
            try:
                val = float(row[1])
            except (TypeError, ValueError):
                continue
            rates.append({"label": label, "rate": val, "excel_row": ri})
    return rates


def _parse_tank_summary(ws) -> dict[str, Any]:
    summary: dict[str, Any] = {"rates": _parse_refund_rates(ws), "tanks": []}
    in_usage = False
    for ri, row in enumerate(ws.iter_rows(min_row=1, max_row=80, values_only=True), 1):
        if not row:
            continue
        label = _cell_str(row[0])
        if label == "Tank" and _cell_str(row[1]) == "Eligible Usage (L)":
            in_usage = True
            continue
        if in_usage and label and label != "Total":
            try:
                summary["tanks"].append(
                    {
                        "tank": label,
                        "eligible_usage": float(row[1] or 0),
                        "non_eligible_usage": float(row[2] or 0),
                        "eligible_volume": float(row[3] or 0),
                        "refund_total": float(row[4] or 0),
                        "excel_row": ri,
                    }
                )
            except (TypeError, ValueError):
                pass
        if label == "Total" and in_usage:
            break
    return summary


def _discover_asset_sheets(wb) -> list[str]:
    asset_sheets: list[str] = []
    for name in wb.sheetnames:
        if name in SYSTEM_SHEETS:
            continue
        ws = wb[name]
        hr = _find_header_row(ws)
        if hr:
            asset_sheets.append(name)
    return asset_sheets


@dataclass
class ParsedWorkbook:
    source_path: str
    sheet_names: list[str] = field(default_factory=list)
    combined_rows: list[dict] = field(default_factory=list)
    combined_headers: list[str] = field(default_factory=list)
    fuel_receipts: list[dict] = field(default_factory=list)
    eligible_review_dispenses: list[dict] = field(default_factory=list)
    asset_sheets: dict[str, list[dict]] = field(default_factory=dict)
    refund_rates: list[dict] = field(default_factory=list)
    tank_summary: dict[str, Any] = field(default_factory=dict)
    parse_warnings: list[str] = field(default_factory=list)


def parse_workbook(path: str | Path) -> ParsedWorkbook:
    path = Path(path)
    if not path.exists():
        raise FileNotFoundError(path)

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    parsed = ParsedWorkbook(source_path=str(path.resolve()), sheet_names=list(wb.sheetnames))

    if "Combined Fuel Transactions" not in wb.sheetnames:
        parsed.parse_warnings.append("Missing sheet: Combined Fuel Transactions")
    else:
        ws = wb["Combined Fuel Transactions"]
        hr = _find_header_row(ws)
        if not hr:
            raise ValueError("Could not find header row on Combined Fuel Transactions")
        if hr != 2:
            parsed.parse_warnings.append(
                f"Combined Fuel Transactions header on row {hr} (expected row 2)"
            )
        parsed.combined_headers, parsed.combined_rows = _rows_from_sheet(ws, hr)

    if "Fuel Receipts" in wb.sheetnames:
        parsed.fuel_receipts = _parse_fuel_receipts(wb["Fuel Receipts"])

    if "Eligible Review - Dispenses" in wb.sheetnames:
        parsed.eligible_review_dispenses = _parse_eligible_review(wb["Eligible Review - Dispenses"])

    if "Combined Tank Summary" in wb.sheetnames:
        parsed.tank_summary = _parse_tank_summary(wb["Combined Tank Summary"])
        parsed.refund_rates = parsed.tank_summary.get("rates", [])

    for sheet_name in _discover_asset_sheets(wb):
        ws = wb[sheet_name]
        hr = _find_header_row(ws)
        if hr:
            _, rows = _rows_from_sheet(ws, hr)
            parsed.asset_sheets[sheet_name] = rows

    wb.close()
    return parsed


def normalize_tx_type(value: Any) -> str:
    return re.sub(r"[\s_]+", "-", _cell_str(value).upper())


def sheet_has_tank_litre_columns(headers: list[str]) -> list[str]:
    found: list[str] = []
    for h in headers:
        hl = h.lower()
        if "tank" in hl and ("before" in hl or "after" in hl):
            found.append(h)
    return found
