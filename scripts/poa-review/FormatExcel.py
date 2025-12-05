"""
FormatExcel.py - Excel Formatting Utility Module

This module provides a Formatter class that wraps openpyxl functionality to simplify
common Excel formatting tasks such as adding borders, colors, fonts, and alignment.

The Formatter class is designed to work with openpyxl Workbook objects and provides
a convenient API for applying consistent formatting across Excel worksheets.
"""

from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, Color
from openpyxl.utils.cell import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
import re
import openpyxl as ox 


class Formatter:
	"""
	Excel formatting utility class that simplifies common formatting operations.
	
	This class provides methods to apply borders, colors, fonts, alignment, and other
	formatting to Excel worksheets. It maintains a reference to a workbook and worksheet,
	allowing you to chain formatting operations easily.
	
	Attributes:
		wb (openpyxl.Workbook): The workbook being formatted
		ws (openpyxl.Worksheet): The active worksheet being formatted
		columns_deleted (list): Tracks which columns have been deleted (for offset calculations)
		ALL_SIDES_THIN_BORDER (Border): Class constant for thin borders on all sides
	"""

	# Class constant: Thin border on all four sides
	ALL_SIDES_THIN_BORDER = Border(left=Side(style='thin'), 
		                         right=Side(style='thin'), 
		                         top=Side(style='thin'), 
		                         bottom=Side(style='thin'))

	def __init__(self, wb: ox.Workbook, sheet_name: str = None):
		"""
		Initialize the Formatter with a workbook and optional sheet name.
		
		Args:
			wb (openpyxl.Workbook): The workbook to format
			sheet_name (str, optional): Name of the sheet to work with. 
			                          If None, uses the active sheet.
			                          If sheet doesn't exist, creates a new one.
		"""
		self.wb = wb
		self.ws = self.set_sheet(sheet_name)
		self.columns_deleted = []  # Track deleted columns for offset calculations

	def set_sheet(self, sheet_name):
		"""
		Set or create the worksheet to work with.
		
		Args:
			sheet_name (str, optional): Name of the sheet. If None, uses active sheet.
			
		Returns:
			openpyxl.Worksheet: The worksheet object
		"""
		if sheet_name in self.wb.sheetnames:
			return self.wb[sheet_name]
		elif sheet_name is None:
			print("Formatter is set to current sheet.")
			return self.wb.active
		else:
			print("Formatter created a new sheet.")
			return self.wb.create_sheet(title=sheet_name)


	def create_borders(self, row_start, num_rows, column_start, num_columns, border_style: Border = ALL_SIDES_THIN_BORDER):
		"""
		Apply borders to a rectangular range of cells.
		
		Args:
			row_start (int): Starting row number (1-indexed)
			num_rows (int): Number of rows to format
			column_start (int): Starting column number (1-indexed)
			num_columns (int): Number of columns to format
			border_style (Border): Border style to apply. Defaults to thin borders on all sides.
			
		Note:
			The range includes rows from row_start to row_start + num_rows (inclusive).
			The +1 in the range ensures the last row is included.
		"""
		for r in range(row_start, row_start + num_rows + 1):
			for c in range(column_start, column_start + num_columns):
				self.ws.cell(row=r, column=c).border = border_style

	def fill_cells(self, row_start, num_rows, column_start, num_columns, color="CCDAF5"):
		"""
		Fill a rectangular range of cells with a solid color.
		
		Args:
			row_start (int): Starting row number (1-indexed)
			num_rows (int): Number of rows to fill
			column_start (int): Starting column number (1-indexed)
			num_columns (int): Number of columns to fill
			color (str): Hex color code (e.g., "CCDAF5" for light blue).
			            Default is light blue (#CCDAF5), commonly used for headers.
			
		Note:
			Color should be a 6-character hex string without the '#' prefix.
		"""
		for r in range(row_start, row_start + num_rows):
			for c in range(column_start, column_start + num_columns):
				# Note: Color expects hex string, but openpyxl's Color(rgb=...) expects tuple
				# This works because openpyxl accepts hex strings directly in some contexts
				self.ws.cell(row=r, column=c).fill = PatternFill(patternType="solid", fgColor=Color(rgb=color))

	def make_bold(self, row_start, num_rows, column_start, num_columns, **kwargs):
		"""
		Make text bold in a rectangular range of cells.
		
		Args:
			row_start (int): Starting row number (1-indexed)
			num_rows (int): Number of rows to format
			column_start (int): Starting column number (1-indexed)
			num_columns (int): Number of columns to format
			**kwargs: Additional font properties (e.g., size=12, italic=True)
		"""
		for r in range(row_start, row_start + num_rows):
			for c in range(column_start, column_start + num_columns):
				self.ws.cell(row=r, column=c).font = Font(bold=True, **kwargs)

	def left_align(self, row_start, num_rows, column_start, num_columns):
		"""
		Left-align text in a rectangular range of cells.
		
		Args:
			row_start (int): Starting row number (1-indexed)
			num_rows (int): Number of rows to format
			column_start (int): Starting column number (1-indexed)
			num_columns (int): Number of columns to format
		"""
		for r in range(row_start, row_start + num_rows):
			for c in range(column_start, column_start + num_columns):
				self.ws.cell(row=r, column=c).alignment = Alignment(horizontal='left')

	def set_font(self, row_start, num_rows, column_start, num_columns, **kwargs):
		"""
		Set font properties for a rectangular range of cells.
		
		Args:
			row_start (int): Starting row number (1-indexed)
			num_rows (int): Number of rows to format
			column_start (int): Starting column number (1-indexed)
			num_columns (int): Number of columns to format
			**kwargs: Font properties (e.g., size=9, name='Arial', bold=True, italic=False)
		"""
		for r in range(row_start, row_start + num_rows):
			for c in range(column_start, column_start + num_columns):
				self.ws.cell(row=r, column=c).font = Font(**kwargs)

	def insert_text(self, row_start, column_start, text: list, default_number_format="#,##0.00"):
		"""
		Insert a 2D array of text/data into the worksheet starting at a specific position.
		
		Args:
			row_start (int): Starting row number (1-indexed)
			column_start (int): Starting column number (1-indexed)
			text (list): List of lists/tuples representing rows of data.
			            Example: [['Header1', 'Header2'], ['Value1', 'Value2']]
			default_number_format (str): Number format to apply to numeric values and formulas.
			                           Default is "#,##0.00" (thousands separator, 2 decimals).
			                           Set to None to skip number formatting.
		
		Raises:
			ValueError: If text is not a list of lists/tuples
			
		Note:
			Automatically detects formulas (starting with =, +, -, *, or containing 'sum')
			and applies number formatting to them as well.
		"""
		if not (isinstance(text[0], list) or isinstance(text[0], tuple)):
			raise ValueError(f"Text must be a list of lists or tuples. Eg [['oh','happy','day'],['Gee','Wizz','Bob']]. Got {text} instead.")

		# Pattern to detect Excel formulas (simple detection - may need refinement)
		formula_pattern = re.compile(r"[=\+\*-]|sum")

		for i, row in enumerate(text, start=row_start):
			for j, value in enumerate(row, start=column_start):
				self.ws.cell(row=i, column=j).value = value
				# Apply number format to numbers and formulas if format is specified
				if default_number_format and (isinstance(value, float) or isinstance(value, int) or formula_pattern.match(str(value))):
					self.ws.cell(row=i, column=j).number_format = default_number_format

	def delete_columns(self, cols: list):
		"""
		Delete one or more columns from the worksheet.
		
		Args:
			cols (list): List of column numbers (1-indexed) to delete
			
		Note:
			Columns are deleted in sorted order. When multiple columns are deleted,
			subsequent column indices shift, so this method tracks deleted columns
			and adjusts indices accordingly to prevent errors.
		"""
		sorted_cols = sorted(cols)
		for col in sorted_cols:
			if col in self.columns_deleted:
				print(f"Column {col} already deleted")
				continue
			self.columns_deleted.append(col)
			# Calculate how many columns before this one have been deleted
			# This adjusts the column index for deletion
			adj_cols = sum(x < col for x in self.columns_deleted)
			self.ws.delete_cols(col - adj_cols)

	def move_cells(self, row_start, num_rows, column_start, num_columns, row_shift, column_shift, **kwargs):
		"""
		Move a rectangular range of cells to a new position.
		
		Args:
			row_start (int): Starting row number (1-indexed)
			num_rows (int): Number of rows to move
			column_start (int): Starting column number (1-indexed)
			num_columns (int): Number of columns to move
			row_shift (int): Number of rows to shift (positive = down, negative = up)
			column_shift (int): Number of columns to shift (positive = right, negative = left)
			**kwargs: Additional arguments passed to openpyxl's move_range method
		"""
		column_start_letter = get_column_letter(column_start)
		column_end_letter = get_column_letter(column_start + num_columns - 1)

		row_end = row_start + num_rows

		self.ws.move_range(f"{column_start_letter}{row_start}:{column_end_letter}{row_end}", rows=row_shift, cols=column_shift, **kwargs)

	def create_abco_style_table(self, row_start, num_rows, column_start, num_columns, text, number_format=None):
		"""
		Create a formatted table in ABCO standard style.
		
		This is a convenience method that creates a complete formatted table with:
		- Data inserted from the text parameter
		- Borders on all cells
		- Header row with light blue background (#CCDAF5) and bold text
		- Optional number formatting
		
		Args:
			row_start (int): Starting row number (1-indexed)
			num_rows (int): Number of data rows (header row is additional)
			column_start (int): Starting column number (1-indexed)
			num_columns (int): Number of columns
			text (list): List of lists where first row is header, rest are data rows
			number_format (str, optional): Number format string (e.g., "#,##0.00")
			
		Raises:
			ValueError: If text dimensions don't match num_rows and num_columns
		"""
		if len(text) - 1 != num_rows: 
			raise ValueError("Text and table dimensions do not correspond.")
		elif len(text[0]) != num_columns:
			raise ValueError("Text and table dimensions do not correspond.")

		# Insert the data
		self.insert_text(row_start, column_start, text, default_number_format=number_format)
		# Add borders to all cells
		self.create_borders(row_start, num_rows, column_start, num_columns, border_style=self.ALL_SIDES_THIN_BORDER)
		# Format header row: light blue background and bold
		self.fill_cells(row_start, 1, column_start, len(text[0]), color="CCDAF5")
		self.make_bold(row_start, 1, column_start, len(text[0]))

	def generate_two_column_add_formulas(self, row_start1, row_start2, num_rows, column1, column2):
		"""
		Generate Excel formulas that add values from two columns.
		
		Creates formulas like "=A1 + B1", "=A2 + B2", etc. for a specified range.
		
		Args:
			row_start1 (int): Starting row for first column (1-indexed)
			row_start2 (int): Starting row for second column (1-indexed)
			num_rows (int): Number of formulas to generate
			column1 (int): First column number (1-indexed)
			column2 (int): Second column number (1-indexed)
			
		Returns:
			list: List of lists, each containing a single formula string.
			      Example: [["=A1 + B1"], ["=A2 + B2"]]
		"""
		col1_letter = get_column_letter(column1)
		col2_letter = get_column_letter(column2)

		formulas = [[f"={col1_letter}{row1} + {col2_letter}{row2}"] 
		            for row1, row2 in zip(range(row_start1, row_start1 + num_rows), 
		                                  range(row_start2, row_start2 + num_rows))]

		return formulas
 
	def single_line_text_entry(self, row_start, column_start, text, merge_columns=1, bold=False, italic=False, fill_color=None, border=None, number_format=None):
		"""
		Insert and format a single line of text, optionally merging cells.
		
		This is a convenience method for creating formatted headers or labels
		that may span multiple columns.
		
		Args:
			row_start (int): Row number (1-indexed)
			column_start (int): Starting column number (1-indexed)
			text (str): Text to insert
			merge_columns (int): Number of columns to merge (default: 1, no merging)
			bold (bool): Make text bold (default: False)
			italic (bool): Make text italic (default: False)
			fill_color: Color object or hex string for cell background (default: None)
			border: Border object to apply (default: None)
			number_format (str): Number format string (default: None)
		"""
		cell = self.ws.cell(row=row_start, column=column_start)

		cell.value = text
		cell.font = Font(bold=bold, italic=italic)

		if fill_color:
			cell.fill = PatternFill(patternType='solid', fgColor=fill_color)

		if border:
			cell.border = border

		if number_format:
			cell.number_format = number_format

		# If merging multiple columns, merge them and apply formatting to all cells
		if merge_columns > 1:
			end_column = column_start + merge_columns - 1
			self.ws.merge_cells(start_row=row_start, start_column=column_start, 
			                  end_row=row_start, end_column=end_column)

			# Apply formatting to all cells in the merged range
			for col in range(column_start, column_start + merge_columns):
				merged_cell = self.ws.cell(row=row_start, column=col)
				merged_cell.font = Font(bold=bold, italic=italic)
				if fill_color:
					merged_cell.fill = PatternFill(patternType='solid', fgColor=fill_color)
				if border:
					merged_cell.border = border

	def dataframe_to_excel(self, df):
		"""
		Convert a pandas DataFrame to Excel with automatic header formatting.
		
		Writes the DataFrame to the worksheet starting at row 1, then formats
		the header row with light blue background and bold text.
		
		Args:
			df (pandas.DataFrame): The DataFrame to write to Excel
		"""
		# Write DataFrame rows (header + data) to worksheet
		for r in dataframe_to_rows(df, index=False, header=True):
			self.ws.append(r)

		# Format header row: light blue background and bold
		self.fill_cells(row_start=1, num_rows=1, column_start=1, num_columns=len(df.columns), color="CCDAF5")
		self.make_bold(row_start=1, num_rows=1, column_start=1, num_columns=len(df.columns))

