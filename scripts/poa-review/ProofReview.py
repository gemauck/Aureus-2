"""
ProofReview.py - Proof of Activity (POA) Review and Analysis Module

This module processes fuel transaction and proof of activity data to analyze compliance
and usage patterns for fuel refund eligibility.

BUSINESS CONTEXT:
- POA (Proof of Activity): Records that prove an asset was in use (e.g., empty tank photos)
- Transactions: Fuel dispense/usage records that require proof to be eligible for refunds
- SMR (Service Meter Reading): Hours or kilometers of asset usage

The script identifies:
1. Assets that have transactions but no proof records (non-compliant)
2. How many proof records exist before each transaction
3. Time gaps between proof records and transactions
4. Total SMR usage grouped by transaction batches

OUTPUT:
Generates formatted Excel reports with conditional formatting highlighting:
- Green rows: Transactions with associated proof records
- Yellow cells: Missing SMR data
- Bold rows: Section headers
"""

import pandas as pd
import numpy as np
from FormatExcel import Formatter
from os import listdir

# Use streaming Excel write for large files to avoid OOM (openpyxl holds ~50x file size in RAM otherwise)
LARGE_FILE_THRESHOLD = 50000

# Styles for write_only path (avoid loading full sheet into memory)
def _write_only_excel(review, output_path, output_cols, bold_rows, green_rows, yellow_col16, yellow_col_index=15):
	from openpyxl import Workbook
	from openpyxl.cell import WriteOnlyCell
	from openpyxl.styles import PatternFill, Font, Alignment, Color
	fill_header = PatternFill(patternType="solid", fgColor=Color(rgb="CCDAF5"))
	fill_green = PatternFill(patternType="solid", fgColor=Color(rgb="D9EAD3"))
	fill_yellow = PatternFill(patternType="solid", fgColor=Color(rgb="FFF2CC"))
	font_9 = Font(size=9)
	font_9_bold = Font(size=9, bold=True)
	align_left = Alignment(horizontal="left")
	num_cols = len(output_cols)
	wb = Workbook(write_only=True)
	ws = wb.create_sheet(title="Details as Assets")
	# Header row
	header_cells = []
	for j, col in enumerate(output_cols):
		c = WriteOnlyCell(ws, value=col)
		c.fill = fill_header
		c.font = font_9_bold
		c.alignment = align_left
		header_cells.append(c)
	ws.append(header_cells)
	# Data rows: use numpy for fast iteration (avoids .iloc[] and row[col] lookups)
	data_arr = review[output_cols].values
	bold_arr = bold_rows.values
	green_arr = green_rows.values
	yellow_arr = yellow_col16.values
	for i in range(len(review)):
		row_vals = data_arr[i]
		row_cells = []
		for j in range(num_cols):
			val = row_vals[j]
			if pd.isna(val):
				val = None
			c = WriteOnlyCell(ws, value=val)
			c.font = font_9_bold if bold_arr[i] else font_9
			c.alignment = align_left
			if green_arr[i]:
				c.fill = fill_green
			if j == yellow_col_index and yellow_arr[i]:
				c.fill = fill_yellow
			row_cells.append(c)
		ws.append(row_cells)
	wb.save(output_path)

# Columns to include in the final review report
review_cols = [
    "Date & Time","Transaction ID","Asset Description","Asset Number","Asset Group","Asset Tank Size (L)","Asset Meter Type (Hr/Km)",
    "Storage Tank","Fuel Pump","Litres","Total Fuel Used (L)","Operation Description / Comment","Refund Eligibility","Opening SMR",
    "Closing SMR","Total SMR Usage","Material","Location.1","Loads / Tonnes","Activity","Comments","Source","Custom Attribute",
    "No POA Asset","Count of proof before transaction","Time since last activity","total smr","Dispense with no proof"
]

class POAReview:
	"""
	Analyzes fuel transaction data and proof of activity records.
	
	This class processes a DataFrame containing both transaction records (fuel dispenses)
	and proof records (evidence of asset usage). It identifies relationships between them,
	calculates compliance metrics, and prepares data for reporting.
	
	BUSINESS RULES:
	- Transactions are identified by having a "Transaction ID"
	- Proof records are identified by having NO "Transaction ID" but having an "Asset Number"
	- Consecutive transactions within 1 hour are grouped together as a single "label"
	- Each transaction should have proof records associated with it for refund eligibility
	
	Attributes:
		data (pd.DataFrame): The input data containing transactions and proof records
		transaction_mask (pd.Series): Boolean mask identifying transaction rows
		proof_mask (pd.Series): Boolean mask identifying proof record rows
	"""

	def __init__(self, per_asset_report: pd.DataFrame):
		"""
		Initialize the POAReview with transaction and proof data.
		
		Args:
			per_asset_report (pd.DataFrame): DataFrame containing both transaction and proof records.
			                                Must have columns: "Transaction ID", "Asset Number", "Date & Time"
		
		The constructor:
		1. Creates boolean masks to identify transaction vs proof records
		2. Converts date columns to datetime format for time-based analysis
		3. Converts high-cardinality object columns to category to reduce memory
		"""
		self.data = per_asset_report
		
		# Create masks to identify different record types
		# Transactions: Have a Transaction ID (but not the header "Transaction ID")
		self.transaction_mask = (self.data["Transaction ID"].notna()) & (self.data["Transaction ID"].astype(str).str.strip() != "Transaction ID")
		
		# Proof records: No Transaction ID, but have an Asset Number
		self.proof_mask = (self.data["Transaction ID"].isna()) & (self.data["Asset Number"].notna())
		
		# Convert "Date & Time" to datetime (single pass)
		self.data["Date & Time"] = pd.to_datetime(self.data["Date & Time"], yearfirst=True, errors="coerce")
		
		# Use category for repeated strings to cut memory (big win at 30k+ rows)
		for col in ("Asset Number", "Transaction ID", "Source"):
			if col in self.data.columns and self.data[col].dtype == object:
				self.data[col] = self.data[col].astype("category")

	def mark_no_poa_assets(self):
		"""
		Mark assets that have transactions but no proof of activity records.
		
		BUSINESS RULE: For fuel refund eligibility, transactions must have associated
		proof records. Assets with transactions but no proof records are non-compliant.
		
		Returns:
			pd.DataFrame: The data with "No POA Asset" column populated
			
		Note:
			Sets "No POA Asset" = "No Proof of Use Asset" for assets that have
			transactions but never appear in proof records.
		"""
		# Set of asset numbers that have proof records (set for fast .isin())
		poa_assets = set(self.data.loc[self.proof_mask, "Asset Number"].dropna().unique())
		# Mark assets that have transactions but no proof records; exclude header and NaN
		self.data.loc[(~self.data["Asset Number"].isin(poa_assets)) &
		              (self.data["Asset Number"].notna()) &
		              (self.data["Asset Number"].astype(str).str.strip() != "Asset Number"),
		              "No POA Asset"] = "No Proof of Use Asset"

		return self.data

	def mark_consecutive_transactions(self):
		"""
		Identify consecutive transactions that occur within 1 hour of each other.
		
		BUSINESS RULE: Transactions within 1 hour are considered part of the same
		dispensing session and should be grouped together for analysis.
		
		Returns:
			pd.DataFrame: The data with "is consec" column populated
			
		Note:
			- "is consec" = 0: Transaction is consecutive (within 1 hour of previous)
			- "is consec" = 1: Transaction starts a new group (more than 1 hour after previous)
			This is used later to create "labels" that group related transactions.
		"""
		# Calculate time difference between consecutive transactions
		time_between_dispenses = self.data.loc[self.transaction_mask, "Date & Time"].diff()

		# Mark as consecutive (0) if within 1 hour, otherwise new group (1); use int8 to save memory
		self.data.loc[self.transaction_mask, "is consec"] = np.where(
			(time_between_dispenses > pd.Timedelta(hours=0)) &
			(time_between_dispenses < pd.Timedelta(hours=1)),
			0,
			1,
		).astype(np.int8)

		return self.data

	def label_rows(self):
		"""
		Create unique labels to group related transactions and their proof records.
		
		LABEL FORMAT: "{AssetNumber}-{GroupNumber}"
		Example: "ASSET001-1", "ASSET001-2", "ASSET002-1"
		
		BUSINESS LOGIC:
		1. Transactions are grouped by asset and consecutive status
		2. Each new transaction group (separated by >1 hour) gets a new label number
		3. Proof records are assigned the same label as the next transaction for that asset
		
		Returns:
			pd.DataFrame: The data with "label" column populated
			
		Note:
			The label is used to link proof records to their associated transactions.
			Proof records are backward-filled with the label of the next transaction
			for the same asset, meaning a proof record is associated with the
			transaction that comes after it chronologically.
		"""
		# Ensure consecutive transactions are marked first
		if "is consec" not in self.data.columns:
			self.data.loc[self.transaction_mask, :] = self.mark_consecutive_transactions()

		# Create labels for transactions: AssetNumber-GroupNumber
		# cumsum() on "is consec" creates incrementing group numbers
		# (each 1 in "is consec" starts a new group)
		self.data.loc[self.transaction_mask, "label"] = (
			self.data.loc[self.transaction_mask, "Asset Number"].astype(str) + 
			"-" + 
			self.data.loc[self.transaction_mask, "is consec"].cumsum().astype(int).astype(str)
		)

		# Assign proof records the same label as the next transaction for that asset
		self.data.loc[self.transaction_mask | self.proof_mask, "label"] = (
			self.data.loc[self.transaction_mask | self.proof_mask, :]
			.groupby("Asset Number")["label"]
			.bfill()
		)
		# Category reduces memory for many repeated labels
		self.data["label"] = self.data["label"].astype("category")
		# Drop temporary column to free memory
		self.data.drop(columns=["is consec"], inplace=True, errors="ignore")

		return self.data

	def count_proof_before_transaction(self):
		"""
		Count how many proof records exist for each transaction label.
		
		BUSINESS METRIC: This indicates compliance - transactions should have
		proof records associated with them. A count of 0 means no proof exists.
		
		Returns:
			pd.DataFrame: The data with "Count of proof before transaction" column populated
			
		Note:
			- Counts proof records grouped by label
			- Maps the count to transactions with the same label
			- Transactions with no proof records get a count of 0
		"""
		# Ensure labels exist
		if "label" not in self.data.columns:
			self.data = self.label_rows()

		# Count proof records per label
		count = self.data.loc[self.proof_mask, "label"].value_counts().to_dict()
		
		# Map counts to transactions; use int32 to save memory
		mapped = self.data.loc[self.transaction_mask, "label"].map(count)
		self.data.loc[self.transaction_mask, "Count of proof before transaction"] = pd.to_numeric(mapped, errors="coerce").fillna(0).astype(np.int32)

		return self.data

	def time_since_last_activity(self):
		"""
		Calculate the time (in hours) since the last proof record for each row.
		
		BUSINESS METRIC: This measures the gap between proof records and transactions.
		Large gaps may indicate compliance issues or missing proof records.
		
		Returns:
			pd.DataFrame: The data with "Time since last activity" column populated (in hours)
			
		PROCESS:
		1. Extract timestamps from proof records (where Transaction ID is NaN)
		2. Forward-fill these timestamps by asset to propagate to subsequent transactions
		3. Calculate time difference between current row and last proof record
		4. Convert to hours (multiply by 24)
		"""
		# Ensure prerequisite columns exist
		if "Count of proof before transaction" not in self.data.columns:
			self.data = self.count_proof_before_transaction()

		# Step 1: Extract timestamps from proof records only
		# .where() keeps values where condition is True, sets others to NaN
		self.data.loc[self.transaction_mask | self.proof_mask, "Last Empty Time"] = (
			self.data.loc[self.transaction_mask | self.proof_mask, "Date & Time"]
			.where(self.data.loc[self.transaction_mask | self.proof_mask, "Transaction ID"].isna())
		)

		# Step 2: Forward-fill proof timestamps by asset
		# This propagates each proof record's timestamp to all subsequent rows for that asset
		self.data.loc[self.transaction_mask | self.proof_mask, "Last Empty Time"] = (
			self.data.loc[self.transaction_mask | self.proof_mask, :]
			.groupby("Asset Number")["Last Empty Time"]
			.ffill()
		)

		# Step 3: Time difference in hours; use float32 to save memory
		delta = (
			self.data.loc[self.transaction_mask | self.proof_mask, "Date & Time"] -
			self.data.loc[self.transaction_mask | self.proof_mask, "Last Empty Time"]
		)
		self.data.loc[self.transaction_mask | self.proof_mask, "Time since last activity"] = (delta.dt.total_seconds() / 3600.0).astype(np.float32)
		# Drop temp column to free memory
		self.data.drop(columns=["Last Empty Time"], inplace=True, errors="ignore")

		return self.data

	def total_smr(self, sources):
		"""
		Calculate total SMR (Service Meter Reading) usage per transaction label.
		
		SMR represents hours or kilometers of asset usage. This aggregates SMR
		from specified sources and groups by transaction label.
		
		Args:
			sources (list): List of source names to include in the calculation.
			               Example: ["Inmine: Daily Diesel Issues"]
			               
		Returns:
			pd.DataFrame: The data with "total smr" column populated
			
		Note:
			- Only includes rows where "Source" is in the provided sources list
			- Sums "Total SMR Usage" grouped by label
			- Maps the total to transactions with matching labels
			- Transactions with no matching SMR data get 0
		"""
		# Sum SMR by label for specified sources without creating a full-column copy (low memory)
		source_mask = self.data["Source"].isin(sources)
		smr_numeric = pd.to_numeric(self.data.loc[source_mask, "Total SMR Usage"], errors="coerce").fillna(0).astype(np.float32)
		hours = (
			self.data.loc[source_mask]
			.assign(_s=smr_numeric)
			.groupby("label", observed=True)["_s"]
			.sum()
			.to_dict()
		)
		del smr_numeric, source_mask

		# Map to transactions; float32 to save memory
		mapped = self.data.loc[self.transaction_mask, "label"].map(hours)
		self.data["total smr"] = np.nan
		self.data.loc[self.transaction_mask, "total smr"] = pd.to_numeric(mapped, errors="coerce").fillna(0).astype(np.float32)
		return self.data

	def mark_dispense_with_no_proof(self):
		"""Set 'Dispense with no proof' = Yes for transaction rows that have zero proof records (gap for follow-up)."""
		if "Count of proof before transaction" not in self.data.columns:
			self.count_proof_before_transaction()
		self.data["Dispense with no proof"] = ""
		count_col = self.data["Count of proof before transaction"]
		self.data.loc[
			self.transaction_mask & (pd.to_numeric(count_col, errors="coerce").fillna(-1) == 0),
			"Dispense with no proof",
		] = "Yes"
		return self.data

def format_review(review, filename, output_path=None, original_columns=None):
	"""
	Format and export the review data to an Excel file with conditional formatting.
	
	CONDITIONAL FORMATTING RULES:
	1. Header row: Light blue background (#CCDAF5), bold, left-aligned
	2. Section headers (non-timestamp in column 0): Bold text
	3. Transactions with proof (has Transaction ID and timestamp): Green background (#D9EAD3)
	4. Missing SMR data (column 15 is empty/0): Yellow background (#FFF2CC) in column 16
	
	Args:
		review (pd.DataFrame): The processed review data
		filename (str): Input filename (used to generate output filename if output_path not provided)
		output_path (str, optional): Full path to output file. If not provided, uses default location.
		original_columns (list, optional): If provided, output keeps these columns in order, then appends
			the computed columns (No POA Asset, Count of proof..., Time since last activity, total smr, Dispense with no proof). No extra columns and no internal "label" column.
	"""
	COMPUTED_COLS = ["No POA Asset", "Count of proof before transaction", "Time since last activity", "total smr", "Dispense with no proof"]
	if original_columns is not None:
		# Output = original columns (same order) + 4 computed only
		output_cols = [c for c in original_columns if c in review.columns] + [
			c for c in COMPUTED_COLS if c in review.columns
		]
	else:
		# Legacy: ensure review_cols exist, then review_cols + rest (excluding internal "label")
		missing = [c for c in review_cols if c not in review.columns]
		if missing:
			for c in missing:
				review[c] = np.nan
		output_cols = list(review_cols) + [c for c in review.columns if c not in review_cols and c != "label"]
	review = review.reindex(columns=output_cols)

	# Generate output path if not provided
	if output_path is None:
		# Generate output filename (remove .csv extension if present, add .xlsx)
		import os
		output_dir = "./Output/Isibonelo Current"
		os.makedirs(output_dir, exist_ok=True)
		output_path = os.path.join(output_dir, filename.rstrip(".csv") + ".xlsx")
	
	# Ensure output directory exists
	import os
	os.makedirs(os.path.dirname(output_path), exist_ok=True)

	n = len(review)
	num_cols = len(output_cols)
	# Precompute format flags once (used by both paths)
	col_dt = review["Date & Time"]
	col_txn = review["Transaction ID"]
	col_smr = review["Total SMR Usage"]
	smr_numeric = pd.to_numeric(col_smr, errors="coerce").fillna(0)
	# Fast path when column is already datetime64
	if pd.api.types.is_datetime64_any_dtype(col_dt):
		is_ts = col_dt.notna()
	else:
		is_ts = pd.Series([isinstance(x, pd.Timestamp) for x in col_dt], index=review.index)
	has_txn = col_txn.notna() & (col_txn.astype(str).str.strip() != "")
	smr_empty = col_smr.isna() | (smr_numeric == 0)
	bold_rows = ~is_ts & col_dt.notna()
	green_rows = has_txn & is_ts
	yellow_col16 = is_ts & ((smr_empty & ~has_txn) | (smr_numeric == 0))
	yellow_col_index = output_cols.index("Total SMR Usage") if "Total SMR Usage" in output_cols else -1
	yellow_col_1based = (yellow_col_index + 1) if yellow_col_index >= 0 else 16

	if n > LARGE_FILE_THRESHOLD:
		# Streaming write: minimal memory so large files don't OOM the server
		_write_only_excel(review, output_path, output_cols, bold_rows, green_rows, yellow_col16, yellow_col_index)
	else:
		# Standard path with Formatter (full sheet in memory)
		with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
			review.to_excel(writer, sheet_name="Details as Assets", index=False)
			wb = writer.book
			ws = writer.sheets["Details as Assets"]
			format_review = Formatter(wb)
			format_review.fill_cells(row_start=1, num_rows=1, column_start=1, num_columns=num_cols + 4)
			format_review.left_align(row_start=1, num_rows=1, column_start=1, num_columns=num_cols + 4)
			format_review.create_borders(row_start=1, num_rows=1, column_start=1, num_columns=num_cols + 4, border_style=None)
			format_review.set_font(row_start=1, num_rows=ws.max_row, column_start=1, num_columns=ws.max_column, size=9)
			def contiguous_ranges(series):
				r, i = [], 0
				while i < len(series):
					if series.iloc[i]:
						start = i
						while i < len(series) and series.iloc[i]:
							i += 1
						r.append((start + 2, i - start))
					else:
						i += 1
				return r
			for start, num_rows in contiguous_ranges(bold_rows):
				format_review.make_bold(row_start=start, num_rows=num_rows, column_start=1, num_columns=num_cols + 4, size=9)
			for start, num_rows in contiguous_ranges(green_rows):
				format_review.fill_cells(row_start=start, num_rows=num_rows, column_start=1, num_columns=num_cols + 4, color="D9EAD3")
			if yellow_col_1based <= num_cols:
				for start, num_rows in contiguous_ranges(yellow_col16):
					format_review.fill_cells(row_start=start, num_rows=num_rows, column_start=yellow_col_1based, num_columns=1, color="FFF2CC")

# ============================================================================
# MAIN EXECUTION SCRIPT
# ============================================================================
# Processes all Excel files in the input directory and generates formatted reports
# ============================================================================

if __name__ == "__main__":
	# Input directory containing Excel files to process
	input_dir = './input/Isibonelo Current'
	
	# Get list of all files in input directory
	inputs = listdir(input_dir)
	
	for file in inputs:
		print("--------starting review---------")
		print("		opening file	")
		print(file)
		
		# Read Excel file (skip first row which may be a header)
		# NOTE: Adjust skiprows if your file format differs
		data = pd.read_excel(f"{input_dir}/{file}", skiprows=1)
		
		print("		reviewing		")
		
		# Initialize review and run analysis
		review = POAReview(data)
		review.mark_no_poa_assets()  # Identify non-compliant assets
		review.time_since_last_activity()  # Calculate time gaps
		review.total_smr(["Inmine: Daily Diesel Issues"])  # Calculate SMR totals
		
		print("		formatting		")
		
		# Generate formatted Excel output
		format_review(review.data, file)
		
		print('		done   	')





# .to_excel("./Output/Mafube/Mafube Coal Mining (Pty) Ltd - Proof of Activity - May 2025.xlsx")