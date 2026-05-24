#!/usr/bin/env python3
"""Convert InsightWare POA Excel export to CSV (detects header row, streams large files)."""
import csv
import sys

import pandas as pd


def find_header_row_preview(input_file: str, max_scan: int = 8) -> tuple[int, list[str]]:
    preview = pd.read_excel(input_file, header=None, nrows=max_scan)
    for i in range(len(preview)):
        cells = [str(v).strip() if pd.notna(v) else '' for v in preview.iloc[i].values]
        row_str = ' '.join(cells).lower()
        if (
            'transaction' in row_str
            and 'asset' in row_str
            and ('date' in row_str or 'time' in row_str)
        ):
            return i, cells
    first = [str(v).strip() if pd.notna(v) else '' for v in preview.iloc[0].values]
    return 0, first


def normalize_headers(headers: list[str]) -> list[str]:
    counts: dict[str, int] = {}
    out: list[str] = []
    for raw in headers:
        base = str(raw or '').strip()
        if not base or base.lower().startswith('unnamed'):
            out.append('')
            continue
        if base not in counts:
            counts[base] = 0
            out.append(base)
        else:
            counts[base] += 1
            out.append(f'{base}.{counts[base]}')
    return out


def _cell_str(value) -> str:
    if value is None:
        return ''
    if isinstance(value, float) and pd.isna(value):
        return ''
    return str(value).strip()


def convert_xlsx_streaming(input_file: str, output_csv: str, header_idx: int, header_row: list[str]) -> int:
    """Low-memory .xlsx → CSV using openpyxl read_only (avoids OOM on large exports)."""
    from openpyxl import load_workbook

    n_cols = len(header_row)
    total_rows = 0
    log_every = 25000

    wb = load_workbook(input_file, read_only=True, data_only=True)
    try:
        ws = wb.active
        with open(output_csv, 'w', encoding='utf-8', newline='') as csv_file:
            writer = csv.writer(csv_file)
            writer.writerow(header_row[:n_cols])

            for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
                if row_idx <= header_idx:
                    continue
                values = [_cell_str(v) for v in row[:n_cols]]
                if len(values) < n_cols:
                    values.extend([''] * (n_cols - len(values)))
                if not any(values):
                    continue
                writer.writerow(values)
                total_rows += 1
                if total_rows % log_every == 0:
                    print(f'Converted {total_rows} data rows...', flush=True)
    finally:
        wb.close()

    print(f'Success! CSV created with {total_rows} data rows', flush=True)
    return total_rows


def convert_excel_pandas_chunks(
    input_file: str,
    output_csv: str,
    header_idx: int,
    header_row: list[str],
    chunk_size: int = 50000,
) -> int:
    """Fallback for .xls and environments without openpyxl read_only."""
    n_cols = len(header_row)
    skip_rows = header_idx + 1
    total_rows = 0

    with open(output_csv, 'w', encoding='utf-8', newline='') as csv_file:
        writer = csv.writer(csv_file)
        writer.writerow(header_row[:n_cols])

        while True:
            chunk = pd.read_excel(
                input_file,
                skiprows=skip_rows,
                nrows=chunk_size,
                header=None,
            )
            if chunk.empty:
                break

            n = chunk.shape[1]
            if n < n_cols:
                for j in range(n, n_cols):
                    chunk[j] = ''
            elif n > n_cols:
                chunk = chunk.iloc[:, :n_cols]

            chunk.columns = list(range(n_cols))
            chunk.to_csv(csv_file, index=False, header=False, mode='a', lineterminator='\n')
            total_rows += len(chunk)
            skip_rows += len(chunk)
            print(f'Converted {total_rows} data rows...', flush=True)
            if len(chunk) < chunk_size:
                break

    print(f'Success! CSV created with {total_rows} data rows', flush=True)
    return total_rows


def convert_excel_to_csv(input_file: str, output_csv: str, chunk_size: int = 50000) -> int:
    header_idx, header_row = find_header_row_preview(input_file)
    header_row = normalize_headers(header_row)
    n_cols = len(header_row)
    print(f'Header row index: {header_idx}, columns: {n_cols}', flush=True)

    ext = input_file.lower().rsplit('.', 1)[-1] if '.' in input_file else ''
    if ext == 'xlsx':
        try:
            return convert_xlsx_streaming(input_file, output_csv, header_idx, header_row)
        except Exception as exc:
            print(f'openpyxl streaming failed ({exc}); falling back to pandas chunks', flush=True)

    return convert_excel_pandas_chunks(input_file, output_csv, header_idx, header_row, chunk_size)


def main() -> int:
    if len(sys.argv) != 3:
        print('Usage: convert_excel_for_poa.py <input.xlsx> <output.csv>', file=sys.stderr)
        return 1
    try:
        convert_excel_to_csv(sys.argv[1], sys.argv[2])
        return 0
    except Exception as exc:
        print(f'Error: {exc}', file=sys.stderr)
        import traceback
        traceback.print_exc(file=sys.stderr)
        return 1


if __name__ == '__main__':
    sys.exit(main())
