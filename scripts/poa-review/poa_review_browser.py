"""
POA Review - standalone script for Pyodide (browser).
Reads /tmp/input.csv, options from /tmp/options.json, writes /tmp/output.xlsx.
Bundled with poaStrengthEvaluator.py via /api/poa-review/browser-script.
"""
import json
import os
import pandas as pd
import numpy as np

try:
    from poaStrengthEvaluator import (
        STRENGTH_FILL,
        STRENGTH_INSUFFICIENT,
        COMPLIANCE_POINTS_FILL,
        SHORTFALLS_COLUMN_FILL,
        POA_COMPLIANCE_POINTS_COL,
        evaluate_all_labels,
        format_shortfalls,
        compliance_points_from_result,
        summarize_strength_results,
    )
except ModuleNotFoundError:
    # Pyodide bundle: evaluator code is exec'd in the same namespace above this import.
    pass

INPUT_CSV = "/tmp/input.csv"
OPTIONS_JSON = "/tmp/options.json"
OUTPUT_XLSX = "/tmp/output.xlsx"

REVIEW_COLS = [
    "Date & Time", "Transaction ID", "Asset Description", "Asset Number", "Asset Group",
    "Asset Tank Size (L)", "Asset Meter Type (Hr/Km)", "Storage Tank", "Fuel Pump", "Litres",
    "Total Fuel Used (L)", "Operation Description / Comment", "Refund Eligibility", "Opening SMR",
    "Closing SMR", "Total SMR Usage", "Material", "Location.1", "Loads / Tonnes", "Activity",
    "Comments", "Source", "Custom Attribute", "No POA Asset", "Count of proof before transaction",
    "Time since last activity", "total smr", "POA Strength", POA_COMPLIANCE_POINTS_COL, "POA Shortfalls"
]


def normalize_column_name(col_name):
    if pd.isna(col_name):
        return None
    return str(col_name).strip().lower()


def find_column(df, target_name):
    n = normalize_column_name(target_name)
    for col in df.columns:
        if normalize_column_name(col) == n:
            return col
    return None


def _write_only_excel(
    review,
    output_path,
    output_cols,
    bold_rows,
    green_rows,
    yellow_col16,
    yellow_col_index=15,
    strength_col_index=-1,
    strength_row_fills=None,
    compliance_col_index=-1,
    shortfalls_col_index=-1,
):
    from openpyxl import Workbook
    from openpyxl.cell import WriteOnlyCell
    from openpyxl.styles import PatternFill, Font, Alignment, Color
    fill_header = PatternFill(patternType="solid", fgColor=Color(rgb="CCDAF5"))
    fill_green = PatternFill(patternType="solid", fgColor=Color(rgb="D9EAD3"))
    fill_yellow = PatternFill(patternType="solid", fgColor=Color(rgb="FFF2CC"))
    strength_fills = {
        hex_color: PatternFill(patternType="solid", fgColor=Color(rgb=hex_color))
        for hex_color in STRENGTH_FILL.values()
    }
    fill_compliance = PatternFill(patternType="solid", fgColor=Color(rgb=COMPLIANCE_POINTS_FILL))
    fill_shortfalls = PatternFill(patternType="solid", fgColor=Color(rgb=SHORTFALLS_COLUMN_FILL))
    font_9 = Font(size=9)
    font_9_bold = Font(size=9, bold=True)
    align_left = Alignment(horizontal="left")
    num_cols = len(output_cols)
    wb = Workbook(write_only=True)
    ws = wb.create_sheet(title="Details as Assets")
    header_cells = []
    for col in output_cols:
        c = WriteOnlyCell(ws, value=col)
        c.fill = fill_header
        c.font = font_9_bold
        c.alignment = align_left
        header_cells.append(c)
    ws.append(header_cells)
    data_arr = review[output_cols].values
    bold_arr = bold_rows.values
    green_arr = green_rows.values
    yellow_arr = yellow_col16.values
    strength_list = strength_row_fills.tolist() if strength_row_fills is not None else None
    for i in range(len(review)):
        row_vals = data_arr[i]
        is_bold = bool(bold_arr[i])
        is_green = bool(green_arr[i])
        is_yellow_smr = bool(yellow_arr[i])
        hex_color = strength_list[i] if strength_list is not None else None
        has_strength = strength_col_index >= 0 and hex_color in strength_fills
        compliance_val = row_vals[compliance_col_index] if compliance_col_index >= 0 else None
        shortfalls_val = row_vals[shortfalls_col_index] if shortfalls_col_index >= 0 else None
        has_compliance_fill = compliance_col_index >= 0 and compliance_val is not None and str(compliance_val).strip() not in ("", "nan")
        has_shortfalls_fill = shortfalls_col_index >= 0 and shortfalls_val is not None and str(shortfalls_val).strip() not in ("", "nan")

        if not (is_bold or is_green or is_yellow_smr or has_strength or has_compliance_fill or has_shortfalls_fill):
            row_list = []
            for j in range(num_cols):
                val = row_vals[j]
                row_list.append(None if val is None or (isinstance(val, float) and pd.isna(val)) else val)
            ws.append(row_list)
            continue

        row_cells = []
        for j in range(num_cols):
            val = row_vals[j]
            if pd.isna(val):
                val = None
            needs_cell_style = (
                is_bold or is_green
                or (j == yellow_col_index and is_yellow_smr)
                or (j == strength_col_index and has_strength)
                or (j == compliance_col_index and has_compliance_fill)
                or (j == shortfalls_col_index and has_shortfalls_fill)
            )
            if not needs_cell_style:
                row_cells.append(val)
                continue
            c = WriteOnlyCell(ws, value=val)
            c.font = font_9_bold if is_bold else font_9
            c.alignment = align_left
            if is_green:
                c.fill = fill_green
            if j == yellow_col_index and is_yellow_smr:
                c.fill = fill_yellow
            if j == strength_col_index and has_strength:
                c.fill = strength_fills[hex_color]
            if j == compliance_col_index and has_compliance_fill:
                c.fill = fill_compliance
            if j == shortfalls_col_index and has_shortfalls_fill:
                c.fill = fill_shortfalls
            row_cells.append(c)
        ws.append(row_cells)
    wb.save(output_path)


class POAReview:
    def __init__(self, per_asset_report):
        self.data = per_asset_report
        txn_id_str = self.data["Transaction ID"].astype(str).str.strip()
        self.transaction_mask = (
            self.data["Transaction ID"].notna()
            & (txn_id_str != "")
            & (txn_id_str != "Transaction ID")
        )
        self.proof_mask = (~self.transaction_mask) & self.data["Asset Number"].notna()
        self.data["Date & Time"] = pd.to_datetime(
            self.data["Date & Time"], yearfirst=True, errors="coerce"
        )
        for col in ("Asset Number", "Transaction ID", "Source"):
            if col in self.data.columns and self.data[col].dtype == object:
                self.data[col] = self.data[col].astype("category")

    def mark_consecutive_transactions(self):
        time_between = self.data.loc[self.transaction_mask, "Date & Time"].diff()
        self.data.loc[self.transaction_mask, "is consec"] = np.where(
            (time_between > pd.Timedelta(hours=0))
            & (time_between < pd.Timedelta(hours=1)),
            0,
            1,
        ).astype(np.int8)
        return self.data

    def label_rows(self):
        if "is consec" not in self.data.columns:
            self.mark_consecutive_transactions()
        self.data.loc[self.transaction_mask, "label"] = (
            self.data.loc[self.transaction_mask, "Asset Number"].astype(str)
            + "-"
            + self.data.loc[self.transaction_mask, "is consec"].cumsum().astype(int).astype(str)
        )
        self.data.loc[self.transaction_mask | self.proof_mask, "label"] = (
            self.data.loc[self.transaction_mask | self.proof_mask, :]
            .groupby("Asset Number")["label"]
            .bfill()
        )
        self.data["label"] = self.data["label"].astype("category")
        self.data.drop(columns=["is consec"], inplace=True, errors="ignore")
        return self.data

    def mark_no_poa_assets(self):
        poa_assets = set(
            self.data.loc[self.proof_mask, "Asset Number"].dropna().unique()
        )
        self.data.loc[
            (~self.data["Asset Number"].isin(poa_assets))
            & (self.data["Asset Number"].notna())
            & (self.data["Asset Number"].astype(str).str.strip() != "Asset Number"),
            "No POA Asset",
        ] = "No Proof of Use Asset"
        return self.data

    def count_proof_before_transaction(self):
        if "label" not in self.data.columns:
            self.label_rows()
        count = self.data.loc[self.proof_mask, "label"].value_counts().to_dict()
        mapped = self.data.loc[self.transaction_mask, "label"].map(count)
        self.data.loc[
            self.transaction_mask, "Count of proof before transaction"
        ] = pd.to_numeric(mapped, errors="coerce").fillna(0).astype(np.int32)
        return self.data

    def time_since_last_activity(self):
        if "Count of proof before transaction" not in self.data.columns:
            self.count_proof_before_transaction()
        mask = self.transaction_mask | self.proof_mask
        self.data.loc[mask, "Last Empty Time"] = self.data.loc[
            mask, "Date & Time"
        ].where(self.proof_mask.loc[mask])
        self.data.loc[mask, "Last Empty Time"] = (
            self.data.loc[mask, :].groupby("Asset Number")["Last Empty Time"].ffill()
        )
        delta = (
            self.data.loc[mask, "Date & Time"]
            - self.data.loc[mask, "Last Empty Time"]
        )
        self.data.loc[mask, "Time since last activity"] = (
            delta.dt.total_seconds() / 3600.0
        ).astype(np.float32)
        self.data.drop(columns=["Last Empty Time"], inplace=True, errors="ignore")
        return self.data

    def total_smr(self, sources):
        source_mask = self.data["Source"].isin(sources)
        smr_numeric = pd.to_numeric(
            self.data.loc[source_mask, "Total SMR Usage"], errors="coerce"
        ).fillna(0).astype(np.float32)
        hours = (
            self.data.loc[source_mask]
            .assign(_s=smr_numeric)
            .groupby("label", observed=True)["_s"]
            .sum()
            .to_dict()
        )
        mapped = self.data.loc[self.transaction_mask, "label"].map(hours)
        self.data["total smr"] = np.nan
        self.data.loc[self.transaction_mask, "total smr"] = (
            pd.to_numeric(mapped, errors="coerce").fillna(0).astype(np.float32)
        )
        return self.data

    def evaluate_poa_strength(self):
        if "label" not in self.data.columns:
            self.label_rows()
        label_results = evaluate_all_labels(self.data, self.proof_mask, self.transaction_mask)
        self._label_results = label_results
        self._apply_label_results(label_results)
        return self.data

    def _apply_label_results(self, label_results):
        strength_map = {label: res["strength"] for label, res in label_results.items()}
        shortfall_map = {
            label: format_shortfalls(res.get("shortfalls") or [])
            for label, res in label_results.items()
        }
        compliance_map = {
            label: compliance_points_from_result(res)
            for label, res in label_results.items()
        }
        self.data["POA Strength"] = pd.Series([None] * len(self.data), index=self.data.index, dtype=object)
        self.data[POA_COMPLIANCE_POINTS_COL] = pd.Series([None] * len(self.data), index=self.data.index, dtype=object)
        self.data["POA Shortfalls"] = pd.Series([None] * len(self.data), index=self.data.index, dtype=object)
        labels_str = self.data["label"].astype(str)
        txn_idx = self.data.index[self.transaction_mask]
        self.data.loc[txn_idx, "POA Strength"] = labels_str.loc[txn_idx].map(strength_map)
        self.data.loc[txn_idx, POA_COMPLIANCE_POINTS_COL] = labels_str.loc[txn_idx].map(compliance_map)
        self.data.loc[txn_idx, "POA Shortfalls"] = labels_str.loc[txn_idx].map(shortfall_map)
        if "Count of proof before transaction" in self.data.columns:
            zero_proof = self.transaction_mask & (
                pd.to_numeric(self.data["Count of proof before transaction"], errors="coerce").fillna(0) == 0
            )
            no_shift_labels = {
                str(label) for label, res in label_results.items() if not res.get("shiftProofApplied")
            }
            override_mask = zero_proof & labels_str.isin(no_shift_labels)
            self.data.loc[override_mask, "POA Strength"] = STRENGTH_INSUFFICIENT
            self.data.loc[override_mask, POA_COMPLIANCE_POINTS_COL] = None
            self.data.loc[override_mask, "POA Shortfalls"] = "No proof-of-activity rows for this batch"
        self.strength_summary = summarize_strength_results(label_results)
        return self.data


def _strength_fill_series(review, output_cols):
    if "POA Strength" not in output_cols or "POA Strength" not in review.columns:
        return None
    fills = []
    for val in review["POA Strength"]:
        if pd.isna(val) or val is None or str(val).strip() == "":
            fills.append(None)
        else:
            fills.append(STRENGTH_FILL.get(str(val).strip(), STRENGTH_FILL[STRENGTH_INSUFFICIENT]))
    return pd.Series(fills, index=review.index)


def _build_review_from_csv(sources):
    data = pd.read_csv(INPUT_CSV, low_memory=True, dtype=str, na_values=[""])
    required = {
        "Transaction ID": ["transaction id", "transactionid", "txn id", "txnid"],
        "Asset Number": ["asset number", "assetnumber", "asset no", "assetno"],
        "Date & Time": ["date & time", "date and time", "datetime", "date", "timestamp"],
    }
    column_mapping = {}
    for expected_col, possible_names in required.items():
        found = expected_col if expected_col in data.columns else None
        if not found:
            for name in possible_names:
                found = find_column(data, name)
                if found:
                    break
        if found and found != expected_col:
            column_mapping[found] = expected_col
        if not found:
            raise ValueError(f"Missing required column: {expected_col}")
    if column_mapping:
        data = data.rename(columns=column_mapping)

    if "Source" not in data.columns and sources:
        data["Source"] = sources[0]
    original_columns = list(data.columns)
    for c in REVIEW_COLS:
        if c not in data.columns:
            data[c] = np.nan
    review = POAReview(data)
    review.mark_consecutive_transactions()
    review.label_rows()
    review.mark_no_poa_assets()
    review.count_proof_before_transaction()
    review.time_since_last_activity()
    review.total_smr(sources)
    return review, original_columns


def _write_output_excel(review, original_columns):
    COMPUTED_COLS = [
        "No POA Asset",
        "Count of proof before transaction",
        "Time since last activity",
        "total smr",
        "POA Strength",
        POA_COMPLIANCE_POINTS_COL,
        "POA Shortfalls",
    ]
    output_cols = [c for c in original_columns if c in review.data.columns and c not in COMPUTED_COLS] + [
        c for c in COMPUTED_COLS if c in review.data.columns
    ]
    review.data = review.data.reindex(columns=output_cols)

    col_dt = review.data["Date & Time"]
    col_txn = review.data["Transaction ID"]
    col_smr = review.data["Total SMR Usage"]
    smr_numeric = pd.to_numeric(col_smr, errors="coerce").fillna(0)
    is_ts = col_dt.notna()
    has_txn = col_txn.notna() & (col_txn.astype(str).str.strip() != "")
    smr_empty = col_smr.isna() | (smr_numeric == 0)
    bold_rows = ~is_ts & col_dt.notna()
    green_rows = has_txn & is_ts
    yellow_col16 = is_ts & ((smr_empty & ~has_txn) | (smr_numeric == 0))
    yellow_col_index = output_cols.index("Total SMR Usage") if "Total SMR Usage" in output_cols else -1
    strength_col_index = output_cols.index("POA Strength") if "POA Strength" in output_cols else -1
    compliance_col_index = output_cols.index(POA_COMPLIANCE_POINTS_COL) if POA_COMPLIANCE_POINTS_COL in output_cols else -1
    shortfalls_col_index = output_cols.index("POA Shortfalls") if "POA Shortfalls" in output_cols else -1
    strength_row_fills = _strength_fill_series(review.data, output_cols)

    _write_only_excel(
        review.data,
        OUTPUT_XLSX,
        output_cols,
        bold_rows,
        green_rows,
        yellow_col16,
        yellow_col_index,
        strength_col_index,
        strength_row_fills,
        compliance_col_index,
        shortfalls_col_index,
    )


def run():
    with open(OPTIONS_JSON, "r") as f:
        options = json.load(f)
    sources = options.get("sources", ["Inmine: Daily Diesel Issues"])
    review, original_columns = _build_review_from_csv(sources)
    review.evaluate_poa_strength()
    _write_output_excel(review, original_columns)


if __name__ == "__main__" and not globals().get("__POA_PYODIDE_WORKER__"):
    run()
