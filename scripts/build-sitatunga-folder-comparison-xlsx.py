#!/usr/bin/env python3
"""Build Sitatunga diesel-rebate folder comparison spreadsheet with content descriptions."""
import hashlib
import re
import zipfile
from datetime import datetime
from pathlib import Path
from xml.etree import ElementTree as ET

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

TINUS = Path("/Users/gemau/Downloads/New Information from Tinus 2")
DISPUTE = Path("/Users/gemau/Downloads/Sitatuga diesel rebate dispute 2")
OUT = Path("/Users/gemau/Desktop/Sitatunga-diesel-rebate-folder-comparison.xlsx")


def sha256_file(p: Path) -> str:
    h = hashlib.sha256()
    with p.open("rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def detect_kind(p: Path) -> str:
    with p.open("rb") as f:
        b = f.read(8)
    if b.startswith(b"%PDF"):
        return "pdf"
    if b.startswith(b"PK"):
        if p.suffix.lower() in (".xlsx", ".xlsm", ".docx"):
            return p.suffix.lower().lstrip(".")
        try:
            with zipfile.ZipFile(p) as z:
                names = z.namelist()
            if any(n.startswith("word/") for n in names):
                return "docx"
            if any("worksheets/" in n for n in names):
                return "xlsx"
        except OSError:
            pass
        return "zip-office"
    if p.suffix.lower() == ".eml":
        return "eml"
    return p.suffix.lower().lstrip(".") or "binary"


def probe_pdf(p: Path) -> str:
    from pypdf import PdfReader

    r = PdfReader(str(p))
    meta = r.metadata or {}
    title = str(meta.get("/Title") or meta.get("title") or "").strip()
    subject = str(meta.get("/Subject") or meta.get("subject") or "").strip()
    pages = len(r.pages)
    text = ""
    if r.pages:
        text = re.sub(r"\s+", " ", (r.pages[0].extract_text() or "").strip())[:220]
    parts = [f"PDF document, {pages} page(s)"]
    if title and title.lower() not in ("untitled",):
        parts.append(f"title «{title[:70]}»")
    if subject:
        parts.append(f"subject «{subject[:70]}»")
    if text:
        parts.append(f"starts: «{text[:180]}»")
    return "; ".join(parts)


def probe_xlsx(p: Path) -> str:
    wb = openpyxl.load_workbook(p, read_only=True, data_only=True)
    sheets = wb.sheetnames
    wb.close()
    preview = ", ".join(sheets[:5])
    if len(sheets) > 5:
        preview += f" (+{len(sheets) - 5} more)"
    return f"Excel workbook ({len(sheets)} sheet(s): {preview})"


def probe_docx(p: Path) -> str:
    with zipfile.ZipFile(p) as z:
        xml = z.read("doc/core.xml")
    root = ET.fromstring(xml)
    ns = {"dc": "http://purl.org/dc/elements/1.1/"}
    title = (root.findtext("dc:title", default="", namespaces=ns) or "").strip()
    subj = (root.findtext("dc:subject", default="", namespaces=ns) or "").strip()
    bits = ["Word document"]
    if title:
        bits.append(f"title «{title[:70]}»")
    if subj:
        bits.append(f"subject «{subj[:70]}»")
    return "; ".join(bits)


def probe_eml(p: Path) -> str:
    head = p.read_text(errors="replace")[:12000]

    def grab(rx: str) -> str:
        m = re.search(rx, head, re.M | re.I)
        return m.group(1).strip() if m else ""

    subj = grab(r"^Subject:\s*(.+)$")
    frm = grab(r"^From:\s*(.+)$")
    date = grab(r"^Date:\s*(.+)$")
    bits = ["Email (.eml)"]
    if subj:
        bits.append(f"subject «{subj[:90]}»")
    if frm:
        bits.append(f"from «{frm[:70]}»")
    if date:
        bits.append(f"sent «{date[:35]}»")
    return "; ".join(bits)


def describe_file(p: Path) -> str:
    kind = detect_kind(p)
    ext = p.suffix.lower()
    try:
        if kind == "pdf":
            return probe_pdf(p)
        if kind in ("xlsx", "xlsm") or (kind == "zip-office" and ext in (".xlsx", ".xlsm")):
            return probe_xlsx(p)
        if kind == "docx" or (kind == "zip-office" and ext == ".docx"):
            return probe_docx(p)
        if kind == "eml":
            return probe_eml(p)
        if kind == "zip-office":
            return (
                f"Office ZIP archive saved as «{p.name}» "
                f"(extension may not match content — likely Excel/Word)"
            )
        return f"File type «{kind}», {p.stat().st_size:,} bytes"
    except Exception as e:
        return (
            f"Could not inspect content ({e.__class__.__name__}); "
            f"{p.stat().st_size:,} bytes, detected kind={kind}"
        )


def fmt_size(n: int) -> str:
    if n >= 1_000_000:
        return f"{n / 1_000_000:.1f} MB"
    if n >= 1_000:
        return f"{n / 1_000:.0f} KB"
    return f"{n} B"


def index_folder(folder: Path) -> tuple[dict, dict]:
    by_name = {}
    by_hash = {}
    for p in sorted(folder.iterdir()):
        if not p.is_file():
            continue
        digest = sha256_file(p)
        by_name[p.name] = {"path": p, "hash": digest, "size": p.stat().st_size}
        by_hash.setdefault(digest, []).append(p.name)
    return by_name, by_hash


def build_why(
    name: str,
    t: dict | None,
    d: dict | None,
    t_by_hash: dict,
    d_by_hash: dict,
    get_desc,
) -> str:
    if t and d and t["hash"] == d["hash"]:
        return "No content difference — files are byte-identical."

    t_desc = get_desc(t["path"], t["hash"]) if t else None
    d_desc = get_desc(d["path"], d["hash"]) if d else None
    lines: list[str] = []

    if t and d and t["hash"] != d["hash"]:
        t_other = [n for n in d_by_hash.get(t["hash"], []) if n != name]
        d_other = [n for n in t_by_hash.get(d["hash"], []) if n != name]

        if t_other or d_other:
            lines.append(
                "WHY DIFFERENT: Filename mismatch in the Dispute folder — both folders use the "
                "same name but the bytes are not the same document. This pattern matches a batch "
                "export/rename error (documents shifted one step), not edited or updated content."
            )
            if t_other:
                lines.append(
                    f"Tinus «{name}» ({fmt_size(t['size'])}) appears in Dispute as: "
                    f"«{'», «'.join(t_other)}»."
                )
            if d_other:
                lines.append(
                    f"Dispute «{name}» ({fmt_size(d['size'])}) is actually the Tinus file: "
                    f"«{'», «'.join(d_other)}»."
                )
            if not t_other:
                lines.append("Tinus content for this name is not found elsewhere in Dispute.")
            if not d_other:
                lines.append("Dispute content at this name is not found elsewhere in Tinus.")
        else:
            lines.append(
                "WHY DIFFERENT: Genuinely different file versions — same filename but different "
                "content, and neither byte stream appears under another name in the other folder "
                "(not a simple rename/shift)."
            )

        lines.append("")
        lines.append(f"Tinus document: {t_desc}")
        lines.append(f"Dispute document at this name: {d_desc}")

        if abs(t["size"] - d["size"]) / max(t["size"], d["size"], 1) > 0.05:
            lines.append(f"Size: Tinus {fmt_size(t['size'])} vs Dispute {fmt_size(d['size'])}.")

        tk, dk = detect_kind(t["path"]), detect_kind(d["path"])
        if tk != dk:
            lines.append(
                f"Format mismatch: Tinus file is {tk}; Dispute file is {dk} "
                f"(extensions: {t['path'].suffix} vs {d['path'].suffix})."
            )

    elif t and not d:
        t_other = d_by_hash.get(t["hash"], [])
        lines.append(f"WHY DIFFERENT: Present in Tinus only ({fmt_size(t['size'])}).")
        if t_other:
            lines.append(
                f"Same content exists in Dispute under: «{'», «'.join(t_other)}»."
            )
        else:
            lines.append("Content does not appear anywhere in the Dispute folder.")
        lines.append(f"Tinus document: {t_desc}")

    elif d and not t:
        d_other = t_by_hash.get(d["hash"], [])
        lines.append(f"WHY DIFFERENT: Present in Dispute only ({fmt_size(d['size'])}).")
        if d_other:
            lines.append(f"Same content exists in Tinus as: «{'», «'.join(d_other)}».")
        else:
            lines.append("Dispute-specific evidence (not in Tinus pack).")
        lines.append(f"Dispute document: {d_desc}")

    return "\n".join(lines)


def build_recommendation(
    name: str,
    t: dict | None,
    d: dict | None,
    status: str,
    t_by_hash: dict,
    d_by_hash: dict,
) -> tuple[str, str]:
    """Return (tier, detail) for whether the user must open/re-read the document."""
    if t and d and t["hash"] == d["hash"]:
        return (
            "No review needed — substantially the same",
            "Byte-identical in both folders. Safe to rely on either copy; no need to open both.",
        )

    if t and d and t["hash"] != d["hash"]:
        t_other = [n for n in d_by_hash.get(t["hash"], []) if n != name]
        d_other = [n for n in t_by_hash.get(d["hash"], []) if n != name]

        if t_other and d_other:
            return (
                "No review needed — substantially the same",
                "Dispute mislabeled (batch shift). Treat Tinus «"
                + name
                + "» as the correct document. The Dispute file at this name is actually «"
                + "», «".join(d_other)
                + "» — do not review it as «"
                + name
                + "». Same Tinus content exists in Dispute as «"
                + "», «".join(t_other)
                + "» if you need that path.",
            )
        if t_other and not d_other:
            return (
                "Use Tinus only — Dispute copy at this name is not the same document",
                "Tinus «"
                + name
                + "» is authoritative. Dispute has the same bytes under «"
                + "», «".join(t_other)
                + "», but the file named «"
                + name
                + "» in Dispute is different content not found in Tinus — ignore Dispute «"
                + name
                + "» unless you deliberately need that orphan file.",
            )
        if d_other and not t_other:
            return (
                "Use Tinus only — Tinus doc missing from Dispute under any name",
                "Open Tinus «"
                + name
                + "». Dispute «"
                + name
                + "» is a mislabeled copy of Tinus «"
                + "», «".join(d_other)
                + "», not «"
                + name
                + "». The real Tinus document does not appear anywhere in Dispute.",
            )
        return (
            "Review required — genuinely different versions",
            "Same filename but different content with no cross-match in the other folder. "
            "Read both Tinus and Dispute versions — one is not a rename of the other.",
        )

    if t and not d:
        t_other = d_by_hash.get(t["hash"], [])
        if t_other:
            return (
                "No review needed — substantially the same",
                "Not listed under this name in Dispute, but identical content is in Dispute as «"
                + "», «".join(t_other)
                + "». Review Tinus «"
                + name
                + "» once; skip hunting in Dispute by this filename.",
            )
        return (
            "Use Tinus only — absent from Dispute",
            "This document exists only in the Tinus pack (not found anywhere in Dispute). "
            "Use Tinus «" + name + "»; note the gap in Dispute.",
        )

    if d and not t:
        d_other = t_by_hash.get(d["hash"], [])
        if d_other:
            return (
                "No review needed — substantially the same",
                "Dispute-only filename, but content matches Tinus «"
                + "», «".join(d_other)
                + "». No need to review as new evidence.",
            )
        return (
            "Review required — Dispute-only (new evidence)",
            "Not in Tinus pack — new or dispute-specific material (e.g. IA series, SARS email). "
            "You should read «" + name + "».",
        )

    return ("Review required", "Unable to classify — inspect manually.")


def main() -> None:
    print("Indexing Tinus...")
    t_by_name, t_by_hash = index_folder(TINUS)
    print("Indexing Dispute...")
    d_by_name, d_by_hash = index_folder(DISPUTE)

    desc_cache: dict[str, str] = {}

    def get_desc(path: Path, digest: str) -> str:
        if digest not in desc_cache:
            desc_cache[digest] = describe_file(path)
        return desc_cache[digest]

    all_hashes = set(t_by_hash) | set(d_by_hash)
    print(f"Extracting content descriptions for {len(all_hashes)} unique files...")
    for digest in all_hashes:
        for n in t_by_hash.get(digest, d_by_hash.get(digest, [])):
            path = TINUS / n
            if not path.is_file():
                path = DISPUTE / n
            if path.is_file():
                get_desc(path, digest)
                break

    all_names = sorted(set(t_by_name) | set(d_by_name))
    rows = []
    for name in all_names:
        t = t_by_name.get(name)
        d = d_by_name.get(name)
        if t and d:
            same = "Yes" if t["hash"] == d["hash"] else "No"
        else:
            same = "N/A"

        if t and d:
            if t["hash"] == d["hash"]:
                status = "Identical (same name)"
            elif t["hash"] in d_by_hash:
                status = "Mislabeled in Dispute"
            elif d["hash"] in t_by_hash:
                status = "Mislabeled in Tinus"
            else:
                status = "Different content (no cross-match)"
        elif t and not d:
            status = (
                "Only in Tinus; absent from Dispute"
                if t["hash"] not in d_by_hash
                else "Only in Tinus by name"
            )
        else:
            status = (
                "Only in Dispute; unique content"
                if d["hash"] not in t_by_hash
                else "Only in Dispute by name"
            )

        why = build_why(name, t, d, t_by_hash, d_by_hash, get_desc)
        rec_tier, rec_detail = build_recommendation(
            name, t, d, status, t_by_hash, d_by_hash
        )
        t_content = get_desc(t["path"], t["hash"]) if t else ""
        d_content = get_desc(d["path"], d["hash"]) if d else ""

        rows.append(
            {
                "Document filename": name,
                "Review recommendation": rec_tier,
                "Recommendation detail": rec_detail,
                "In Tinus folder": "Yes" if t else "No",
                "In Dispute folder": "Yes" if d else "No",
                "Same content at same filename": same,
                "Content match status": status,
                "Why different / content explanation": why,
                "Tinus content summary": t_content,
                "Dispute content summary": d_content,
                "Tinus file size": fmt_size(t["size"]) if t else "",
                "Dispute file size": fmt_size(d["size"]) if d else "",
                "Tinus SHA-256 (first 16)": t["hash"][:16] if t else "",
                "Dispute SHA-256 (first 16)": d["hash"][:16] if d else "",
                "Tinus content in Dispute filed as": "; ".join(
                    n for n in d_by_hash.get(t["hash"], []) if n != name
                )
                if t
                else "",
                "Dispute content in Tinus filed as": "; ".join(
                    n for n in t_by_hash.get(d["hash"], []) if n != name
                )
                if d
                else "",
            }
        )

    headers = list(rows[0].keys())
    wb = openpyxl.Workbook()
    ws_sum = wb.active
    ws_sum.title = "Summary"
    ws_sum.append(["Sitatunga diesel rebate — folder comparison"])
    ws_sum["A1"].font = Font(bold=True, size=14)
    ws_sum.append(["Generated", datetime.now().strftime("%Y-%m-%d %H:%M")])
    ws_sum.append(["Tinus folder", str(TINUS)])
    ws_sum.append(["Dispute folder", str(DISPUTE)])
    ws_sum.append([])
    ws_sum.append(["Metric", "Count"])
    ws_sum.append(["Total unique filenames", len(all_names)])
    ws_sum.append(
        [
            "Identical at same filename",
            sum(1 for r in rows if r["Same content at same filename"] == "Yes"),
        ]
    )
    ws_sum.append(
        [
            "Mislabeled / shifted in Dispute",
            sum(1 for r in rows if r["Content match status"] == "Mislabeled in Dispute"),
        ]
    )
    ws_sum.append(
        [
            "Genuinely different versions",
            sum(1 for r in rows if "no cross-match" in r["Content match status"]),
        ]
    )
    ws_sum.append(
        [
            "Only in Dispute (unique)",
            sum(1 for r in rows if r["Content match status"] == "Only in Dispute; unique content"),
        ]
    )
    ws_sum.append([])
    ws_sum.append(["Review recommendation", "Count"])
    rec_counts: dict[str, int] = {}
    for r in rows:
        rec_counts[r["Review recommendation"]] = rec_counts.get(r["Review recommendation"], 0) + 1
    for tier in sorted(rec_counts.keys()):
        ws_sum.append([tier, rec_counts[tier]])
    ws_sum.append([])
    ws_sum.append(["Guidance", ""])
    ws_sum.append(
        [
            "No review needed",
            "Safe to treat as substantially the same; use Tinus naming where Dispute is mislabeled.",
        ]
    )
    ws_sum.append(
        [
            "Use Tinus only",
            "Do not trust Dispute at this filename; open the Tinus file (or cross-mapped Tinus name).",
        ]
    )
    ws_sum.append(
        [
            "Review required",
            "Open and compare — new evidence or genuinely different file versions.",
        ]
    )

    ws = wb.create_sheet("Document comparison")
    ws.append(headers)
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF")
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    fills = {
        "safe": PatternFill("solid", fgColor="C6EFCE"),
        "tinus": PatternFill("solid", fgColor="FFEB9C"),
        "review": PatternFill("solid", fgColor="FFC7CE"),
    }

    for r in rows:
        ws.append([r[h] for h in headers])
        row_idx = ws.max_row
        rec = r["Review recommendation"]
        if rec.startswith("No review needed"):
            fill = fills["safe"]
        elif rec.startswith("Use Tinus only"):
            fill = fills["tinus"]
        else:
            fill = fills["review"]
        for col in range(1, len(headers) + 1):
            ws.cell(row=row_idx, column=col).fill = fill
        ws.row_dimensions[row_idx].height = 72

    widths = [44, 38, 48, 12, 12, 14, 28, 58, 42, 42, 12, 12, 18, 18, 36, 36]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = min(w, 60)
    ws.column_dimensions["C"].width = 52
    ws.column_dimensions["H"].width = 72
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(rows) + 1}"

    wb.save(OUT)
    print(f"Wrote {OUT} ({len(rows)} rows)")


if __name__ == "__main__":
    main()
