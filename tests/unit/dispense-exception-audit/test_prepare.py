"""Unit tests for dispense exception prep."""
from __future__ import annotations

import json
from datetime import datetime
from pathlib import Path

import pandas as pd
import pytest

import sys

ROOT = Path(__file__).resolve().parents[3]
FIXTURES = ROOT / "tests" / "fixtures" / "dispense-exception-audit"
sys.path.insert(0, str(ROOT / "scripts" / "dispense-exception-audit"))

from enrichment import (
    enrich_transactions,
    is_mining_eligible,
    review_reason,
    should_escalate,
    suggested_abco_comment,
)
from exception_rules import RuleFlags, compute_flags_for_asset
from month_diff import compare_month_over_month
from parse_workbook import parse_details_sheet
from site_rules import get_site_profile


def test_is_mining_eligible_excludes_non_mining():
    assert is_mining_eligible({"asset_group": "Mining- Eligible", "refund_eligibility": "Eligible"})
    assert not is_mining_eligible(
        {"asset_group": "Non Mining - Non Eligible", "refund_eligibility": "Non-Eligible"}
    )


def test_consecutive_dispense_flags():
    rows = [
        {
            "transaction_id": "a1",
            "date_time": datetime(2026, 5, 1, 10, 0),
            "opening_odo_num": 100.0,
            "closing_odo_num": 100.0,
            "total_usage_num": 0.0,
            "litres": 500.0,
            "tank_size_l": 2000.0,
        },
        {
            "transaction_id": "a2",
            "date_time": datetime(2026, 5, 1, 10, 30),
            "opening_odo_num": 100.0,
            "closing_odo_num": 110.0,
            "total_usage_num": 10.0,
            "litres": 400.0,
            "tank_size_l": 2000.0,
        },
    ]
    computed = compute_flags_for_asset(rows)
    assert len(computed) == 2
    assert computed[1].flags.consec_60
    assert computed[1].flags.consec_120
    assert computed[0].flags.odo_non_positive


def test_fill_outside_exception_escalates():
    row = {
        "transaction_id": "x1",
        "exception_60": "Fill outside of one hour from start",
    }
    assert should_escalate(row, None)


def test_routine_split_escalates_after_zero_odo_starter():
    first = {
        "transaction_id": "x1",
        "exception_60": "Odo difference <= 0",
    }
    second = {
        "transaction_id": "x2",
        "exception_60": "Odo difference <= 0, Consecutive dispenses within 60 minutes",
    }
    assert should_escalate(first, None)
    assert should_escalate(second, None, prev_row=first)
    assert not should_escalate(second, None)


def test_odo_gt_50_respects_just_ok_comment():
    row = {
        "transaction_id": "x1",
        "exception_60": "Odo difference > 50 hrs",
        "abco_comment": "Just ok",
    }
    assert not should_escalate(row, None)


def test_economy_variance_alone_does_not_escalate_belfast_profile():
    row = {
        "transaction_id": "x1",
        "refund_eligibility": "Eligible",
        "total_usage_num": 10.0,
        "pct_variance": 1.5,
    }
    assert not should_escalate(row, None, profile=get_site_profile("belfast"))


def test_suggested_abco_comment_fill_outside():
    row = {"exception_60": "Fill outside of one hour from start"}
    assert suggested_abco_comment(row, None) == "Check 120 min"


def test_review_reason_fill_outside():
    row = {"exception_60": "Fill outside of one hour from start"}
    assert review_reason(row, None) == "Fill outside one hour"


def test_repeated_asset_header_preserves_exception_columns():
    rows = [
        ["Date & Time", "Transaction ID", None, "Asset Number", None, None, None, None, None, None, "Litres", None, None, "Total Usage Km/Hr", "Exception Reason (120 min)", "Exception Reason (60 min)"],
        ["BD110", None, None, None, None, None, None, None, None, None, None, None, None, None, None, None],
        ["Date & Time", "Transaction ID", None, "Asset Number", None, None, None, None, None, None, "Litres", None, None, "Total Usage Km/Hr", None, None],
        [datetime(2026, 5, 9, 9, 4), "59-20260509-9-1-165-663056", None, "CBK304", None, None, None, "hr", None, None, 307.3, None, None, "1682.0 hr", "Odo difference > 50 hrs", "Odo difference > 50 hrs"],
    ]
    df = pd.DataFrame(rows)
    parsed = parse_details_sheet(df)
    assert len(parsed) == 1
    assert parsed[0]["exception_60"] == "Odo difference > 50 hrs"
    assert parsed[0]["exception_120"] == "Odo difference > 50 hrs"


def test_month_over_month_diff(tmp_path):
    import openpyxl

    prior_path = tmp_path / "prior.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Transactions deemed ineligible"
    ws.append(["Transaction ID", "Litres"])
    ws.append(["a", 5])
    wb.save(prior_path)

    current = [
        {"transaction_id": "a", "asset_number": "BD110", "litres": 10},
        {"transaction_id": "b", "asset_number": "BD119", "litres": 20},
    ]
    diff = compare_month_over_month(current, str(prior_path))
    assert diff is not None
    assert diff["prior_review_count"] == 1
    assert diff["new_in_review_count"] == 1
    assert diff["repeat_in_review_count"] == 1


@pytest.mark.integration
def test_belfast_golden_review_queue():
    expected_path = FIXTURES / "belfast-may-2026-expected.json"
    expected = json.loads(expected_path.read_text(encoding="utf-8"))
    source_candidates = [
        Path("/Users/gemau/Downloads/Exxaro Belfast Mine - Transaction Exceptions - In Context - May 2026 (1).xlsx"),
        FIXTURES / "belfast-may-2026-source.xlsx",
    ]
    source = next((path for path in source_candidates if path.exists()), None)
    if source is None:
        pytest.skip("Belfast source workbook not available")

    from parse_workbook import load_workbook

    parsed = load_workbook(str(source))
    enriched = enrich_transactions(parsed["transactions"], {}, [], profile=get_site_profile("belfast"))
    review_ids = sorted(row["transaction_id"] for row in enriched["review_queue"])

    assert len(review_ids) == expected["review_queue_count"]
    assert set(review_ids) == set(expected["review_transaction_ids"])

    causes = {
        row["exception_reason"]: row
        for row in __import__("enrichment").build_possible_cause_summary(enriched["review_queue"])
    }
    for reason, totals in expected["possible_cause_totals"].items():
        assert reason in causes
        assert causes[reason]["transaction_count"] == totals["transactions"]
        assert round(causes[reason]["litres"], 2) == totals["litres"]
