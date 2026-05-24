"""Unit tests for Details as Assets audit column writer."""

import os
import sys
import tempfile
import unittest

sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', '..', 'scripts', 'dispense-exception-audit'))

from detailsSheetAudit import (  # noqa: E402
    apply_audit_columns_to_details_sheet,
    build_txn_audit_lookup,
    format_findings_for_cell,
)
from parseWorkbook import load_workbook  # noqa: E402
from runAudit import write_audit_excel  # noqa: E402
from auditDecisions import run_all_audits  # noqa: E402

SAMPLE_PATH = os.environ.get(
    'DISPENSE_EXCEPTION_SAMPLE',
    '/Users/gemau/Downloads/Exxaro Belfast Mine - Transaction Exceptions - In Context - Apr 2026.xlsx',
)


class DetailsSheetAuditHelpersTest(unittest.TestCase):
    def test_format_findings_for_cell_skips_info(self):
        text = format_findings_for_cell([
            {'severity': 'warning', 'check': 'review_queue', 'expected_value': 'Should be in queue'},
            {'severity': 'info', 'check': 'economy_outlier', 'expected_value': 'High variance'},
        ])
        self.assertIn('review_queue', text)
        self.assertNotIn('economy_outlier', text)

    def test_build_txn_audit_lookup_includes_comment(self):
        lookup = build_txn_audit_lookup(
            [{'transaction_id': 'TX1', 'max_severity': 'warning', 'findings': []}],
            {'TX1': 'Looks acceptable'},
        )
        self.assertEqual(lookup['TX1']['comment'], 'Looks acceptable')
        self.assertEqual(lookup['TX1']['severity'], 'warning')


@unittest.skipUnless(os.path.isfile(SAMPLE_PATH), 'sample workbook not available')
class DetailsSheetAuditIntegrationTest(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.data = load_workbook(SAMPLE_PATH)
        cls.audit_result = run_all_audits(cls.data)

    def test_write_audit_excel_adds_details_columns(self):
        with tempfile.TemporaryDirectory() as tmp:
            output_path = os.path.join(tmp, 'audit.xlsx')
            write_audit_excel(SAMPLE_PATH, output_path, self.audit_result)

            from openpyxl import load_workbook as load_wb

            wb = load_wb(output_path)
            ws = wb['Details as Assets']

            header_row = None
            for row_idx in range(1, min(ws.max_row, 200)):
                values = [ws.cell(row=row_idx, column=c).value for c in range(1, ws.max_column + 1)]
                joined = ' '.join(str(v or '').lower() for v in values)
                if 'transaction id' in joined and 'audit severity' in joined:
                    header_row = row_idx
                    break

            self.assertIsNotNone(header_row, 'Expected Audit Severity header on Details as Assets')

            flagged = self.audit_result['review_transactions'][0]
            tid = flagged['transaction_id']
            found_value = False
            for row_idx in range(1, ws.max_row + 1):
                txn_cell = ws.cell(row=row_idx, column=2).value
                if str(txn_cell or '').strip() != str(tid):
                    continue
                severity_cell = None
                for col_idx in range(1, ws.max_column + 1):
                    if ws.cell(row=header_row, column=col_idx).value == 'Audit Severity':
                        severity_cell = ws.cell(row=row_idx, column=col_idx).value
                        break
                if severity_cell:
                    found_value = True
                    break
            self.assertTrue(found_value, 'Expected audit severity on a flagged transaction row')

            apply_audit_columns_to_details_sheet(
                output_path,
                self.audit_result['review_transactions'],
                {tid: 'Analyst sign-off'},
            )
            wb = load_wb(output_path)
            ws = wb['Details as Assets']
            comment_value = None
            for row_idx in range(1, ws.max_row + 1):
                if str(ws.cell(row=row_idx, column=2).value or '').strip() != str(tid):
                    continue
                for col_idx in range(1, ws.max_column + 1):
                    if ws.cell(row=header_row, column=col_idx).value == 'Auditor Comment':
                        comment_value = ws.cell(row=row_idx, column=col_idx).value
                        break
            self.assertEqual(comment_value, 'Analyst sign-off')


if __name__ == '__main__':
    unittest.main()
