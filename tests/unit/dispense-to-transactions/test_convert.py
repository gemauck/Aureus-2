"""Contract tests for dispense → transactions converter."""
from __future__ import annotations

import sys
from pathlib import Path

import openpyxl
import pytest

REPO_ROOT = Path(__file__).resolve().parents[3]
SCRIPT_DIR = REPO_ROOT / "scripts" / "dispense-to-transactions"
if str(SCRIPT_DIR) not in sys.path:
    sys.path.insert(0, str(SCRIPT_DIR))

from convert import (
    DEFAULT_GILBARCO_TEMPLATE,
    convert_row_to_gilbarco,
    convert_vehicle_row,
    format_transaction_datetime,
    load_pump_config,
    lookup_fuel_pump_route,
    merge_fuel_pump_routes,
    resolve_bowser_fleet_id,
)
from fleet_lookup import TRANSACTION_HEADERS, build_fuel_pump_routes, load_reference_indexes
from parse_dispense import (
    DISPENSE_HEADERS,
    GILBARCO_SECTION_TITLE,
    ORIGINAL_SECTION_TITLE,
    OUTPUT_SHEET_NAME,
    build_output_filename,
    collapse_redundant_model,
    dispense_period_range,
    is_footer_row,
    model_has_redundant_duplicate,
    normalize_asset_group,
    parse_consumption,
    resolve_make_model,
)


def test_footer_row_detection():
    assert is_footer_row({"Pump After": "Total Litres:", "Date & Time": None})
    assert not is_footer_row({"Date & Time": "2026-05-29 11:49:09", "Asset Number": "CAM1513"})


def test_format_transaction_datetime():
    dt, d, t = format_transaction_datetime("2026-05-29 13:51:43")
    assert dt == "29/05/2026 13:51:43"
    assert d == "29/05/2026"
    assert t == "13:51:43"


def test_parse_consumption_hours():
    meter, ctype, reading = parse_consumption("2083 hr", "0.19 L/hr")
    assert meter == "Hours"
    assert ctype == "L/HR"
    assert reading == 2083.0


def test_normalize_asset_group():
    assert normalize_asset_group("Non-Eligible") == "Not Eligible"
    assert normalize_asset_group("Eligible") == "Eligible"


def test_build_output_filename():
    assert build_output_filename(("20260529", "20260602")) == "Fuel Dispense Report 20260529 - 20260602.xlsx"
    assert build_output_filename(None) == "Fuel Dispense Report.xlsx"


def test_dispense_period_range():
    rows = [
        {"Date & Time": "2026-05-29 11:49:09"},
        {"Date & Time": "2026-06-02 14:00:00"},
    ]
    assert dispense_period_range(rows) == ("20260529", "20260602")


def test_bundled_gilbarco_template_exists():
    assert DEFAULT_GILBARCO_TEMPLATE.exists(), "gilbarco-template.xlsx must be committed in scripts/dispense-to-transactions/"


def test_convert_vehicle_row_uses_fleet_template():
    config = load_pump_config()
    fleet = {
        "CAM1513": {
            "Make": "BELL",
            "Model": "DUMP TRUCK B40D",
            "Location": "ANDRU MINING",
            "Device ID": "A MINING - CONTR",
            "Pump": "A MINING  - CONTR P1",
            "Pump Controller": "490",
            "VehicleCategory Description": "ARTICULATED DUMP TRUCK",
            "Group1": None,
            "Group2": "Not Eligible",
            "Group3": "ANDRU MINING",
            "Group4": "306080",
            "Group5": "4003338577",
            "Product Name": "DIESEL",
            "Consumption Meter": "Hours",
            "Consumption Type": "L/HR",
        }
    }
    row = {
        "Date & Time": "2026-05-29 13:51:43",
        "API ID": 1736741,
        "Fuel Pump": "OTK105M (FuelTrack)",
        "Asset Description": "BELL-B40E-ADT",
        "Asset Number": "CAM1513",
        "Asset Registration": "CAM1513",
        "Group1": "Andru Mining HDV",
        "Asset Group": "Non-Eligible",
        "Asset Owner": "ANDRU MINING",
        "Internal Order Number": 4003338577,
        "Operation": "Rehap - Haul Materials",
        "Litres": 404.739,
        "Odometer": "2083 hr",
        "Economy": "0.19 L/hr",
    }
    warnings: list[str] = []
    out = convert_vehicle_row(row, fleet, {}, config, warnings)
    assert out is not None
    assert out["Fleet ID"] == "CAM1513"
    assert out["Make"] == "BELL"
    assert out["Pump"] == "A MINING  - CONTR P1"
    assert out["Voucher Number"] == "1736741"
    assert out["Liters"] == pytest.approx(404.739)


def test_unallocated_row_always_converts():
    config = load_pump_config()
    row = {
        "Date & Time": "2026-05-29 12:13:29",
        "API ID": 1736671,
        "Fuel Pump": "OTK105M (FuelTrack)",
        "Litres": 12.044,
        "Override": "Code: 192867; Test 2nd fill; User: Greg Keague",
    }
    warnings: list[str] = []
    out = convert_row_to_gilbarco(row, {}, {}, {}, config, warnings)
    assert out["Liters"] == pytest.approx(12.044)
    assert out["Voucher Number"] == "1736671"


def test_side_by_side_workbook_layout(tmp_path):
    from convert import write_side_by_side_workbook

    dispense = [
        {
            "Date & Time": "2026-05-29 13:51:43",
            "API ID": 1,
            "Fuel Pump": "OTK105M",
            "Asset Number": "CAM1513",
            "Litres": 100.0,
        },
        {
            "Date & Time": "2026-05-29 12:13:29",
            "API ID": 2,
            "Fuel Pump": "OTK105M",
            "Litres": 12.0,
        },
    ]
    gilbarco = [
        {"Fleet ID": "CAM1513", "Liters": 100.0, "Voucher Number": "1"},
        {"Fleet ID": "OTK105M", "Liters": 12.0, "Voucher Number": "2"},
    ]
    out = tmp_path / "out.xlsx"
    write_side_by_side_workbook(dispense, gilbarco, out)

    wb = openpyxl.load_workbook(out, read_only=True, data_only=True)
    assert wb.sheetnames == [OUTPUT_SHEET_NAME]
    ws = wb[OUTPUT_SHEET_NAME]
    assert ws.cell(1, 1).value == GILBARCO_SECTION_TITLE
    dispense_start = len(TRANSACTION_HEADERS) + 2
    assert ws.cell(1, dispense_start).value == ORIGINAL_SECTION_TITLE
    assert ws.cell(2, 1).value == TRANSACTION_HEADERS[0]
    assert ws.cell(2, dispense_start).value == DISPENSE_HEADERS[0]
    assert ws.cell(3, 1).value == "CAM1513"
    assert ws.cell(3, dispense_start).value == "2026-05-29 13:51:43"
    api_col = dispense_start + DISPENSE_HEADERS.index("API ID")
    assert ws.cell(4, api_col).value == 2
    wb.close()


def test_otk105m_bulk_alias_uses_bowser_template():
    config = load_pump_config()
    bowser = {
        "OTK105M": {
            "Make": "BULK DIESEL TANK",
            "Model": "BULK TANK BULK TANK",
            "Location": "Service Bay",
            "Device ID": "Mafube HDV Bay",
            "Pump": "SB P1",
            "VehicleCategory Description": "BOWSER",
        }
    }
    row = {
        "Date & Time": "2026-04-01 08:55:04",
        "API ID": 99,
        "Fuel Pump": "Service Bay - 1 Tank",
        "Asset Number": "OTK105M - BULK",
        "Asset Registration": "OTK105M-BULK",
        "Asset Description": "HINO-MOBILE-DIESEL BOWSER",
        "Group1": "Andru Mining HDV",
        "Asset Owner": "ANDRU MINING",
        "Litres": 5212.0,
    }
    warnings: list[str] = []
    out = convert_row_to_gilbarco(row, {}, {}, bowser, config, warnings)
    assert out["Fleet ID"] == "OTK105M"
    assert out["Pump"] == "SB P1"
    assert out["Location"] == "Service Bay"


def test_dbotk010_bowser_asset_uses_nooitgedacht_pump():
    config = load_pump_config()
    bowser = {
        "DBOTK010": {
            "Make": "BULK DIESEL TANK",
            "Model": "BULK TANK BULK TANK",
            "Location": "Nooitgedact",
            "Device ID": "Mafube - Nooitgedact",
            "Pump": "Mafube - Nooitgedact - HDV P2",
            "VehicleCategory Description": "BOWSER",
        }
    }
    row = {
        "Date & Time": "2026-04-01 10:50:25",
        "API ID": 100,
        "Fuel Pump": "Nooitgedact P2",
        "Asset Number": "DBOTK010",
        "Asset Registration": "DBOTK010",
        "Asset Owner": "MAFUBE",
        "Litres": 8358.3,
    }
    warnings: list[str] = []
    out = convert_row_to_gilbarco(row, {}, {"MAFUBE": {"pump": "OTK010 DB - DSL P1"}}, bowser, config, warnings)
    assert out["Fleet ID"] == "DBOTK010"
    assert out["Pump"] == "Mafube - Nooitgedact - HDV P2"


def test_odt001m_alias_resolves_vehicle_template():
    config = load_pump_config()
    fleet = {
        "370ODT001": {
            "Make": "HITACHI",
            "Model": "REAR DUMP TRUCK EH3500",
            "Location": "Nooitgedact",
            "Device ID": "Mafube - Nooitgedact",
            "Pump": "Mafube - Nooitgedact - HDV P2",
            "VehicleCategory Description": "REAR DUMP TRUCK",
            "Group3": "MAFUBE",
            "Product Name": "DIESEL",
            "Consumption Meter": "Hours",
            "Consumption Type": "L/HR",
        }
    }
    row = {
        "Date & Time": "2026-04-01 12:53:59",
        "API ID": 101,
        "Fuel Pump": "OTK010 DB",
        "Asset Number": "ODT001M",
        "Asset Registration": "ODT001M",
        "Group1": "MAFUBE - HDV",
        "Asset Owner": "MAFUBE",
        "Litres": 2523.7,
    }
    routes = merge_fuel_pump_routes({}, config)
    warnings: list[str] = []
    out = convert_row_to_gilbarco(row, fleet, {}, {}, config, warnings, routes)
    assert out["Fleet ID"] == "370ODT001"
    assert out["Pump"] == "Mafube - Nooitgedact - HDV P2"


def test_fuel_pump_route_before_owner_route():
    config = load_pump_config()
    row = {
        "Date & Time": "2026-04-01 09:15:56",
        "API ID": 102,
        "Fuel Pump": "OTK010 DB",
        "Asset Number": "OLB001M",
        "Asset Registration": "OLB001M",
        "Group1": "MAFUBE - HDV",
        "Asset Owner": "MAFUBE",
        "Litres": 933.2,
    }
    fleet = {
        "307OLB001": {
            "Make": "CATERPILLAR",
            "Model": "TLB 426",
            "Location": "OTK010 DB",
            "Device ID": "OTK010 DB",
            "Pump": "OTK010 DB - DSL P1",
            "Product Name": "DIESEL",
            "Consumption Meter": "Hours",
            "Consumption Type": "L/HR",
        }
    }
    owner = {
        "MAFUBE": {
            "location": "Nooitgedact",
            "device_id": "Mafube - Nooitgedact",
            "pump": "Mafube - Nooitgedact - HDV P2",
        }
    }
    routes = merge_fuel_pump_routes({}, config)
    warnings: list[str] = []
    out = convert_vehicle_row(row, fleet, owner, config, warnings, routes)
    assert out["Fleet ID"] == "307OLB001"
    assert out["Pump"] == "OTK010 DB - DSL P1"


def test_lookup_fuel_pump_route_from_bundled_reference():
    routes = build_fuel_pump_routes(DEFAULT_GILBARCO_TEMPLATE)
    config = load_pump_config()
    merged = merge_fuel_pump_routes(routes, config)
    route = lookup_fuel_pump_route("Nooitgedact P2", config, merged)
    assert route.get("pump") == "Mafube - Nooitgedact - HDV P2"


def test_resolve_bowser_fleet_id_for_bulk_suffix():
    config = load_pump_config()
    bowser = {"OTK105M": {"Pump": "SB P1"}}
    row = {"Asset Number": "OTK105M - BULK", "Asset Description": "HINO-MOBILE-DIESEL BOWSER"}
    assert resolve_bowser_fleet_id(row, config, bowser) == "OTK105M"


def test_model_has_redundant_duplicate():
    assert model_has_redundant_duplicate("DOZER DOZER")
    assert model_has_redundant_duplicate("HILUX  HILUX")
    assert not model_has_redundant_duplicate("DOZER D11T")
    assert not model_has_redundant_duplicate("EX3600 6 EXCAVATOR")


def test_collapse_redundant_model():
    assert collapse_redundant_model("DOZER DOZER") == "DOZER"
    assert collapse_redundant_model("HILUX  HILUX") == "HILUX"


def test_resolve_make_model_prefers_sparrow_description():
    template = {"Make": "CATERPILLAR", "Model": "DOZER DOZER"}
    make, model = resolve_make_model("CATERPILLAR-D11R-DOZER", template)
    assert make == "CATERPILLAR"
    assert model == "D11R DOZER"


def test_resolve_make_model_keeps_good_template():
    template = {"Make": "CATERPILLAR", "Model": "DOZER D11T"}
    make, model = resolve_make_model("CATERPILLAR-D11R-DOZER", template)
    assert make == "CATERPILLAR"
    assert model == "DOZER D11T"


def test_winshuttle_report_date_format():
    from winshuttle import format_winshuttle_report_date

    assert format_winshuttle_report_date("20260529") == "29/5/2026"
    assert format_winshuttle_report_date("20260602") == "2/6/2026"


def test_winshuttle_workbook_layout(tmp_path):
    from winshuttle import WINSHUTTLE_HEADERS, WINSHUTTLE_SAP_FIELDS, write_winshuttle_workbook

    dispense = [
        {
            "Date & Time": "2026-05-29 13:51:43",
            "Internal Order Number": 4003338574,
            "Asset Number": "UR13",
            "Litres": 369.0,
        },
        {
            "Date & Time": "2026-05-29 12:13:29",
            "Asset Number": "DBOTK010",
            "Litres": 10354.5,
        },
    ]
    gilbarco = [
        {"Fleet ID": "UR13", "Liters": 369.0, "Group5": 4003338574},
        {"Fleet ID": "DBOTK010", "Liters": 10354.5, "Group5": 4009348222},
    ]
    ws_cfg = {
        "material": "am0193746",
        "storage_location": "mf07",
        "goods_recipient": "mpho",
        "product": 261,
        "name": "tm01",
    }
    out = tmp_path / "winshuttle.xlsx"
    write_winshuttle_workbook(
        dispense, gilbarco, out, report_date="29/5/2026", ws_cfg=ws_cfg
    )

    wb = openpyxl.load_workbook(out, read_only=True, data_only=True)
    ws = wb.active
    assert ws.cell(1, 1).value == WINSHUTTLE_HEADERS[0]
    assert ws.cell(2, 1).value == WINSHUTTLE_SAP_FIELDS[0]
    assert ws.cell(3, 2).value == "29/5/2026"
    assert ws.cell(4, 1).value == "UR13"
    assert ws.cell(4, 2).value == "am0193746"
    assert ws.cell(4, 3).value == pytest.approx(369.0)
    assert ws.cell(4, 4).value == "mf07"
    assert ws.cell(4, 5).value == "mpho"
    assert ws.cell(4, 6).value == 4003338574
    assert ws.cell(4, 7).value == 261
    assert ws.cell(4, 8).value == "tm01"
    assert ws.cell(5, 1).value == "DBOTK010"
    assert ws.cell(5, 6).value == 4009348222
    assert ws.cell(1, 1).fill.fgColor.rgb in ("00729FB3", "729FB3")
    assert ws.cell(2, 7).font.bold is False
    assert ws.cell(3, 2).fill.fgColor.rgb in ("0000B050", "00B050")
    assert ws.cell(3, 2).font.bold is True
    assert ws.cell(4, 1).fill.patternType is None
    wb.close()


def test_convert_vehicle_row_fixes_duplicate_template_model():
    config = load_pump_config()
    fleet = {
        "DZ100030": {
            "Make": "CATERPILLAR",
            "Model": "DOZER DOZER",
            "Location": "OTK010 DB",
            "Device ID": "OTK010 DB",
            "Pump": "OTK010 DB - DSL P1",
            "Product Name": "DIESEL",
            "Consumption Meter": "Hours",
            "Consumption Type": "L/HR",
        }
    }
    row = {
        "Date & Time": "2026-05-29 05:04:31",
        "Asset Number": "DZ100030",
        "Asset Description": "CATERPILLAR-D11R-DOZER",
        "Fuel Pump": "OTK010 DB",
        "Litres": 100.0,
    }
    routes = merge_fuel_pump_routes({}, config)
    warnings: list[str] = []
    out = convert_vehicle_row(row, fleet, {}, config, warnings, routes)
    assert out["Model"] == "D11R DOZER"
    assert out["Make"] == "CATERPILLAR"
