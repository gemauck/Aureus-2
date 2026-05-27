"""Tests for re-audit workbook normalization."""
import os
import sys
import tempfile
from pathlib import Path

import openpyxl
import pytest

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "..", "..", "scripts", "fuel-refund-report-audit"))

from parse_workbook import parse_workbook  # noqa: E402
from workbook_normalize import prepare_output_workbook  # noqa: E402


def test_prepare_strips_audit_columns_before_parse():
  with tempfile.TemporaryDirectory() as tmp:
    src = Path(tmp) / "source.xlsx"
    out = Path(tmp) / "out.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Combined Fuel Transactions"
    ws.cell(1, 1, "Banner")
    headers = [
      "Audit Result",
      "Audit Severity",
      "Findings Count",
      "Audit Comments",
      "Checks Failed",
      "Transaction Type",
      "Date & Time",
      "Transaction ID",
      "Asset Group",
      "Fuel Dispensed or Received (L)",
    ]
    for col, h in enumerate(headers, 1):
      ws.cell(2, col, h)
    ws.cell(3, 6, "DISPENSE")
    ws.cell(3, 7, "2026-04-01 10:00:00")
    ws.cell(3, 8, "TX-1")
    ws.cell(3, 9, "Mining - Eligible")
    ws.cell(3, 10, -50)
    wb.create_sheet("Audit Findings")
    wb.save(src)
    wb.close()

    warnings = prepare_output_workbook(src, out)
    assert any("audit columns" in w.lower() for w in warnings)
    parsed = parse_workbook(out)
    assert len(parsed.combined_rows) == 1
    assert parsed.combined_rows[0]["Transaction ID"] == "TX-1"
    assert "Audit Result" not in parsed.combined_headers
